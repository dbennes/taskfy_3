import csv
import io
import os
import re
import hashlib
import math
from datetime import datetime, date
from typing import Dict, List, Optional, Tuple
from xml.etree import ElementTree as ET

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils.text import slugify

from .forms import ScheduleImportForm
from .models import JobCard, ScheduleActivity, ScheduleWBS

try:
    from .models import ScheduleLink  # opcional
except Exception:
    ScheduleLink = None


def _reset_schedule() -> None:
    """
    Zera o cronograma (links -> atividades -> WBS), em ordem segura.
    Deve ser chamado DENTRO de uma transação (atomic).
    """
    if ScheduleLink:
        ScheduleLink.objects.all().delete()
    ScheduleActivity.objects.all().delete()
    ScheduleWBS.objects.all().delete()


# ==================== Config ====================

DATE_FORMATS = (
    "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y",
    "%d-%b-%y", "%d-%b-%Y"  # ex: 31-May-24 A
)
_WBS_CODE_MAX = 760  # (model usa 768; deixo folga p/ sufixo hash)

# ==================== Utils ====================

def _clean(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()

def _expand_tabs(s: str) -> str:
    # normaliza TAB e NBSP para espaço
    return (s or "").replace("\t", "    ").replace("\xa0", " ")

def _parse_date(val) -> Optional[datetime.date]:
    """Aceita date/datetime reais (openpyxl), ISO strings e formatos comuns."""
    if val is None:
        return None
    if isinstance(val, date):
        return val if not isinstance(val, datetime) else val.date()
    txt = str(val).strip()
    if not txt:
        return None
    txt = re.sub(r"\s+", " ", txt)
    # corta timezone e sufixos " A"/" P", e normaliza "T"
    txt = txt.replace("T", " ")
    txt = txt.split(" A")[0].split(" P")[0].strip()
    # tenta YYYY-MM-DD HH:MM:SS e YYYY-MM-DD
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(txt, fmt).date()
        except ValueError:
            pass
    # tenta os demais formatos
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d-%b-%y", "%d-%b-%Y"):
        try:
            return datetime.strptime(txt, fmt).date()
        except ValueError:
            continue
    return None

def _to_float(x) -> Optional[float]:
    if x is None:
        return None
    s = str(x).strip().replace(",", "")
    if not s:
        return None
    m = re.search(r"[-+]?\d*\.?\d+", s)
    return float(m.group(0)) if m else None

def _to_percent(x) -> float:
    v = _to_float(x)
    return float(v) if v is not None else 0.0

def _is_activity_id(text: str) -> bool:
    """Regra segura: ID não tem espaços e tem hífen/underscore + dígito, ou letras+2+ dígitos."""
    if not text:
        return False
    t = text.strip()
    if re.search(r"\s", t):
        return False
    if re.search(r"[-_]", t) and re.search(r"\d", t):
        return True
    if re.match(r"^[A-Za-z]+[0-9]{2,}$", t):
        return True
    return False

def _wbs_code_from_path(path_names: List[str]) -> str:
    slugs = [slugify(_clean(p)) or "node" for p in path_names if _clean(p)]
    code = "/".join(slugs) or "root"
    if len(code) > _WBS_CODE_MAX:
        h = hashlib.sha1(code.encode("utf-8")).hexdigest()[:10]
        code = f"{code[:_WBS_CODE_MAX-11]}-{h}"
    return code

def _ensure_wbs_by_path(path_names: List[str]) -> ScheduleWBS:
    parent = None
    built: List[str] = []
    obj = None
    for name in path_names:
        name = _clean(name)
        if not name:
            continue
        built.append(name)
        code = _wbs_code_from_path(built)
        obj, created = ScheduleWBS.objects.get_or_create(
            code=code, defaults={"name": name, "parent": parent}
        )
        if not created and obj.name != name:
            obj.name = name
            obj.save(update_fields=["name"])
        parent = obj
    return obj

def _link_jobcard_number(activity_id: str) -> str:
    if not activity_id:
        return ""
    jc = JobCard.objects.filter(activity_id=activity_id).only("job_card_number").first()
    return jc.job_card_number if jc else ""

def _update_jobcard_dates(activity_id: str, start, finish, do_update: bool) -> int:
    """
    Atualiza TODAS as JobCards com o activity_id informado para as datas vindas do P6.
    Match robusto: ignora diferenças de caixa e espaços à esquerda/direita.
    Retorna a quantidade de JobCards efetivamente atualizadas.
    """
    if not do_update:
        return 0
    aid = (activity_id or "").strip()
    if not aid:
        return 0

    qs = JobCard.objects.filter(
        Q(activity_id__iexact=aid) |
        Q(activity_id__iregex=rf'^\s*{re.escape(aid)}\s*$')
    )

    updated = 0
    for jc in qs:
        fields = []
        if start and jc.start != start:
            jc.start = start
            fields.append("start")
        if finish and jc.finish != finish:
            jc.finish = finish
            fields.append("finish")
        if fields:
            jc.save(update_fields=fields)
            updated += 1
    return updated

def _norm(h: str) -> str:
    return _clean(h).lower()


# ==================== Parser por LEVEL (autoridade) ====================

def _normalize_level(raw_level: str, base: int) -> int:
    """
    Converte o Level do arquivo (pode começar em 0,1,2...) em um nível interno 1-based.
    Ex.: base=0 -> 0,1,2 vira 1,2,3; base=1 -> 1,2,3 vira 1,2,3.
    """
    try:
        L = int(float(str(raw_level).strip()))
    except Exception:
        L = 0
    return max(1, L - base + 1)

def parse_with_levels(rows: List[Dict[str, str]], update_jobcards: bool) -> Tuple[int, int, int]:
    """
    Espera colunas:
      Level | Activity ID | Activity Name | HH | Activity % Complete | Start | Finish | Pontos | (Original Duration opcional)
    - NÃO cria WBS; apenas lê o Level e armazena em ScheduleActivity.level.
    - Mantém a ordem do arquivo em ScheduleActivity.sort_index.
    - **HH é a métrica oficial**. Se 'HH' não existir, tentamos usar 'Original Duration' como fallback para HH.
    Retorna (created, updated, jobcards_updated).
    """
    if not rows:
        return (0, 0, 0)

    headers = list(rows[0].keys())
    H = { _norm(h): h for h in headers }

    H_LVL = H.get("level") or H.get("nivel")
    H_AID = H.get("activity id") or H.get("id")
    H_ANM = H.get("activity name") or H.get("name")
    H_OD  = H.get("original duration") or H.get("orig duration")
    H_PC  = H.get("activity % complete") or H.get("% complete") or H.get("percent complete")
    H_ST  = H.get("start")
    H_FN  = H.get("finish")
    H_PTS = H.get("pontos") or H.get("points")
    H_HH  = H.get("hh") or H.get("manhours") or H.get("man hours")

    if not H_LVL or not H_AID:
        raise RuntimeError("Template inválido: colunas 'Level' e 'Activity ID' são obrigatórias.")

    created, updated, jc_updates = 0, 0, 0
    order_counter = 0

    with transaction.atomic():
        for row in rows:
            order_counter += 1

            raw_level = str(row.get(H_LVL, "")).strip()
            try:
                level = int(float(raw_level))
            except Exception:
                level = 0

            aid = _clean(_expand_tabs(str(row.get(H_AID, "") or "")))
            if not aid:
                continue  # linha inválida

            name  = _clean(str(row.get(H_ANM, "") or "")) if H_ANM else ""
            start = _parse_date(row.get(H_ST)) if H_ST else None
            fin   = _parse_date(row.get(H_FN)) if H_FN else None

            # ========= HH (oficial) =========
            hh_val = _to_float(row.get(H_HH)) if H_HH else None
            if hh_val is None and H_OD:
                # fallback: se vier só "Original Duration", usamos como HH
                hh_val = _to_float(row.get(H_OD))

            pc    = _to_percent(row.get(H_PC)) if H_PC else 0.0
            pts   = _to_float(row.get(H_PTS)) if H_PTS else None

            duration_days = (fin - start).days if (start and fin) else None
            jobcard_number = _link_jobcard_number(aid)

            defaults = dict(
                name=name or aid,
                start=start,
                finish=fin,
                duration_days=duration_days,
                original_duration_days=None,   # não usamos mais OD para lógica
                percent_complete=pc,
                points=pts,
                hh=hh_val,                     # <<=== HH oficial
                wbs=None,                      # ignoramos WBS
                level=level,                   # guarda level do arquivo
                sort_index=order_counter,
                jobcard_number=jobcard_number,
            )
            _, is_created = ScheduleActivity.objects.update_or_create(activity_id=aid, defaults=defaults)
            created += 1 if is_created else 0
            updated += 0 if is_created else 1

            jc_updates += _update_jobcard_dates(aid, start, fin, update_jobcards)

    return created, updated, jc_updates


# ==================== Parsers de arquivo ====================

def parse_csv(file, update_jobcards: bool) -> Tuple[int, int, int]:
    raw = file.read().decode("utf-8-sig", errors="ignore")
    if not raw.strip():
        return (0, 0, 0)
    # Excel pt-BR costuma usar ';'
    try:
        sample = "\n".join(raw.splitlines()[:5])
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
    except Exception:
        class _D: delimiter = ';'
        dialect = _D()
    reader = csv.DictReader(io.StringIO(raw), dialect=dialect)
    rows = list(reader)
    return parse_with_levels(rows, update_jobcards)

def parse_xlsx(file, update_jobcards: bool) -> Tuple[int, int, int]:
    try:
        import openpyxl
    except Exception as e:
        raise RuntimeError("Para importar .xlsx/.xlsm instale openpyxl: pip install openpyxl") from e

    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: List[Dict[str, str]] = []
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row = {}
        for i, cell in enumerate(r[:len(headers)]):
            row[headers[i]] = cell.value if cell.value is not None else ""
        rows.append(row)
    return parse_with_levels(rows, update_jobcards)

def _find_tag(el, *names):
    for n in names:
        node = el.find(f".//{n}")
        if node is not None and node.text:
            return node.text
        for child in el.iter():
            if child.tag.split('}')[-1].lower() == n.lower() and child.text:
                return child.text
    return ""

def parse_xml(file, update_jobcards: bool) -> Tuple[int, int, int]:
    # Importa atividades "flat" (XML geralmente não traz Level)
    tree = ET.parse(file)
    root = tree.getroot()
    created, updated, jc_updates = 0, 0, 0
    acts = root.findall(".//Activity") or [el for el in root.iter() if el.tag.split('}')[-1].lower() == "activity"]
    with transaction.atomic():
        for a in acts:
            aid   = _clean(_find_tag(a, "ActivityID", "Id"))
            if not aid:
                continue
            name  = _clean(_find_tag(a, "ActivityName", "Name"))
            start = _parse_date(_find_tag(a, "StartDate", "EarlyStartDate", "PlannedStartDate"))
            fin   = _parse_date(_find_tag(a, "FinishDate", "EarlyFinishDate", "PlannedFinishDate"))
            pc    = _to_percent(_find_tag(a, "PercentComplete", "CompletePercent") or "0")
            duration_days = (fin - start).days if (start and fin) else None
            jobcard_number = _link_jobcard_number(aid)
            defaults = dict(
                name=name, start=start, finish=fin, duration_days=duration_days,
                percent_complete=pc, wbs=None, jobcard_number=jobcard_number,
                sort_index=0,
                hh=None,                        # XML normalmente não traz HH
                original_duration_days=None,    # não usamos OD
            )
            _, is_created = ScheduleActivity.objects.update_or_create(activity_id=aid, defaults=defaults)
            created += 1 if is_created else 0
            updated += 0 if is_created else 1
            jc_updates += _update_jobcard_dates(aid, start, fin, update_jobcards)
    return created, updated, jc_updates

# ==================== Views ====================

@login_required(login_url="login")
def schedule_template(request):
    """
    Template CSV com separador ';' (Excel pt-BR abre em colunas).
    Cabeçalho:
    Level;Activity ID;Activity Name;HH;Activity % Complete;Start;Finish;Pontos;Original Duration
    Nota: O sistema usa **HH** como oficial. Se 'HH' vier vazio, 'Original Duration'
    será usado como fallback para HH.
    """
    headers = [
        "Level", "Activity ID", "Activity Name", "HH",
        "Activity % Complete", "Start", "Finish", "Pontos",
        "Original Duration"  # opcional (fallback)
    ]
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=';')  # ponto-e-vírgula
    writer.writerow(headers)
    resp = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = 'attachment; filename="schedule_template_br.csv"'
    return resp


@login_required(login_url="login")
@permission_required("jobcards.change_jobcard", raise_exception=True)
def schedule_upload(request):
    """
    Full refresh das atividades. Opcionalmente, propaga Start/Finish do P6
    para TODAS as JobCards com o mesmo Activity ID.
    Suporta AJAX (?ajax=1) retornando JSON {created, updated, jobcards_updated}.
    """
    if request.method != "POST":
        return redirect(f"{reverse('schedule_gantt')}?open_import=1")

    form = ScheduleImportForm(request.POST, request.FILES)
    if not form.is_valid():
        if request.GET.get("ajax"):
            return JsonResponse({"ok": False, "error": "invalid form"}, status=400)
        messages.error(request, "Formulário inválido. Verifique os campos e tente novamente.")
        return redirect(f"{reverse('schedule_gantt')}?import_error=1")

    mode = form.cleaned_data["mode"]
    update_jobcards = form.cleaned_data["update_jobcards"]
    f = request.FILES["file"]

    try:
        fname = (f.name or "").lower()
        _, ext = os.path.splitext(fname)
        data = f.read()  # guarda os bytes para reuso

        with transaction.atomic():
            # 1) zera o cronograma atual
            _reset_schedule()

            # 2) importa tudo de novo
            if mode == "csv" or (mode == "auto" and ext in (".csv",)):
                created, updated, jc_upd = parse_csv(io.BytesIO(data), update_jobcards)
            elif mode == "excel" or (mode == "auto" and ext in (".xlsx", ".xlsm")):
                created, updated, jc_upd = parse_xlsx(io.BytesIO(data), update_jobcards)
            elif mode == "auto" and ext == ".xls":
                raise RuntimeError("Arquivo .xls (Excel antigo) não suportado. Salve como .xlsx/.xlsm.")
            else:
                created, updated, jc_upd = parse_xml(io.BytesIO(data), update_jobcards)

            # Se nada veio, aborta (rollback automático)
            if (created + updated) == 0:
                raise RuntimeError("Template vazio/inesperado — nada importado.")

        if request.GET.get("ajax"):
            return JsonResponse({"ok": True, "created": created, "updated": updated, "jobcards_updated": jc_upd})

        messages.success(request, f"Importação concluída. Criados: {created} | Atualizados: {updated} | JobCards atualizadas: {jc_upd}.")
        return redirect(f"{reverse('schedule_gantt')}?imported=1")

    except Exception as e:
        if request.GET.get("ajax"):
            return JsonResponse({"ok": False, "error": str(e)}, status=500)
        messages.error(request, f"Falha ao importar: {e}")
        return redirect(f"{reverse('schedule_gantt')}?import_error=1")


@login_required(login_url="login")
def schedule_gantt(request):
    disciplines = list(
        JobCard.objects.exclude(discipline__isnull=True).exclude(discipline__exact="")
        .values_list("discipline", flat=True).distinct()
    )
    return render(request, "sistema/schedule/schedule_gantt.html", {"disciplines": disciplines})



@login_required(login_url="login")
def schedule_api(request):
    """
    Retorna atividades na ordem importada com level e TODAS as JobCards ligadas.
    FILTRO DE DATAS (estrito):
      - start={YYYY-MM-DD}  -> filtra por start__gte (somente campo start)
      - finish={YYYY-MM-DD} -> filtra por finish__lte (somente campo finish)
      - ambos               -> start__gte & finish__lte
    """
    q = request.GET.get("q", "").strip()
    discipline = request.GET.get("discipline", "").strip()
    start_f = request.GET.get("start")
    finish_f = request.GET.get("finish")

    acts = ScheduleActivity.objects.all()

    if q:
        acts = acts.filter(
            Q(activity_id__icontains=q) |
            Q(name__icontains=q) |
            Q(jobcard_number__icontains=q)
        )

    # --------- Filtro de datas (estrito) ----------
    # start filtra estritamente pelo campo start (>=)
    if start_f:
        try:
            sdate = datetime.strptime(start_f, "%Y-%m-%d").date()
            acts = acts.filter(start__gte=sdate)
        except ValueError:
            pass

    # finish filtra estritamente pelo campo finish (<=)
    if finish_f:
        try:
            fdate = datetime.strptime(finish_f, "%Y-%m-%d").date()
            acts = acts.filter(finish__lte=fdate)
        except ValueError:
            pass
    # ---------------------------------------------

    if discipline:
        ids = list(
            JobCard.objects.filter(discipline__iexact=discipline)
            .values_list("activity_id", flat=True)
        )
        acts = acts.filter(activity_id__in=ids)

    acts = acts.order_by("sort_index", "pk")

    # ============== JobCards por activity_id (com campos completos) ==============
    act_ids = list(acts.values_list("activity_id", flat=True))
    jmap = {}
    if act_ids:
        jc_qs = JobCard.objects.filter(activity_id__in=act_ids).values(
            "activity_id",
            "job_card_number",
            "job_card_description",
            "working_code_description",
            "discipline",
            "start", "finish",
            "total_man_hours", "total_duration_hs", "indice_kpi",
            "status",
        )
        for jc in jc_qs:
            aid = jc.get("activity_id") or ""
            lst = jmap.setdefault(aid, [])
            s = jc["start"].isoformat() if jc.get("start") else ""
            f = jc["finish"].isoformat() if jc.get("finish") else ""
            lst.append({
                "job_card_number": jc.get("job_card_number") or "",
                "job_card_description": (jc.get("job_card_description") or "").strip(),
                "working_code_description": (jc.get("working_code_description") or "").strip(),
                "discipline": jc.get("discipline") or "",
                "start": s,
                "finish": f,
                "total_man_hours": jc.get("total_man_hours") or "",
                "total_duration_hs": jc.get("total_duration_hs") or "",
                "indice_kpi": jc.get("indice_kpi") or "",
                "status": jc.get("status") or "",
                "orig_duration": "",
            })

    def _safe_int(v, default=0):
        try:
            if v is None:
                return default
            x = float(v)
            if not math.isfinite(x):
                return default
            return int(round(x))
        except Exception:
            return default

    def _safe_num(v):
        try:
            if v is None:
                return None
            x = float(v)
            if not math.isfinite(x):
                return None
            return x
        except Exception:
            return None

    data = []
    for a in acts:
        start = a.start.isoformat() if a.start else ""
        end   = a.finish.isoformat() if a.finish else ""

        prog = _safe_int(getattr(a, "percent_complete", 0), 0)
        prog = max(0, min(100, prog))

        hh_val = _safe_num(getattr(a, "hh", None))

        data.append({
            "id": a.activity_id or "",
            "name": (a.name or a.activity_id or ""),
            "level": int(a.level or 0),
            "start": start,
            "end": end,
            "progress": prog,
            "hh": hh_val,
            "points": _safe_num(getattr(a, "points", None)),
            "orig_duration": "",
            "jobcards": jmap.get(a.activity_id, []),
        })

    resp = JsonResponse(
        {"tasks": data},
        json_dumps_params={"ensure_ascii": False, "allow_nan": False}
    )
    resp["Cache-Control"] = "no-store"
    return resp


# --- EXPORT: Excel SEM outline e SEM indent visual — apenas ESPAÇOS reais,
# ---         SEM linhas de grade e com borda pontilhada cinza por linha ---
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import date as _date
import math
import re

# ===== Config: somente espaços reais =====
INDENT_UNIT = "    "   # 4 espaços por nível (ajuste para "  " se preferir)
INDENT_MAX_LEVEL = 200 # teto de segurança

# ===== Borda base (pontilhada cinza médio) =====
ROW_BORDER_COLOR = "A6A6A6"  # cinza médio
ROW_BORDER_STYLE = "dotted"  # pontilhado

def _excel_date(v):
    """Converte valor em date real para o Excel."""
    if v is None:
        return None
    if isinstance(v, _date):
        return v
    try:
        d = _parse_date(v)  # usa seu helper do arquivo
        return d
    except Exception:
        return None

def _num_or_none(v):
    try:
        if v is None:
            return None
        x = float(v)
        return x if math.isfinite(x) else None
    except Exception:
        return None

def _is_canceled_py(obj: dict) -> bool:
    """Detecção de JC cancelada (PT/EN), igual à tela."""
    if not obj:
        return False
    # flags booleanas
    for k in ("is_canceled", "is_cancelled", "canceled", "cancelled", "void", "is_void"):
        v = obj.get(k)
        if v in (True, 1) or str(v).strip().lower() == "true":
            return True
    # status textual
    for k in ("status", "jobcard_status", "situacao", "status_code"):
        v = obj.get(k)
        if not v:
            continue
        s = str(v)
        try:
            s = s.strip().lower().encode("ascii", "ignore").decode("ascii")
        except Exception:
            s = s.strip().lower()
        if re.search(r"cancelad|cancel+ed|void(ed)?|anulad|aborted?", s):
            return True
    return False

@login_required(login_url="login")
def schedule_export_excel(request):
    """
    XLSX com:
      - Hierarquia por ESPAÇOS na coluna A (sem outline/indent visual).
      - Sem linhas de grade; borda pontilhada cinza por linha.
      - 'ID Level' e 'Raw ID' ocultos.
      - HH da atividade = ScheduleActivity.hh.
      - Filtra JC canceladas.
      - Datas como date real.
    FILTRO DE DATAS (estrito, igual ao schedule_api):
      - start={YYYY-MM-DD}  -> start__gte
      - finish={YYYY-MM-DD} -> finish__lte
    """
    q = request.GET.get("q", "").strip()
    discipline = request.GET.get("discipline", "").strip()
    start_f = request.GET.get("start")
    finish_f = request.GET.get("finish")

    acts = ScheduleActivity.objects.all()

    if q:
        acts = acts.filter(
            Q(activity_id__icontains=q) |
            Q(name__icontains=q) |
            Q(jobcard_number__icontains=q)
        )

    # --------- Filtro de datas (estrito) ----------
    if start_f:
        try:
            sdate = datetime.strptime(start_f, "%Y-%m-%d").date()
            acts = acts.filter(start__gte=sdate)
        except ValueError:
            pass

    if finish_f:
        try:
            fdate = datetime.strptime(finish_f, "%Y-%m-%d").date()
            acts = acts.filter(finish__lte=fdate)
        except ValueError:
            pass
    # ---------------------------------------------

    if discipline:
        ids = list(
            JobCard.objects.filter(discipline__iexact=discipline)
            .values_list("activity_id", flat=True)
        )
        acts = acts.filter(activity_id__in=ids)

    acts = acts.order_by("sort_index", "pk")

    # ===== JobCards por activity_id (filtrando canceladas) =====
    act_ids = list(acts.values_list("activity_id", flat=True))
    jmap = {}
    if act_ids:
        jc_qs = JobCard.objects.filter(activity_id__in=act_ids).values(
            "activity_id",
            "job_card_number",
            "job_card_description",
            "working_code_description",
            "discipline",
            "start", "finish",
            "total_man_hours", "total_duration_hs", "indice_kpi",
            "status", "jobcard_status",
        )
        for jc in jc_qs:
            if _is_canceled_py(jc):
                continue
            jmap.setdefault(jc.get("activity_id") or "", []).append(jc)

    def _jc_hh(j):
        return _num_or_none(j.get("total_man_hours")) \
               or _num_or_none(j.get("total_duration_hs")) \
               or _num_or_none(j.get("orig_duration"))

    # ===== Linhas flat com nível =====
    rows = []
    for a in acts:
        lvl = int(a.level or 0)

        rows.append({
            "lvl": lvl,
            "is_jc": False,
            "id": a.activity_id or "",
            "name": a.name or "",
            "hh": _num_or_none(getattr(a, "hh", None)),
            "pts": _num_or_none(getattr(a, "points", None)),
            "pct": max(0, min(100, int(_num_or_none(getattr(a, "percent_complete", 0)) or 0))) / 100.0,
            "start": a.start,
            "finish": a.finish,
            "wk_desc": "",
        })

        jcs = jmap.get(a.activity_id or "", [])
        hh_array = [(_jc_hh(j) or 0) for j in jcs]
        hh_sum = sum(hh_array) if hh_array else 0.0
        act_pts = _num_or_none(getattr(a, "points", None))

        for j in jcs:
            j_h = _jc_hh(j) or 0.0
            share = (j_h / hh_sum) if hh_sum > 0 else 0.0
            pts_calc = (act_pts * share) if (act_pts is not None) else None

            rows.append({
                "lvl": lvl + 1,
                "is_jc": True,
                "id": (j.get("job_card_number") or "").strip() or (a.activity_id or ""),
                "name": (j.get("job_card_description") or "").strip(),
                "hh": j_h,
                "pts": pts_calc,
                "pct": 0.0,
                "start": j.get("start"),
                "finish": j.get("finish"),
                "wk_desc": (j.get("working_code_description") or "").strip(),
            })

    # ===== workbook (resto permanece igual) =====
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    ws.sheet_view.showGridLines = False
    ws.print_options.gridLines = False

    headers = [
        "Activity / JobCard",
        "Name / Description",
        "HH (h)",
        "Pts",
        "%",
        "Start",
        "Finish",
        "ID Level",
        "Raw ID",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="111111")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col, value=headers[col - 1])
        c.font = header_font
        c.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(2, len(rows)+1)}"

    widths = [34, 56, 14, 12, 8, 12, 12, 8, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    dotted_bottom = Side(style=ROW_BORDER_STYLE, color=ROW_BORDER_COLOR)

    def apply_row_dotted_border(row_idx: int, first_col: int, last_col: int):
        for col in range(first_col, last_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            b = cell.border
            cell.border = Border(
                left=b.left, right=b.right, top=b.top, bottom=dotted_bottom
            )

    def left_bar(color_hex):
        return Border(
            left=Side(style="thin", color=color_hex),
            bottom=dotted_bottom
        )

    fill_gray  = PatternFill("solid", fgColor="F3F4F6")
    fill_beige = PatternFill("solid", fgColor="F5F0E6")

    first_row = 2
    for i, r in enumerate(rows, start=first_row):
        lvl = max(0, int(r["lvl"]))
        lvl_spaces = min(lvl, INDENT_MAX_LEVEL)
        raw_id = r["id"] or ""
        display_id = (INDENT_UNIT * lvl_spaces) + raw_id

        c_id = ws.cell(row=i, column=1, value=display_id)
        c_id.alignment = Alignment(indent=0, vertical="center")

        ws.cell(row=i, column=2, value=(r["name"] or r["wk_desc"] or ""))

        c_hh = ws.cell(row=i, column=3, value=(r["hh"] if r["hh"] is not None else None))
        if r["hh"] is not None:
            c_hh.number_format = '#,##0.0" h"'

        c_pts = ws.cell(row=i, column=4, value=(r["pts"] if r["pts"] is not None else None))
        if r["pts"] is not None:
            c_pts.number_format = '#,##0.##'

        c_pct = ws.cell(row=i, column=5, value=r["pct"])
        c_pct.number_format = "0.00%"

        st = _excel_date(r["start"])
        fn = _excel_date(r["finish"])
        if st:
            cs = ws.cell(row=i, column=6, value=st)
            cs.number_format = "DD/MM/YYYY"
        if fn:
            cf = ws.cell(row=i, column=7, value=fn)
            cf.number_format = "DD/MM/YYYY"

        ws.cell(row=i, column=8, value=lvl)     # ID Level
        ws.cell(row=i, column=9, value=raw_id)  # Raw ID

        apply_row_dotted_border(i, 1, len(headers))

    ws.column_dimensions[get_column_letter(8)].hidden = True
    ws.column_dimensions[get_column_letter(9)].hidden = True

    last_col = len(headers)
    n = len(rows)
    for idx in range(n):
        excel_row = first_row + idx
        cur_lvl = int(rows[idx]["lvl"])
        next_lvl = int(rows[idx + 1]["lvl"]) if (idx + 1) < n else cur_lvl
        if next_lvl > cur_lvl >= 0:
            rng = ws[f"A{excel_row}:{get_column_letter(last_col)}{excel_row}"]
            for cell in rng[0]:
                cell.font = Font(bold=True, color="111111")
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=dotted_bottom
                )
            pal = cur_lvl % 3
            if pal == 0:
                for cell in rng[0]:
                    cell.border = Border(
                        left=Side(style="thin", color="111111"),
                        right=cell.border.right, top=cell.border.top,
                        bottom=dotted_bottom
                    )
            elif pal == 1:
                for cell in rng[0]:
                    cell.fill = fill_gray
                    cell.border = Border(
                        left=Side(style="thin", color="9CA3AF"),
                        right=cell.border.right, top=cell.border.top,
                        bottom=dotted_bottom
                    )
            else:
                for cell in rng[0]:
                    cell.fill = fill_beige
                    cell.border = Border(
                        left=Side(style="thin", color="B49C78"),
                        right=cell.border.right, top=cell.border.top,
                        bottom=dotted_bottom
                    )

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="taskfy_schedule.xlsx"'
    return resp


