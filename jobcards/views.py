from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login as auth_login
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils import timezone
from django.db import transaction
from collections import defaultdict
from datetime import datetime
from decimal import Decimal
import os
import re
import pdfkit
import barcode
from barcode.writer import ImageWriter
from io import BytesIO
from django.views.generic import ListView, CreateView
from django.urls import reverse_lazy
from .models import Discipline, Area, WorkingCode, System
from .forms import DisciplineForm, AreaForm, WorkingCodeForm, SystemForm, ImpedimentsForm, JobCardImageForm
import datetime
from django.conf import settings
from .models import (
    JobCard, TaskBase, ManpowerBase, MaterialBase, ToolsBase, EngineeringBase,
    AllocatedEngineering, AllocatedManpower, AllocatedMaterial, AllocatedTool, AllocatedTask,
    Discipline, Area, WorkingCode, System, Impediments, PMTOBase, MRBase, ProcurementBase, WarehouseStock
)
import tempfile
from django.db.models import Sum
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from .models import MaterialBase, JobCard
import pandas as pd
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST
from datetime import datetime
from datetime import date
import io
from jobcards.models import DocumentoRevisaoAlterada, DocumentoControle, DailyFieldReport

from django.shortcuts import render, get_object_or_404, redirect
from django.db import transaction
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
import re
from collections import defaultdict
from .models import (
    JobCard, TaskBase, ManpowerBase, AllocatedManpower, AllocatedTask,
    MaterialBase, AllocatedMaterial, ToolsBase, AllocatedTool,
    EngineeringBase, AllocatedEngineering
)

from django.shortcuts import render, get_object_or_404, redirect
from django.db import transaction
from django.db.models import Sum
from django.contrib.auth.decorators import login_required
from collections import defaultdict
import re
from .models import JobCard, TaskBase, ManpowerBase, AllocatedManpower, AllocatedTask, MaterialBase, AllocatedMaterial, ToolsBase, AllocatedTool, EngineeringBase, AllocatedEngineering
import json
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.decorators import user_passes_test
from django.db.models.functions import TruncDate
from django.db.models import F
from .models import WarehouseStock

import json

from .models import ProcurementBase, WarehouseStock, WarehousePiece
from django.db import models
from django.contrib import messages
from .serializers import AllocatedManpowerSerializer, JobCardSerializer
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication

from django.http import FileResponse, Http404
from django.utils.encoding import smart_str

from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import WarehousePiece
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from .models import Impediments
from .serializers import ImpedimentSerializer

from collections import defaultdict
import re
from django.db.models import Sum
from .forms import JobCardForm, EDITABLE_FIELDS
import csv
from typing import List
from django.db import models as djm

from django.utils.http import url_has_allowed_host_and_scheme
from django.conf import settings
from django.urls import reverse, NoReverseMatch

# - PARTE DA INDEX DO SISTEMA
def index(request):
    return render(request, 'index.html')

# - PERMISSIONAMENTO POR GRUPO
def group_required(group_name):
    def in_group(u):
        return u.is_authenticated and u.groups.filter(name=group_name).exists()
    return user_passes_test(in_group, login_url='login', redirect_field_name=None)

# Caminho absoluto do executável dentro do projeto
path_wkhtmltopdf = os.path.join(settings.BASE_DIR, 'wkhtmltopdf', 'bin', 'wkhtmltopdf.exe')
config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)

# - PARTE DA INDEX DO SISTEMA
def _is_safe_next(target_url: str, request) -> bool:
    """
    Garante que o 'next' é seguro:
    - Mesmo host
    - HTTPS se a requisição for HTTPS
    - Não aponte para as rotas de login (evita loop)
    """
    if not target_url:
        return False

    # evita loop para as rotas de login
    login_urls = []
    try:
        login_urls.append(reverse('login'))
    except NoReverseMatch:
        pass
    try:
        login_urls.append(reverse('accounts_login'))
    except NoReverseMatch:
        pass
    if any(target_url.startswith(u) for u in login_urls):
        return False

    return url_has_allowed_host_and_scheme(
        target_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    )

def login(request):
    # Captura o next (GET tem prioridade quando chegou redirecionado por @login_required)
    next_url = request.GET.get('next') or request.POST.get('next') or ''

    # Se já está autenticado, respeita o next seguro, senão manda para dashboard
    if request.user.is_authenticated:
        if _is_safe_next(next_url, request):
            return redirect(next_url)
        return redirect('dashboard')

    if request.method == 'POST':
        username = (request.POST.get('username') or '').strip()
        password = request.POST.get('password') or ''

        user = authenticate(request, username=username, password=password)
        if user is not None:
            auth_login(request, user)

            # Após logar, respeita o next seguro
            if _is_safe_next(next_url, request):
                return redirect(next_url)
            return redirect('dashboard')

        # Credenciais inválidas → renderiza com erro e mantém o next
        return render(request, 'login.html', {'erro': 'Login Inválido', 'next': next_url})

    # GET normal → renderiza login e injeta o next (se houver)
    return render(request, 'login.html', {'next': next_url})

# - PARTE DO DASHBOARD

# ====== Constantes de status (cadeia) ======
STATUS_NO     = 'NO CHECKED'
STATUS_PRE    = 'PRELIMINARY JOBCARD CHECKED'
STATUS_PLAN   = 'PLANNING JOBCARD CHECKED'
STATUS_OFF    = 'OFFSHORE FIELD JOBCARD CHECKED'
STATUS_DONE   = 'JOBCARD COMPLETED'

# Conjuntos cumulativos
STAGE_PRE_OR_MORE   = [STATUS_PRE, STATUS_PLAN, STATUS_OFF, STATUS_DONE]
STAGE_PLAN_OR_MORE  = [STATUS_PLAN, STATUS_OFF, STATUS_DONE]
STAGE_OFF_OR_MORE   = [STATUS_OFF, STATUS_DONE]
STAGE_DONE_ONLY     = [STATUS_DONE]

@login_required(login_url='login')
def dashboard(request):
    # Excluir canceladas sempre
    CANCEL_Q = Q(jobcard_status__icontains='CANCEL')
    qs_active = JobCard.objects.exclude(CANCEL_Q)

    # Totais principais (somente ativos)
    total_jobcards    = qs_active.count()
    not_checked_count = qs_active.filter(jobcard_status=STATUS_NO).count()
    checked_count     = total_jobcards - not_checked_count

    # JobCards with Material: considere só as ativas
    active_numbers = qs_active.values_list('job_card_number', flat=True)
    jobcards_with_material = (
        MaterialBase.objects
        .exclude(job_card_number__isnull=True)
        .exclude(job_card_number__exact='')
        .filter(job_card_number__in=active_numbers)
        .values('job_card_number').distinct().count()
    )

    # Disciplinas (por discipline_code) - só ativas
    codigos       = list(qs_active.values_list('discipline_code', flat=True).distinct().order_by('discipline_code'))
    labels_total  = codigos
    data_total    = [qs_active.filter(discipline_code=cod).count() for cod in codigos]

    # NÃO CHECKED por disciplina (só ativas)
    not_checked_qs = (
        qs_active.filter(jobcard_status=STATUS_NO)
                 .values('discipline').annotate(count=Count('id')).order_by('discipline')
    )
    labels_not_checked = [r['discipline'] for r in not_checked_qs]
    data_not_checked   = [r['count'] for r in not_checked_qs]

    # AWP (só ativas)
    awp_data = defaultdict(lambda: defaultdict(list))
    for jc in qs_active.order_by('system', 'workpack_number', 'job_card_number'):
        if jc.system and jc.workpack_number:
            awp_data[jc.system][jc.workpack_number].append(jc)

    # Alertas de campos obrigatórios (só ativas)
    jobcards_incompletos = qs_active.filter(
        Q(job_card_number__isnull=True) | Q(job_card_number__exact='') |
        Q(prepared_by__isnull=True)     | Q(prepared_by__exact='')     |
        Q(discipline__isnull=True)      | Q(discipline__exact='')      |
        Q(job_card_description__isnull=True) | Q(job_card_description__exact='') |
        Q(date_prepared__isnull=True)
    )
    alerta_count = jobcards_incompletos.count()

    # Áreas (só ativas)
    area_qs     = qs_active.values('location').annotate(count=Count('id')).order_by('location')
    labels_areas = [x['location'] or '—' for x in area_qs]
    data_areas   = [x['count'] for x in area_qs]

    # Mini métricas (só ativas)
    level_xx_count                 = qs_active.exclude(level__iexact='XX').count()
    activity_to_be_verified_count  = qs_active.exclude(activity_id__iexact='to be verified').count()
    start_1900_count               = qs_active.exclude(start=date(1900, 1, 1)).count()
    finish_1900_count              = qs_active.exclude(finish=date(1900, 1, 1)).count()
    system_to_be_verified_count    = qs_active.exclude(system__iexact='to be verified').count()
    subsystem_to_be_verified_count = qs_active.exclude(subsystem__iexact='to be verified').count()

    # ===== Contagens CUMULATIVAS por estágio =====
    preliminary_checked_count = qs_active.filter(jobcard_status__in=STAGE_PRE_OR_MORE).count()
    planning_checked_count    = qs_active.filter(jobcard_status__in=STAGE_PLAN_OR_MORE).count()
    offshore_checked_count    = qs_active.filter(jobcard_status__in=STAGE_OFF_OR_MORE).count()
    finalized_count           = qs_active.filter(jobcard_status__in=STAGE_DONE_ONLY).count()

    # (opcional) outros status fora da cadeia podem continuar isolados
    approved_to_execute_count = qs_active.filter(jobcard_status='RELEASED FOR EXECUTION').count()

    # Legend de disciplinas (só ativas)
    discipline_legend = qs_active.values_list('discipline_code', 'discipline').distinct().order_by('discipline_code')

    # Eng. sincronizados
    engineering_synced_docs = EngineeringBase.objects.filter(
        document__in=DocumentoControle.objects.values_list('codigo', flat=True)
    )

    # Ordem visual de disciplinas
    discipline_order = [
        'LOGISTIC','SCAFFOLDING','STRUCTURAL','MECHANICAL','PIPING','TIE-IN',
        'HVAC','INSTRUMENTATION & AUTOMATION','ELECTRICAL','TELECOM','PAINTING','COMMISSIONING',
    ]

    # ===== Discipline summary (cumulativo >= PRELIMINARY) =====
    discipline_summary = []
    for d in qs_active.values('discipline').distinct():
        total   = qs_active.filter(discipline=d['discipline']).count()
        checked = qs_active.filter(discipline=d['discipline'],
                                   jobcard_status__in=STAGE_PRE_OR_MORE).count()
        percent = (checked / total * 100) if total else 0.0
        discipline_summary.append({
            'discipline': d['discipline'],
            'total_jobcard': total,
            'total_checked': checked,
            'percent_checked': percent,
        })

    # Ordena conforme ordem fixa (preenche zeros p/ faltantes)
    disc_dict = {d['discipline']: d for d in discipline_summary}
    discipline_summary_sorted = []
    for name in discipline_order:
        discipline_summary_sorted.append(disc_dict.get(name, {
            'discipline': name, 'total_jobcard': 0, 'total_checked': 0, 'percent_checked': 0.0
        }))

    # ===== Area summary (cumulativo >= PRELIMINARY) =====
    area_groups = defaultdict(list)
    for a in Area.objects.all():
        area_groups[a.area_code].append(a)

    area_summary = []
    for area_code in sorted(area_groups):
        codes   = [a.area_code for a in area_groups[area_code]]
        total   = qs_active.filter(location__in=codes).count()
        checked = qs_active.filter(location__in=codes,
                                   jobcard_status__in=STAGE_PRE_OR_MORE).count()
        percent = (checked / total * 100) if total else 0.0
        area_obj = Area.objects.filter(area_code=area_code).order_by('pk').first()
        area_summary.append({
            'area_code': area_code,
            'area_description': area_obj.location if area_obj else '',
            'total_jobcard': total,
            'total_checked': checked,
            'percent_checked': percent,
        })

    # ===== Workpack summary (cumulativo >= PRELIMINARY) =====
    workpack_summary = []
    distinct_workpacks = (qs_active
                          .exclude(workpack_number__isnull=True)
                          .exclude(workpack_number__exact='')
                          .values('workpack_number').distinct())
    for w in distinct_workpacks:
        wp = w['workpack_number']
        total   = qs_active.filter(workpack_number=wp).count()
        checked = qs_active.filter(workpack_number=wp,
                                   jobcard_status__in=STAGE_PRE_OR_MORE).count()
        percent = (checked / total * 100) if total else 0.0
        workpack_summary.append({
            'workpack': wp,
            'total_jobcard': total,
            'total_checked': checked,
            'percent_checked': percent,
        })

    # Totais/percentuais globais
    workpack_total_jobcard   = sum(w['total_jobcard'] for w in workpack_summary)
    workpack_total_checked   = sum(w['total_checked'] for w in workpack_summary)
    workpack_percent_checked = (workpack_total_checked / workpack_total_jobcard * 100) if workpack_total_jobcard else 0.0

    discipline_total_jobcard   = sum(d['total_jobcard'] for d in discipline_summary_sorted)
    discipline_total_checked   = sum(d['total_checked'] for d in discipline_summary_sorted)
    discipline_percent_checked = (discipline_total_checked / discipline_total_jobcard * 100) if discipline_total_jobcard else 0.0

    area_total_jobcard   = sum(a['total_jobcard'] for a in area_summary)
    area_total_checked   = sum(a['total_checked'] for a in area_summary)
    area_percent_checked = (area_total_checked / area_total_jobcard * 100) if area_total_jobcard else 0.0

    # ===== Gráfico: PRELIMINARY per day (usar o carimbo, não o status atual) =====
    checked_daily = (
        qs_active.filter(checked_preliminary_at__isnull=False)
                 .annotate(day=TruncDate('checked_preliminary_at'))
                 .values('day').annotate(count=Count('id')).order_by('day')
    )
    checked_days   = [d['day'].strftime('%d/%m') for d in checked_daily if d['day']]
    checked_counts = [d['count'] for d in checked_daily if d['day']]

    # “Recently Checked JobCards” pelo carimbo, independente do status atual
    recent_checked_jobcards = (
        qs_active.filter(checked_preliminary_at__isnull=False)
                 .order_by('-checked_preliminary_at')[:5]
    )

    # Amostra
    jobcards = qs_active.all()[:20]

    # PDFs disponíveis
    backups_dir     = os.path.join(settings.BASE_DIR, 'jobcard_backups')
    available_pdfs  = {f for f in os.listdir(backups_dir) if f.endswith('.pdf')} if os.path.isdir(backups_dir) else set()

    # Percentuais (cada estágio sozinho e cumulativo)
    preliminary_percent = f"{(preliminary_checked_count / total_jobcards * 100):.2f}" if total_jobcards else "0.00"
    planning_percent    = f"{(planning_checked_count    / total_jobcards * 100):.2f}" if total_jobcards else "0.00"
    offshore_percent    = f"{(offshore_checked_count    / total_jobcards * 100):.2f}" if total_jobcards else "0.00"
    approved_percent    = f"{(approved_to_execute_count / total_jobcards * 100):.2f}" if total_jobcards else "0.00"

    # ======================================================================
    # ===============  NOVAS TABELAS: APENAS PRE/TAM 2026  =================
    # ======================================================================
    WORKPACK_FILTER = (
        Q(workpack_number__icontains='PRE TAM 2026') |
        Q(workpack_number__icontains='TAM 2026')
    )
    qs_tam = qs_active.filter(WORKPACK_FILTER)

    # ---- Discipline summary (PRE/TAM 2026; cumulativo >= PRELIMINARY)
    discipline_summary_tam = []
    for d in qs_tam.values('discipline').distinct():
        disc_name = d['discipline']
        total_t   = qs_tam.filter(discipline=disc_name).count()
        checked_t = qs_tam.filter(discipline=disc_name,
                                  jobcard_status__in=STAGE_PRE_OR_MORE).count()
        percent_t = (checked_t / total_t * 100) if total_t else 0.0
        discipline_summary_tam.append({
            'discipline': disc_name,
            'total_jobcard': total_t,
            'total_checked': checked_t,
            'percent_checked': percent_t,
        })

    # Ordena pela mesma ordem fixa
    disc_dict_tam = {d['discipline']: d for d in discipline_summary_tam}
    discipline_summary_tam_sorted = []
    for name in discipline_order:
        discipline_summary_tam_sorted.append(disc_dict_tam.get(name, {
            'discipline': name, 'total_jobcard': 0, 'total_checked': 0, 'percent_checked': 0.0
        }))

    discipline_total_jobcard_tam   = sum(d['total_jobcard'] for d in discipline_summary_tam_sorted)
    discipline_total_checked_tam   = sum(d['total_checked'] for d in discipline_summary_tam_sorted)
    discipline_percent_checked_tam = (discipline_total_checked_tam / discipline_total_jobcard_tam * 100) if discipline_total_jobcard_tam else 0.0

    # ---- Area summary (PRE/TAM 2026; cumulativo >= PRELIMINARY)
    area_summary_tam = []
    for area_code in sorted(area_groups):
        codes    = [a.area_code for a in area_groups[area_code]]
        total_a  = qs_tam.filter(location__in=codes).count()
        checked_a= qs_tam.filter(location__in=codes, jobcard_status__in=STAGE_PRE_OR_MORE).count()
        percent_a= (checked_a / total_a * 100) if total_a else 0.0
        area_obj = Area.objects.filter(area_code=area_code).order_by('pk').first()
        area_summary_tam.append({
            'area_code': area_code,
            'area_description': area_obj.location if area_obj else '',
            'total_jobcard': total_a,
            'total_checked': checked_a,
            'percent_checked': percent_a,
        })

    area_total_jobcard_tam   = sum(a['total_jobcard'] for a in area_summary_tam)
    area_total_checked_tam   = sum(a['total_checked'] for a in area_summary_tam)
    area_percent_checked_tam = (area_total_checked_tam / area_total_jobcard_tam * 100) if area_total_jobcard_tam else 0.0
    # ======================================================================

    context = {
        'total_jobcards'  : total_jobcards,
        'not_checked_count': not_checked_count,
        'checked_count'   : checked_count,

        'jobcards_with_material': jobcards_with_material,

        'labels_not_checked': labels_not_checked,
        'data_not_checked'  : data_not_checked,

        'labels_total': labels_total,
        'data_total'  : data_total,

        'awp_data'             : awp_data,
        'jobcards_incompletos' : jobcards_incompletos,
        'alerta_count'         : alerta_count,

        'labels_areas': labels_areas,
        'data_areas'  : data_areas,

        'level_xx_count'                : level_xx_count,
        'activity_to_be_verified_count' : activity_to_be_verified_count,
        'start_1900_count'              : start_1900_count,
        'finish_1900_count'             : finish_1900_count,
        'system_to_be_verified_count'   : system_to_be_verified_count,
        'subsystem_to_be_verified_count': subsystem_to_be_verified_count,

        # >>> CUMULATIVOS <<<
        'preliminary_checked_count': preliminary_checked_count,
        'planning_checked_count'   : planning_checked_count,
        'offshore_checked_count'   : offshore_checked_count,
        'approved_to_execute_count': approved_to_execute_count,
        'finalized_count'          : finalized_count,

        'discipline_legend'      : discipline_legend,
        'engineering_synced_docs': engineering_synced_docs,

        'preliminary_percent': preliminary_percent,
        'planning_percent'   : planning_percent,
        'offshore_percent'   : offshore_percent,
        'approved_percent'   : approved_percent,

        'discipline_summary'      : discipline_summary_sorted,
        'area_summary'            : area_summary,
        'workpack_summary'        : workpack_summary,

        'workpack_total_jobcard'  : workpack_total_jobcard,
        'workpack_total_checked'  : workpack_total_checked,
        'workpack_percent_checked': workpack_percent_checked,

        'discipline_total_jobcard'  : discipline_total_jobcard,
        'discipline_total_checked'  : discipline_total_checked,
        'discipline_percent_checked': discipline_percent_checked,

        'area_total_jobcard'  : area_total_jobcard,
        'area_total_checked'  : area_total_checked,
        'area_percent_checked': area_percent_checked,

        'checked_days'  : checked_days,
        'checked_counts': checked_counts,

        'recent_checked_jobcards': recent_checked_jobcards,
        'available_pdfs'        : available_pdfs,
        'jobcards'              : jobcards,

        # === NOVOS CONTEXTOS (apenas PRE/TAM 2026) ===
        'discipline_summary_tam'      : discipline_summary_tam_sorted,
        'discipline_total_jobcard_tam': discipline_total_jobcard_tam,
        'discipline_total_checked_tam': discipline_total_checked_tam,
        'discipline_percent_checked_tam': discipline_percent_checked_tam,

        'area_summary_tam'      : area_summary_tam,
        'area_total_jobcard_tam': area_total_jobcard_tam,
        'area_total_checked_tam': area_total_checked_tam,
        'area_percent_checked_tam': area_percent_checked_tam,
    }
    return render(request, 'sistema/dashboard.html', context)


# views.py
import os, re
from datetime import datetime
from typing import List

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db import models as djm
from django.db.models import Q
from django.shortcuts import render

from .models import JobCard


# ===== Helpers para a busca global =====
def _field_names_by_type(model, djangotypes: List[type]) -> List[str]:
    """Retorna nomes de campos concretos do model, filtrando por tipos Django."""
    names = []
    for f in model._meta.get_fields():
        if not hasattr(f, "attname"):  # ignora relações reversas
            continue
        if any(isinstance(f, t) for t in djangotypes):
            names.append(f.name)
    return names


def _build_global_search_q(model, raw_query: str) -> Q:
    """
    Busca global:
      - tokeniza por espaço (tokens AND)
      - texto -> icontains em todos Char/Text
      - número -> tenta casar em Integer/Float/Decimal (igualdade)
      - data -> dd/mm/YYYY ou YYYY-MM-DD em Date/DateTime (igualdade)
    """
    tokens = [t for t in re.split(r"\s+", (raw_query or "").strip()) if t]
    if not tokens:
        return Q()

    text_fields   = _field_names_by_type(model, [djm.CharField, djm.TextField])
    int_fields    = _field_names_by_type(model, [djm.IntegerField])
    num_fields    = _field_names_by_type(model, [djm.FloatField, djm.DecimalField])
    date_fields   = _field_names_by_type(model, [djm.DateField])
    datetime_flds = _field_names_by_type(model, [djm.DateTimeField])

    # prioriza campos comuns para otimizar o plano do DB
    preferred = [
        "job_card_number", "activity_id", "discipline", "location",
        "system", "subsystem", "working_code", "tag",
        "job_card_description", "working_code_description",
        "jobcard_status", "prepared_by", "approved_br", "comments", "status",
    ]
    text_fields = [f for f in preferred if f in text_fields] + [f for f in text_fields if f not in preferred]

    final_q = Q()
    for token in tokens:
        token_q = Q()

        # 1) Texto
        for fname in text_fields:
            token_q |= Q(**{f"{fname}__icontains": token})

        # 2) Número (int/float)
        if token.isdigit():
            val_int = int(token)
            for fname in int_fields:
                token_q |= Q(**{f"{fname}": val_int})
        else:
            try:
                val_float = float(token.replace(",", "."))
                for fname in num_fields:
                    token_q |= Q(**{f"{fname}": val_float})
            except Exception:
                pass

        # 3) Data (dd/mm/YYYY ou YYYY-MM-DD)
        parsed_date = None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                parsed_date = datetime.strptime(token, fmt).date()
                break
            except Exception:
                continue
        if parsed_date:
            for fname in date_fields:
                token_q |= Q(**{f"{fname}": parsed_date})
            for fname in datetime_flds:
                token_q |= Q(**{f"{fname}__date": parsed_date})

        # AND entre tokens
        final_q &= token_q

    return final_q


@login_required(login_url='login')
def jobcards_list(request):
    qs = JobCard.objects.all()

    # parâmetros
    search          = (request.GET.get('search') or '').strip()
    system_param    = (request.GET.get('system') or '').strip()
    discipline_param= (request.GET.get('discipline') or '').strip()  # << NOVO
    start_date_s    = request.GET.get('start_date', '')
    end_date_s      = request.GET.get('end_date', '')
    items_s         = request.GET.get('items_per_page', '10')

    # busca global (todos os campos)
    if search:
        qs = qs.filter(_build_global_search_q(JobCard, search))

    # filtro System (match exato, case-insensitive)
    if system_param:
        qs = qs.filter(system__iexact=system_param)

    # filtro Discipline (match exato, case-insensitive)  << NOVO
    if discipline_param:
        qs = qs.filter(discipline__iexact=discipline_param)

    # range de datas (start, finish, date_prepared, date_approved)
    def parse_input_date(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except Exception:
            return None

    d_start = parse_input_date(start_date_s)
    d_end   = parse_input_date(end_date_s)

    if d_start and d_end:
        qs = qs.filter(
            Q(start__range=(d_start, d_end)) |
            Q(finish__range=(d_start, d_end)) |
            Q(date_prepared__range=(d_start, d_end)) |
            Q(date_approved__range=(d_start, d_end))
        )
    elif d_start:
        qs = qs.filter(
            Q(start__gte=d_start) | Q(finish__gte=d_start) |
            Q(date_prepared__gte=d_start) | Q(date_approved__gte=d_start)
        )
    elif d_end:
        qs = qs.filter(
            Q(start__lte=d_end) | Q(finish__lte=d_end) |
            Q(date_prepared__lte=d_end) | Q(date_approved__lte=d_end)
        )

    # itens por página
    try:
        items_per_page = max(1, int(items_s))
    except (ValueError, TypeError):
        items_per_page = 10

    paginator = Paginator(qs.order_by('job_card_number'), items_per_page)
    page = request.GET.get('page')
    try:
        jobcards_page = paginator.page(page)
    except PageNotAnInteger:
        jobcards_page = paginator.page(1)
    except EmptyPage:
        jobcards_page = paginator.page(paginator.num_pages)

    # PDFs disponíveis (seguro se a pasta não existir)
    backups_dir = os.path.join(getattr(settings, 'BASE_DIR', ''), 'jobcard_backups')
    try:
        available_pdfs = {f for f in os.listdir(backups_dir) if f.lower().endswith('.pdf')}
    except Exception:
        available_pdfs = set()

    # listas distintas para selects (filtradas e ordenadas)
    systems_qs = JobCard.objects.values_list('system', flat=True).distinct()
    systems = sorted({(s or '').strip() for s in systems_qs if (s or '').strip()})

    # << NOVO: disciplinas distintas diretamente do JobCard
    disciplines_qs = JobCard.objects.values_list('discipline', flat=True).distinct()
    disciplines = sorted({(d or '').strip() for d in disciplines_qs if (d or '').strip()})

    context = {
        'jobcards': jobcards_page,
        'search': search,
        'system': system_param,
        'discipline': discipline_param,     # << NOVO
        'start_date': start_date_s,
        'end_date': end_date_s,
        'items_per_page': items_per_page,
        'available_pdfs': available_pdfs,
        'systems': systems,
        'disciplines': disciplines,         # << NOVO
    }
    return render(request, 'sistema/jobcards.html', context)



# - PARTE DO DASHBOARD
@login_required(login_url='login') # EXIGE O USUARIO A ESTAR LOGADO
@permission_required('jobcards.add_jobcard', raise_exception=True)
def create_jobcard(request, jobcard_id=None):
    return render(request, 'sistema/create_jobcard.html')

@login_required(login_url='login')
@permission_required('jobcards.change_jobcard', raise_exception=True)
def edit_jobcard(request, jobcard_id=None):
    

    job = get_object_or_404(JobCard, job_card_number=jobcard_id) if jobcard_id else None

    def safe_float(s):
        return float(s.replace('%', '').replace(',', '.')) if s and str(s).strip() else 0.0

    if request.method == 'POST' and job:
        # Atualiza campos principais
        job.jobcard_status = request.POST.get('JOBCARD_STATUS', job.jobcard_status)
        job.discipline = request.POST.get('DISCIPLINE', job.discipline)
        job.discipline_code = request.POST.get('DISCIPLINE_CODE', job.discipline_code)
        job.location = request.POST.get('LOCATION', job.location)
        job.level = request.POST.get('LEVEL', job.level)
        job.activity_id = request.POST.get('ACTIVITY_ID', job.activity_id)
        job.system = request.POST.get('SYSTEM', job.system)
        job.subsystem = request.POST.get('SUBSYSTEM', job.subsystem)
        job.workpack_number = request.POST.get('WORKPACK_NUMBER', job.workpack_number)
        job.start = request.POST.get('START') or None
        job.finish = request.POST.get('FINISH') or None
        job.working_code_description = request.POST.get('WORKING_CODE_DESCRIPTION', job.working_code_description)
        job.tag = request.POST.get('TAG', job.tag)
        job.job_card_description = request.POST.get('JOB_CARD_DESCRIPTION', job.job_card_description)
        job.comments = request.POST.get('COMMENTS', job.comments)
        job.total_weight = request.POST.get('TOTAL_WEIGHT', job.total_weight)
        job.unit = request.POST.get('UNIT', job.unit)
        job.indice_kpi = request.POST.get('INDICE_KPI', job.indice_kpi)
        job.prepared_by = request.POST.get('PREPARED_BY', job.prepared_by)
        job.date_prepared = request.POST.get('DATE_PREPARED') or None
        job.approved_br = request.POST.get('APPROVED_BR', job.approved_br)
        job.date_approved = request.POST.get('DATE_APPROVED') or None
        job.hot_work_required = request.POST.get('HOT_WORK_REQUIRED', job.hot_work_required)
        job.last_modified_by = request.user.username
        job.seq_number = f"{job.job_card_number.split('-')[-1]}"
        
        job.total_duration_hs = f"{float(request.POST.get('total_duration_hs', 0) or 0):.2f}"
        job.total_man_hours = f"{float(request.POST.get('total_man_hours', 0) or 0):.2f}"

        job.save()
        
        # Atualiza checked_preliminary_by e at se for o status correto
        if job.jobcard_status == 'PRELIMINARY JOBCARD CHECKED':
            # Só registra se nunca foi preenchido!
            if not job.checked_preliminary_by and not job.checked_preliminary_at:
                job.checked_preliminary_by = request.user.username
                job.checked_preliminary_at = timezone.now()
                job.save(update_fields=['checked_preliminary_by', 'checked_preliminary_at'])

            # … depois de salvar os campos do job e do safe_float …

        ### --- AllocatedManpower --- ###
        mp_pattern = re.compile(r'^mp-(\d+)-(\d+)-qty$')
        posted_manpowers = {}

        for key, val in request.POST.items():
            m = mp_pattern.match(key)
            if not m:
                continue
            task_order, mp_id = map(int, m.groups())
            qty = safe_float(val)
            hours = safe_float(request.POST.get(f"mp-{task_order}-{mp_id}-hh", '0'))
            if hours <= 0 or qty <= 0:
                continue
            posted_manpowers[(task_order, mp_id)] = (qty, hours)

        # Lista atual no banco
        existing_allocated = {
            (a.task_order, a.direct_labor): a
            for a in AllocatedManpower.objects.filter(jobcard_number=job.job_card_number)
        }

        # Processa o que veio no POST
        for (task_order, mp_id), (qty, hours) in posted_manpowers.items():
            mp = ManpowerBase.objects.filter(pk=mp_id).first()
            if not mp:
                continue
            key = (task_order, mp.direct_labor)
            if key in existing_allocated:
                obj = existing_allocated.pop(key)
                if obj.qty != qty or obj.hours != hours:
                    obj.qty = qty
                    obj.hours = hours
                    obj.save()
            else:
                AllocatedManpower.objects.create(
                    jobcard_number = job.job_card_number,
                    discipline     = mp.discipline,
                    working_code   = mp.working_code,
                    direct_labor   = mp.direct_labor,
                    qty            = qty,
                    hours          = hours,
                    task_order     = task_order,
                )

       
        ### --- fim AllocatedManpower --- ###



         ### --- AllocatedTask --- ###
        # Sempre sincronizar a descrição com a TaskBase atual
        task_list = TaskBase.objects.filter(working_code=job.working_code).order_by('order')

        # Mapa: order -> descrição atual da TaskBase
        task_desc = {t.order: t.typical_task for t in task_list}

        # Tarefas já alocadas para esta JobCard
        existing_tasks = {
            t.task_order: t
            for t in AllocatedTask.objects.filter(jobcard_number=job.job_card_number)
        }

        updated_orders = set()

        for task in task_list:
            order = task.order
            max_hh = safe_float(request.POST.get(f"hh-max-{order}", '0'))
            total_hh = safe_float(request.POST.get(f"hh-total-{order}", '0'))
            percent = safe_float(request.POST.get(f"hh-percent-{order}", '0'))
            not_applicable = request.POST.get(f"task-not-applicable-{order}") == 'on'
            desc = task_desc.get(order, '').strip()  # descrição 100% da base

            if order in existing_tasks:
                obj = existing_tasks.pop(order)
                changed = False

                if (obj.description or '') != desc:
                    obj.description = desc
                    changed = True
                if (obj.max_hours or 0.0) != max_hh:
                    obj.max_hours = max_hh
                    changed = True
                if (obj.total_hours or 0.0) != total_hh:
                    obj.total_hours = total_hh
                    changed = True
                if (obj.percent or 0.0) != percent:
                    obj.percent = percent
                    changed = True
                if bool(obj.not_applicable) != bool(not_applicable):
                    obj.not_applicable = not_applicable
                    changed = True

                if changed:
                    obj.save(update_fields=[
                        "description", "max_hours", "total_hours", "percent", "not_applicable"
                    ])
            else:
                # Cria já com a descrição ATUAL da TaskBase
                AllocatedTask.objects.create(
                    jobcard_number=job.job_card_number,
                    task_order=order,
                    description=desc,
                    max_hours=max_hh,
                    total_hours=total_hh,
                    percent=percent,
                    not_applicable=not_applicable,
                )

            updated_orders.add(order)

        # Remove tarefas que não existem mais na TaskBase (foram excluídas ou reordenadas)
        for order, obj in existing_tasks.items():
            if order not in updated_orders:
                obj.delete()
        ### --- fim AllocatedTask --- ###




        ### --- AllocatedMaterial --- ###
        project_codes = request.POST.getlist('project_code[]')
        descriptions = request.POST.getlist('description[]')
        jobcard_required_qtys = request.POST.getlist('jobcard_required_qty[]')
        comments = request.POST.getlist('comments[]')
        nps1_list = request.POST.getlist('nps1[]')
        existing_materials = list(AllocatedMaterial.objects.filter(jobcard_number=job.job_card_number))
        new_keys = []
        for code, desc, qty, comment, nps1 in zip(project_codes, descriptions, jobcard_required_qtys, comments, nps1_list):
            key = (code, desc, nps1)
            new_keys.append(key)
            obj = next((m for m in existing_materials if (m.pmto_code, m.description, m.nps1) == key), None)
            if obj:
                if obj.qty != safe_float(qty):
                    obj.qty = safe_float(qty)
                    obj.comments = comment
                    obj.save()
                existing_materials.remove(obj)
            else:
                AllocatedMaterial.objects.create(
                    jobcard_number=job.job_card_number,
                    discipline=job.discipline,
                    working_code=job.working_code,
                    pmto_code=code,
                    description=desc,
                    qty=safe_float(qty),
                    comments=comment,
                    nps1=nps1
                )
        for m in existing_materials:
            m.delete()

        ### --- AllocatedTool --- ###
        tool_items = request.POST.getlist('tool_item[]')
        tool_disciplines = request.POST.getlist('tool_discipline[]')
        tool_working_codes = request.POST.getlist('tool_working_code[]')
        tool_direct_labors = request.POST.getlist('tool_direct_labor[]')
        tool_qty_direct_labors = request.POST.getlist('tool_qty_direct_labor[]')
        tool_special_toolings = request.POST.getlist('tool_special_tooling[]')
        tool_qtys = request.POST.getlist('tool_qty[]')
        existing_tools = list(AllocatedTool.objects.filter(jobcard_number=job.job_card_number))
        new_tools_keys = []
        for item, discipline, working_code, direct_labor, qty_dl, special_tooling, qty in zip(
            tool_items, tool_disciplines, tool_working_codes, tool_direct_labors,
            tool_qty_direct_labors, tool_special_toolings, tool_qtys
        ):
            key = (direct_labor, special_tooling)
            new_tools_keys.append(key)
            obj = next((t for t in existing_tools if (t.direct_labor, t.special_tooling) == key), None)
            if obj:
                if obj.qty != safe_float(qty):
                    obj.qty = safe_float(qty)
                    obj.qty_direct_labor = safe_float(qty_dl)
                    obj.save()
                existing_tools.remove(obj)
            else:
                AllocatedTool.objects.create(
                    jobcard_number=job.job_card_number,
                    discipline=discipline,
                    working_code=working_code,
                    direct_labor=direct_labor,
                    qty_direct_labor=safe_float(qty_dl),
                    special_tooling=special_tooling,
                    qty=safe_float(qty)
                )

        for t in existing_tools:
            t.delete()

        ### --- AllocatedEngineering --- ###
        eng_disciplines = request.POST.getlist('eng_discipline[]')
        eng_documents = request.POST.getlist('eng_document[]')
        eng_tags = request.POST.getlist('eng_tag[]')
        eng_revs = request.POST.getlist('eng_rev[]')
        eng_statuses = request.POST.getlist('eng_status[]')
        existing_engs = list(AllocatedEngineering.objects.filter(jobcard_number=job.job_card_number))
        new_eng_keys = []
        for discipline, document, tag, rev, status in zip(eng_disciplines, eng_documents, eng_tags, eng_revs, eng_statuses):
            key = (document, tag, rev)
            new_eng_keys.append(key)
            obj = next((e for e in existing_engs if (e.document, e.tag, e.rev) == key), None)
            if obj:
                if obj.status != status:
                    obj.status = status
                    obj.save()
                existing_engs.remove(obj)
            else:
                AllocatedEngineering.objects.create(
                    jobcard_number=job.job_card_number,
                    discipline=discipline,
                    document=document,
                    tag=tag,
                    rev=rev,
                    status=status,
                )
        for e in existing_engs:
            e.delete()

        # Atualiza totais
        #total_duration_hs = AllocatedTask.objects.filter(jobcard_number=job.job_card_number).aggregate(total=Sum('max_hours'))['total'] or 0
        #total_man_hours = AllocatedTask.objects.filter(jobcard_number=job.job_card_number).aggregate(total=Sum('total_hours'))['total'] or 0
        #job.total_duration_hs = f'{total_duration_hs:.2f}'
        #job.total_man_hours = f'{total_man_hours:.2f}'
        job.rev = '1'
        job.save(update_fields=['total_duration_hs', 'total_man_hours', 'rev'])

        # --- Upload das imagens (salva com nomes baseados na jobcard, sobrescrevendo antigas) ---
        for i in range(1, 5):
            uploaded = request.FILES.get(f'image_{i}')
            if uploaded:
                try:
                    _save_jobcard_image(job, uploaded, i)
                except Exception as e:
                    # opcional: log/notify; não interrompe o fluxo principal
                    print(f"Erro salvando imagem {i} para {job.job_card_number}: {e}")
        # garante que campos image_* sejam persistidos
        job.save()

        return redirect(f"{reverse('generate_pdf', args=[job.job_card_number])}?src=edit&apply_status=1")

    # --- GET ---
    disciplinas = JobCard.objects.values_list('discipline', flat=True).distinct().order_by('discipline')
    discipline_codes = JobCard.objects.values_list('discipline_code', flat=True).distinct().order_by('discipline_code')
    locations = JobCard.objects.values_list('location', flat=True).distinct().order_by('location')
    levels = JobCard.objects.values_list('level', flat=True).distinct().order_by('level')
    activity_ids = JobCard.objects.values_list('activity_id', flat=True).distinct().order_by('activity_id')
    systems = JobCard.objects.values_list('system', flat=True).distinct().order_by('system')
    subsystems = JobCard.objects.values_list('subsystem', flat=True).distinct().order_by('subsystem')
    workpacks = JobCard.objects.values_list('workpack_number', flat=True).distinct().order_by('workpack_number')

    task_list = TaskBase.objects.filter(working_code=job.working_code).order_by('order') if job else []
    manpowers = ManpowerBase.objects.filter(working_code__in=[t.working_code for t in task_list])
    manpowers_dict = defaultdict(list)
    for mp in manpowers:
        manpowers_dict[mp.working_code].append(mp)

    materials_list = MaterialBase.objects.filter(job_card_number=job.job_card_number) if job else []
    tools_list = ToolsBase.objects.filter(working_code=job.working_code) if job else []
    engineering_list = EngineeringBase.objects.filter(jobcard_number=job.job_card_number) if job else []

    all_manpowers = ManpowerBase.objects.all().order_by('direct_labor', 'id')
    unique_manpowers = {}
    for mp in all_manpowers:
        if mp.direct_labor not in unique_manpowers:
            unique_manpowers[mp.direct_labor] = mp

    allocated_tasks_qs = AllocatedTask.objects.filter(jobcard_number=job.job_card_number)
    allocated_tasks_dict = {t.task_order: t for t in allocated_tasks_qs}

    allocated_manpowers_qs = AllocatedManpower.objects.filter(jobcard_number=job.job_card_number)
    allocated_manpowers_dict = defaultdict(list)
    for mp in allocated_manpowers_qs:
        allocated_manpowers_dict[mp.task_order].append(mp)

    def manpower_list_for_task(task):
        # Se houver allocated, mostra allocated, senão mostra base
        return allocated_manpowers_dict.get(task.order) or manpowers_dict.get(task.working_code, [])
    
    

    context = {
        'job': job,
        'disciplinas': disciplinas,
        'discipline_codes': discipline_codes,
        'locations': locations,
        'levels': levels,
        'activity_ids': activity_ids,
        'systems': systems,
        'subsystems': subsystems,
        'workpacks': workpacks,
        'task_list': task_list,
        'manpowers_dict': manpowers_dict,
        'all_manpowers': ManpowerBase.objects.all(),
        'materials_list': materials_list,
        'tools_list': tools_list,
        'engineering_list': engineering_list,
        'unique_manpowers': unique_manpowers.values(),
        'allocated_tasks_dict': allocated_tasks_dict,
        'allocated_manpowers_dict': allocated_manpowers_dict,
        'manpower_list_for_task': manpower_list_for_task,
    }
    
    print('POST DATA:', request.POST)
    return render(request, 'sistema/create_jobcard.html', context)

@login_required(login_url='login')
@permission_required('jobcards.change_jobcard', raise_exception=True)
def allocate_resources(request, jobcard_id):
    job = get_object_or_404(JobCard, job_card_number=jobcard_id)

    # Alocar Manpower
    for mp in ManpowerBase.objects.filter(working_code=job.working_code):
        AllocatedManpower.objects.update_or_create(
            jobcard_number=job.job_card_number,
            manpower_name=mp.manpower_name,
            defaults={
                'working_code': mp.working_code,
                'qty': mp.qty,
                'description': mp.description,
            }
        )

    # Alocar Materials
    for mat in MaterialBase.objects.filter(job_card_number=job.job_card_number):
        AllocatedMaterial.objects.update_or_create(
            jobcard_number=job.job_card_number,
            item=mat.item,
            defaults={
                'working_code': mat.working_code,
                'discipline': mat.discipline,
                'description': mat.description,
                'qty': mat.qty,
                'unit': mat.unit,
            }
        )

    # Alocar Tools
    for tool in ToolsBase.objects.filter(working_code=job.working_code):
        AllocatedTool.objects.update_or_create(
            jobcard_number=job.job_card_number,
            item=tool.item,
            defaults={
                'working_code': tool.working_code,
                'special_tooling': tool.special_tooling,
                'qty': tool.qty,
            }
        )

    return redirect('generate_pdf', jobcard_id=job.job_card_number)


from django.http import JsonResponse



@login_required(login_url='login')
@permission_required('jobcards.change_jobcard', raise_exception=True)
def generate_pdf(request, jobcard_id):
    """
    Regras de status:
      - Só altera status se vier ?apply_status=1
        - src=edit   : força PRELIMINARY JOBCARD CHECKED (com carimbo na 1ª vez).
        - src=modify : se vier ?status=..., sobrescreve o status no banco (e carimba se PRELIMINARY nunca carimbado).
      - Sem apply_status: NÃO toca no status; apenas renderiza o PDF.
    """
    from django.core.files.storage import default_storage  # garante import local se necessário

    job = get_object_or_404(JobCard, job_card_number=jobcard_id)
    area_info = Area.objects.filter(area_code=job.location).first() if job.location else None

    src            = (request.GET.get('src') or '').lower().strip()
    apply_status   = (request.GET.get('apply_status') == '1')
    incoming_status= request.GET.get('status')  # usado quando src=modify

    # -------- Código de barras --------
    barcode_folder = os.path.join(settings.BASE_DIR, 'static', 'barcodes')
    os.makedirs(barcode_folder, exist_ok=True)
    barcode_filename = f'{job.job_card_number}.png'
    barcode_path = os.path.join(barcode_folder, barcode_filename)
    if not os.path.exists(barcode_path):
        CODE128 = barcode.get_barcode_class('code128')
        code128 = CODE128(job.job_card_number, writer=ImageWriter())
        with open(barcode_path, 'wb') as bf:
            code128.write(bf, options={'write_text': False})
    barcode_url = f'file:///{barcode_path.replace("\\", "/")}'

    # -------- Regras de status (condicionais) --------
    if apply_status:
        if src == 'edit' or not src:
            if job.jobcard_status != 'PRELIMINARY JOBCARD CHECKED':
                job.jobcard_status = 'PRELIMINARY JOBCARD CHECKED'
                if not job.checked_preliminary_by and not job.checked_preliminary_at:
                    job.checked_preliminary_by = request.user.username
                    job.checked_preliminary_at = timezone.now()
                    job.save(update_fields=['jobcard_status', 'checked_preliminary_by', 'checked_preliminary_at'])
                else:
                    job.save(update_fields=['jobcard_status'])

        elif src == 'modify':
            if incoming_status:
                s = incoming_status.strip()
                if s and s != job.jobcard_status:
                    job.jobcard_status   = s
                    job.last_modified_by = request.user.username
                    job.last_modified_at = timezone.now()
                    fields = ['jobcard_status', 'last_modified_by', 'last_modified_at']

                    # se virou PRELIMINARY e nunca carimbou, carimba agora
                    if (
                        s == 'PRELIMINARY JOBCARD CHECKED'
                        and not job.checked_preliminary_by
                        and not job.checked_preliminary_at
                    ):
                        job.checked_preliminary_by = request.user.username
                        job.checked_preliminary_at = timezone.now()
                        fields += ['checked_preliminary_by', 'checked_preliminary_at']

                    job.save(update_fields=fields)
            # se não vier status, mantém o que já está salvo

    # -------- Query das alocações --------
    allocated_manpowers    = AllocatedManpower.objects.filter(jobcard_number=jobcard_id).order_by('task_order')
    allocated_materials    = AllocatedMaterial.objects.filter(jobcard_number=job.job_card_number)
    allocated_tools        = AllocatedTool.objects.filter(jobcard_number=jobcard_id)
    allocated_tasks        = AllocatedTask.objects.filter(jobcard_number=jobcard_id).order_by('task_order')
    allocated_engineerings = AllocatedEngineering.objects.filter(jobcard_number=job.job_card_number)

    # -------- Imagem de fundo --------
    image_path = os.path.join(settings.BASE_DIR, 'static', 'assets', 'img', '3.jpg')
    image_url  = f'file:///{image_path.replace("\\", "/")}'

    # -------- Marca d'água --------
    watermark_path = os.path.join(settings.BASE_DIR, 'static', 'assets', 'img', 'utc_vazio.png')
    watermark_url  = f'file:///{watermark_path.replace("\\", "/")}'
    

    # -------- Limpa image_* inexistentes --------
    updated_fields = []
    for i in range(1, 5):
        field = f'image_{i}'
        f = getattr(job, field, None)
        if f:
            exists = False
            try:
                if hasattr(f, 'path') and os.path.exists(f.path):
                    exists = True
                elif hasattr(f, 'name') and default_storage.exists(f.name):
                    exists = True
            except Exception:
                exists = False
            if not exists:
                setattr(job, field, None)
                updated_fields.append(field)
    if updated_fields:
        job.save(update_fields=updated_fields)

    # -------- Monta image_files para template --------
    image_files = {}
    for i in range(1, 5):
        field = f'image_{i}'
        f = getattr(job, field, None)
        if f:
            try:
                if hasattr(f, 'path') and os.path.exists(f.path):
                    image_files[field] = 'file:///' + f.path.replace('\\', '/')
                elif hasattr(f, 'name') and default_storage.exists(f.name):
                    storage_path = default_storage.path(f.name)
                    image_files[field] = 'file:///' + storage_path.replace('\\', '/')
                elif hasattr(f, 'url'):
                    image_files[field] = request.build_absolute_uri(f.url)
                else:
                    image_files[field] = None
            except Exception:
                image_files[field] = getattr(f, 'url', None)
        else:
            image_files[field] = None

    half = len(allocated_tools) // 2 if allocated_tools else 0

    # -------- Config wkhtmltopdf --------
    local_config = None
    try:
        if path_wkhtmltopdf and os.path.exists(path_wkhtmltopdf):
            local_config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
    except Exception:
        local_config = None

    # -------- Contexto --------
    context = {
        'job': job,
        'allocated_manpowers': allocated_manpowers,
        'allocated_materials': allocated_materials,
        'allocated_tools': allocated_tools,
        'allocated_tools_left': allocated_tools[:half],
        'allocated_tools_right': allocated_tools[half:],
        'allocated_tasks': allocated_tasks,
        'allocated_engineerings': allocated_engineerings,
        'image_path': image_url,
        'watermark_url': watermark_url,
        'barcode_image': barcode_url,
        'area_info': area_info,
        'image_files': image_files,
    }

    # -------- Render HTML --------
    html_string         = render_to_string('sistema/jobcard_pdf.html', context, request=request)
    header_html_string  = render_to_string('sistema/header.html', context, request=request)
    footer_html_string  = render_to_string('sistema/footer.html', context, request=request)
    header_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.html')
    footer_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.html')
    header_temp.write(header_html_string.encode('utf-8'))
    footer_temp.write(footer_html_string.encode('utf-8'))
    header_temp.close()
    footer_temp.close()

    # -------- Gera PDF --------
    try:
        pdf_file = pdfkit.from_string(
            html_string,
            False,
            configuration=local_config,
            options={
                'enable-local-file-access': None,
                'margin-top': '35mm',
                'margin-bottom': '30mm',
                'header-html': f'file:///{header_temp.name.replace("\\", "/")}',
                'footer-html': f'file:///{footer_temp.name.replace("\\", "/")}',
                'header-spacing': '5',
                'footer-spacing': '5',
                'quiet': ''
            }
        )
    except Exception as e:
        raise RuntimeError(
            f"wkhtmltopdf failed: {e}. "
            f"Verifique se imagens referenciadas existem no storage e se os campos image_* do JobCard foram atualizados."
        )

    # -------- Backup --------
    backup_filename = f'JobCard_{jobcard_id}_Rev_{job.rev}.pdf'
    backups_dir = os.path.join(settings.BASE_DIR, 'jobcard_backups')
    os.makedirs(backups_dir, exist_ok=True)
    backup_path = os.path.join(backups_dir, backup_filename)
    with open(backup_path, 'wb') as f:
        f.write(pdf_file)

    # -------- Limpa temporários --------
    try:
        os.unlink(header_temp.name)
        os.unlink(footer_temp.name)
    except Exception:
        pass

    # -------- Resposta --------
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=JobCard_{jobcard_id}_Rev_{job.rev}.pdf'
    return response


# --------- DATABASE EXIBIÇÕES --------------- #

def jobcards_overview(request):
    # Filtros e busca
    jobcards = JobCard.objects.all()
    search_number = request.GET.get('search_number', '').strip()
    search_discipline = request.GET.get('search_discipline', '').strip()
    search_prepared_by = request.GET.get('search_prepared_by', '').strip()
    search_location = request.GET.get('search_location', '').strip()
    search_status = request.GET.get('search_status', '').strip()
    global_search = request.GET.get('global_search', '').strip()

    if search_number:
        jobcards = jobcards.filter(job_card_number__icontains=search_number)
    if search_discipline:
        jobcards = jobcards.filter(discipline__icontains=search_discipline)
    if search_prepared_by:
        jobcards = jobcards.filter(prepared_by__icontains=search_prepared_by)
    if search_location:
        jobcards = jobcards.filter(location__icontains=search_location)
    if search_status:
        jobcards = jobcards.filter(jobcard_status__icontains=search_status)
    if global_search:
        jobcards = jobcards.filter(
            Q(job_card_number__icontains=global_search) |
            Q(discipline__icontains=global_search) |
            Q(prepared_by__icontains=global_search) |
            Q(location__icontains=global_search) |
            Q(jobcard_status__icontains=global_search)
        )

   # MULTIORDENAÇÃO AQUI (depois dos filtros, antes da paginação)
    order_by = request.GET.get('order_by', '')
    order_dir = request.GET.get('order_dir', '')
    order_by_list = order_by.split(',') if order_by else []
    order_dir_list = order_dir.split(',') if order_dir else []

    ordering = []
    for i, field in enumerate(order_by_list):
        dir = order_dir_list[i] if i < len(order_dir_list) else 'desc'
        ordering.append('-'+field if dir == 'desc' else field)

    if ordering:
        jobcards = jobcards.order_by(*ordering)
    else:
        jobcards = jobcards.order_by('-date_prepared')

    # Agora pagina
    page = request.GET.get('page', 1)
    per_page = int(request.GET.get('per_page', 18))
    paginator = Paginator(jobcards, per_page)
    jobcards_page = paginator.get_page(page)

    filters = {
        'search_number': search_number,
        'search_discipline': search_discipline,
        'search_prepared_by': search_prepared_by,
        'search_location': search_location,
        'search_status': search_status,
        'global_search': global_search,
        'per_page': per_page,
    }
    
    jobcard_columns = [
        ('item','Item'), ('seq_number','Seq Number'), ('discipline','Discipline'), ('discipline_code','Discipline Code'),
        ('location','Location'), ('level','Level'), ('activity_id','Activity ID'), ('start','Start'), ('finish','Finish'),
        ('system','System'), ('subsystem','Subsystem'), ('workpack_number','Workpack Number'), ('working_code','Working Code'),
        ('tag','Tag'), ('working_code_description','Working Code Description'), ('job_card_number','Job Card Number'),
        ('rev','Rev'), ('jobcard_status','Status'), ('job_card_description','Job Card Description'), ('completed','Completed'),
        ('total_weight','Total Weight'), ('unit','Unit'), ('total_duration_hs','Total Duration Hs'), ('indice_kpi','Indice KPI'),
        ('total_man_hours','Total Man Hours'), ('prepared_by','Prepared By'), ('date_prepared','Date Prepared'), ('approved_br','Approved By'),
        ('date_approved','Date Approved'), ('hot_work_required','Hot Work Required'), ('status','Status (Optional)'),
        ('comments','Comments'), ('last_modified_by','Last Modified By'), ('last_modified_at','Last Modified At')
    ]
    

    context = {
        'jobcards': jobcards_page,
        'filters': filters,
        'paginator': paginator,
        'order_by': order_by,
        'order_dir': order_dir,
        'order_by_list': order_by_list,
        'order_dir_list': order_dir_list,
        'jobcard_columns': jobcard_columns,
        
    }

    # AJAX: retorna só a tabela
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        table_html = render_to_string('sistema/databases/jobcards_table.html', context, request)
        pagination_html = render_to_string('sistema/databases/jobcards_pagination.html', context, request)
        return JsonResponse({'table_html': table_html, 'pagination_html': pagination_html})

    # Requisição normal: página completa
    return render(request, 'sistema/databases/jobcards_overview.html', context)



# --------- EXIBIÇÃO DE MATERIAIS BASE --------------- #

from django.db.models import OuterRef, Subquery

def materials_list(request):
    # Subconsulta correlacionada por job_card_number
    jc_qs = JobCard.objects.filter(job_card_number=OuterRef('job_card_number'))

    materials = (
        MaterialBase.objects
        .annotate(
            jc_status=Subquery(jc_qs.values('jobcard_status')[:1]),   # Status "novo" do JobCard
            jc_status_legacy=Subquery(jc_qs.values('status')[:1]),    # Status "legacy" (se existir no model)
            jc_rev=Subquery(jc_qs.values('rev')[:1])                  # Revisão do JobCard
        )
        .order_by('job_card_number', 'item')
    )

    return render(request, 'sistema/databases/materials_list.html', {'materials': materials})

# --------------------------------------------------- #

def manpower_list(request):
    manpowers = ManpowerBase.objects.all()
    context = {
        'manpowers': manpowers
    }
    return render(request, 'sistema/databases/manpower_list.html', context)

def tools_list(request):
    tools = ToolsBase.objects.all()
    context = {
        'tools': tools
    }
    return render(request, 'sistema/databases/tools_list.html', context)

def engineering_list(request):
    engineering = EngineeringBase.objects.all()
    context = {
        'engineering': engineering
    }
    return render(request, 'sistema/databases/engineering_list.html', context)

def task_list(request):
    tasks = TaskBase.objects.all()
    # Cria um dicionário: código → descrição
    code_to_desc = {wc.code.upper(): wc.description for wc in WorkingCode.objects.all()}
    context = {
        'tasks': tasks,
        'code_to_desc': code_to_desc
    }
    return render(request, 'sistema/databases/task_list.html', context)

# --------- DATABASE ALOCAÇÕES --------------- #


def allocated_tool_list(request):
    allocated_tools = AllocatedTool.objects.all()
    context = {
        'allocated_tools': allocated_tools
    }
    return render(request, 'sistema/allocated/allocated_tool_list.html', context)

def allocated_material_list(request):
    allocated_materials = AllocatedMaterial.objects.all()
    context = {
        'allocated_materials': allocated_materials
    }
    return render(request, 'sistema/allocated/allocated_material_list.html', context)

def allocated_engineering_list(request):
    allocated_engineering = AllocatedEngineering.objects.all()
    context = {
        'allocated_engineering': allocated_engineering
    }
    return render(request, 'sistema/allocated/allocated_engineering_list.html', context)


def allocated_task_list(request):
    allocated_tasks = AllocatedTask.objects.all()
    context = {
        'allocated_tasks': allocated_tasks
    }
    return render(request, 'sistema/allocated/allocated_task_list.html', context)

# --------- IMPORTAÇÕES BANCOS --------------- #

@csrf_exempt
def import_materials(request):
    if request.method == "POST":
        overwrite = request.POST.get('overwrite') == '1'
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({'status': 'error', 'message': 'No file uploaded.'})

        # Lê o arquivo (csv ou excel)
        try:
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Could not read file: {str(e)}'})

        # Lista das colunas obrigatórias (atualizadas!)
        required_columns = [
            'job_card_number', 'working_code', 'discipline', 'tag_jobcard_base',
            'jobcard_required_qty', 'unit_req_qty', 'weight_kg', 'material_segmentation',
            'comments', 'sequenc_no_procurement', 'status_procurement', 'mr_number',
            'basic_material', 'description', 'project_code', 'nps1', 'qty', 'unit', 'po',
            'reference_documents'
        ]
        missing = [col for col in required_columns if col not in df.columns]
        if missing:
            return JsonResponse({'status': 'error', 'message': f"Missing columns: {', '.join(missing)}"})

        # Padroniza os jobcards da planilha
        df['job_card_number'] = df['job_card_number'].astype(str).str.strip().str.upper()
        jobcards_in_file = df['job_card_number'].unique()
        errors = []
        # Valida existência de todos os jobcards antes de qualquer alteração
        for jc in jobcards_in_file:
            if not JobCard.objects.filter(job_card_number__iexact=jc).exists():
                errors.append(f"Job Card '{jc}' does not exist.")

        if errors:
            return JsonResponse({'status': 'error', 'message': " | ".join(errors)})

        # NOVA VALIDAÇÃO: cada job_card_number só pode ter 1 working_code e 1 discipline
        for jc in jobcards_in_file:
            working_codes = df[df['job_card_number'].str.upper() == jc.upper()]['working_code'].astype(str).str.upper().unique()
            disciplines = df[df['job_card_number'].str.upper() == jc.upper()]['discipline'].astype(str).str.upper().unique()
            if len(working_codes) > 1:
                return JsonResponse({'status': 'error', 'message': f"Job Card '{jc}' has more than one Working Code in the file: {', '.join(working_codes)}."})
            if len(disciplines) > 1:
                return JsonResponse({'status': 'error', 'message': f"Job Card '{jc}' has more than one Discipline in the file: {', '.join(disciplines)}."})

        # Confirma duplicidade para qualquer jobcard
        materials_exist = any(
            MaterialBase.objects.filter(job_card_number__iexact=jc).exists()
            for jc in jobcards_in_file
        )
        if materials_exist and not overwrite:
            return JsonResponse({'status': 'duplicate', 'message': 'Materials already registered for one or more Job Cards in your file. Overwrite?'})

        # Se overwrite, apaga antes de importar os novos
        if overwrite:
            for jc in jobcards_in_file:
                MaterialBase.objects.filter(job_card_number__iexact=jc).delete()

        # Insere todos os materiais de todas as jobcards válidas
        for idx, row in df.iterrows():
            try:
                jc = str(row['job_card_number']).strip().upper()
                MaterialBase.objects.create(
                    job_card_number = jc,
                    working_code = str(row.get('working_code', '')).strip().upper(),
                    discipline = str(row.get('discipline', '')).strip().upper(),
                    tag_jobcard_base = str(row.get('tag_jobcard_base', '')).strip().upper(),
                    jobcard_required_qty = float(row.get('jobcard_required_qty', 0)) if row.get('jobcard_required_qty') not in [None, ''] else None,
                    unit_req_qty = str(row.get('unit_req_qty', '')).strip().upper(),
                    weight_kg = float(row.get('weight_kg', 0)) if row.get('weight_kg') not in [None, ''] else None,
                    material_segmentation = str(row.get('material_segmentation', '')).strip().upper(),
                    comments = str(row.get('comments', '')).strip(),
                    sequenc_no_procurement = str(row.get('sequenc_no_procurement', '')).strip().upper(),
                    status_procurement = str(row.get('status_procurement', '')).strip().upper(),
                    mr_number = str(row.get('mr_number', '')).strip().upper(),
                    basic_material = str(row.get('basic_material', '')).strip().upper(),
                    description = str(row.get('description', '')).strip().upper(),
                    project_code = str(row.get('project_code', '')).strip().upper(),
                    nps1 = str(row.get('nps1', '')).strip().upper(),
                    qty = float(row.get('qty', 0)) if row.get('qty') not in [None, ''] else None,
                    unit = str(row.get('unit', '')).strip().upper(),
                    po = str(row.get('po', '')).strip().upper(),
                    reference_documents = str(row.get('reference_documents', '')).strip()
                )
            except Exception as e:
                return JsonResponse({
                    'status': 'error',
                    'message': f"Error on row {idx+2}: {str(e)}"
                })

        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request.'})


def import_jobcard(request):
    if request.method == "POST":
        overwrite = request.POST.get('overwrite') == '1'
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({'status': 'error', 'message': 'No file uploaded.'})

        try:
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Could not read file: {str(e)}'})

        required_fields = [
            "item",
            "seq_number",
            "discipline",
            "discipline_code",
            "location",
            "level",
            "activity_id",
            "start",
            "finish",
            "system",
            "subsystem",
            "workpack_number",
            "working_code",
            "tag",
            "working_code_description",
            "job_card_number",
            "job_card_description",
            "total_weight",
            "unit",
            "total_duration_hs",
            "indice_kpi",
            "total_man_hours",
            "comments",
            "hot_work_required",  # <- Agora faz parte dos required_fields!
        ]

        file_fields = list(df.columns)
        missing = [f for f in required_fields if f not in file_fields]
        extra = [f for f in file_fields if f not in required_fields]
        if missing:
            return JsonResponse({'status': 'error', 'message': f"Missing columns: {', '.join(missing)}"})
        if extra:
            return JsonResponse({'status': 'error', 'message': f"Extra/unexpected columns: {', '.join(extra)}"})

        empty_fields = []
        for field in required_fields:
            if df[field].isnull().any() or (df[field] == '').any():
                empty_fields.append(field)
        if empty_fields:
            return JsonResponse({'status': 'error', 'message': f"Please fill all required fields: {', '.join(empty_fields)}"})

        for _, row in df.iterrows():
            # ----------- Validação do formato do Job Card Number -----------
            job_card_number = str(row['job_card_number']).strip()
            pattern = r"^[A-Z0-9]{3}\.[A-Z0-9]{2}-[A-Z0-9]{2}-\d{4}$"
            if not re.match(pattern, job_card_number):
                return JsonResponse({
                    'status': 'error',
                    'message': f"Job Card Number '{job_card_number}' is invalid. The expected format is: A09.PS-EL-0001"
                })

            # --- Validação do campo hot_work_required ---
            hot_work_value = str(row.get('hot_work_required', '')).strip().capitalize()
            if hot_work_value not in ['Yes', 'No']:
                return JsonResponse({
                    'status': 'error',
                    'message': f"Invalid value for Hot Work Required in JobCard '{job_card_number}'. Only 'Yes' or 'No' are allowed."
                })

            jobcard_exist = JobCard.objects.filter(job_card_number=job_card_number).exists()

            # --- Converte campos de data, se necessário ---
            for date_field in ["start", "finish"]:
                if isinstance(row[date_field], str) and row[date_field]:
                    try:
                        row[date_field] = pd.to_datetime(row[date_field]).date()
                    except Exception:
                        return JsonResponse({'status': 'error', 'message': f"Invalid date format in {date_field} for Job Card '{job_card_number}'"})

            if jobcard_exist:
                if not overwrite:
                    return JsonResponse({
                        'status': 'duplicate',
                        'message': f"JobCard {job_card_number} already exists. Do you want to update this JobCard?"
                    })
                else:
                    # Atualiza, mas não mexe em rev!
                    jobcard = JobCard.objects.get(job_card_number=job_card_number)
                    for field in required_fields:
                        if field == 'hot_work_required':
                            setattr(jobcard, field, hot_work_value)
                        else:
                            setattr(jobcard, field, row[field])
                    jobcard.last_modified_by = request.user.username
                    jobcard.last_modified_at = timezone.now()
                    jobcard.save()
            else:
                # Cria nova JobCard normalmente
                data = {field: row[field] for field in required_fields}
                data["rev"] = "0"
                data["hot_work_required"] = hot_work_value
                data["jobcard_status"] = "NO CHECKED"
                data["completed"] = "NO"
                data["prepared_by"] = request.user.username
                data["date_prepared"] = timezone.now()
                data["approved_br"] = ""
                data["date_approved"] = None
                data["status"] = "NO"
                data["last_modified_by"] = request.user.username
                data["last_modified_at"] = timezone.now()
                data["offshore_field_check"] = "NO"
                jobcard = JobCard.objects.create(**data)

        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request.'})


def import_manpower(request):
    if request.method == "POST":
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({'status': 'error', 'message': 'No file uploaded.'})

        try:
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Could not read file: {str(e)}'})

        required_fields = [
            "item",
            "discipline",
            "working_code",
            "working_description",
            "direct_labor",
            "qty"
        ]

        file_fields = list(df.columns)
        missing = [f for f in required_fields if f not in file_fields]
        extra = [f for f in file_fields if f not in required_fields]
        if missing:
            return JsonResponse({'status': 'error', 'message': f"Missing columns: {', '.join(missing)}"})
        if extra:
            return JsonResponse({'status': 'error', 'message': f"Extra/unexpected columns: {', '.join(extra)}"})

        empty_fields = []
        for field in required_fields:
            if df[field].isnull().any() or (df[field] == '').any():
                empty_fields.append(field)
        if empty_fields:
            return JsonResponse({'status': 'error', 'message': f"Please fill all required fields: {', '.join(empty_fields)}"})

        # Descobrir todos os pares discipline+working_code do arquivo, sempre em maiúsculo
        discipline_codes = set(
            (str(row["discipline"]).strip().upper(), str(row["working_code"]).strip().upper())
            for _, row in df.iterrows()
        )

        # Apagar todos os registros existentes para esses pares, usando sempre upper
        for discipline, working_code in discipline_codes:
            ManpowerBase.objects.filter(
                discipline=discipline,
                working_code=working_code
            ).delete()

        # Checar unicidade dentro do arquivo e criar registros em CAPS LOCK
        seen = set()
        for _, row in df.iterrows():
            discipline = str(row["discipline"]).strip().upper()
            working_code = str(row["working_code"]).strip().upper()
            direct_labor = str(row["direct_labor"]).strip().upper()
            working_description = str(row["working_description"]).strip().upper()
            key = (discipline, working_code, direct_labor)
            if key in seen:
                return JsonResponse({'status': 'error', 'message': f"Duplicated direct_labor '{direct_labor}' for working_code '{working_code}' and discipline '{discipline}' in the file."})
            seen.add(key)

            # working_code deve existir (sempre upper)
            if not WorkingCode.objects.filter(code=working_code).exists():
                return JsonResponse({
                    'status': 'error',
                    'message': f"Working code '{working_code}' is not registered in the WorkingCode base. Only registered codes are allowed."
                })

            ManpowerBase.objects.create(
                item=row["item"],
                discipline=discipline,
                working_code=working_code,
                working_description=working_description,
                direct_labor=direct_labor,
                qty=row["qty"]
            )

        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request.'})


def import_toolsbase(request):
    if request.method == "POST":
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({'status': 'error', 'message': 'No file uploaded.'})

        try:
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Could not read file: {str(e)}'})

        required_fields = [
            "item",
            "discipline",
            "working_code",
            "direct_labor",
            "qty_direct_labor",
            "special_tooling",
            "qty"
        ]

        file_fields = list(df.columns)
        missing = [f for f in required_fields if f not in file_fields]
        extra = [f for f in file_fields if f not in required_fields]
        if missing:
            return JsonResponse({'status': 'error', 'message': f"Missing columns: {', '.join(missing)}"})
        if extra:
            return JsonResponse({'status': 'error', 'message': f"Extra/unexpected columns: {', '.join(extra)}"})

        empty_fields = []
        for field in required_fields:
            if df[field].isnull().any() or (df[field].astype(str).str.strip() == '').any():
                empty_fields.append(field)
        if empty_fields:
            return JsonResponse({'status': 'error', 'message': f"Please fill all required fields: {', '.join(empty_fields)}"})

        # Descobrir todos os pares discipline+working_code do arquivo
        discipline_codes = set(
            (str(row["discipline"]).strip().lower(), str(row["working_code"]).strip().lower())
            for _, row in df.iterrows()
        )

        # Apagar todos os registros existentes para esses pares
        for discipline, working_code in discipline_codes:
            ToolsBase.objects.filter(
                discipline__iexact=discipline,
                working_code__iexact=working_code
            ).delete()

        # Criar os novos registros (garante unicidade no arquivo também)
        seen = set()
        for _, row in df.iterrows():
            discipline = str(row["discipline"]).strip().upper()
            working_code = str(row["working_code"]).strip().upper()
            direct_labor = str(row["direct_labor"]).strip().upper()
            special_tooling = str(row["special_tooling"]).strip().upper()

            # Checa unicidade dentro do próprio arquivo (opcional, mas recomendado)
            key = (discipline, working_code, direct_labor, special_tooling)
            if key in seen:
                return JsonResponse({'status': 'error', 'message': f"Duplicated tool for direct labor '{direct_labor}', special tooling '{special_tooling}' in file."})
            seen.add(key)

            # working_code deve existir
            if not WorkingCode.objects.filter(code__iexact=working_code).exists():
                return JsonResponse({
                    'status': 'error',
                    'message': f"Working code '{row['working_code']}' is not registered in the WorkingCode base. Only registered codes are allowed."
                })

            ToolsBase.objects.create(
                item=int(row["item"]),
                discipline=discipline,
                working_code=working_code,
                direct_labor=direct_labor,
                qty_direct_labor=float(row["qty_direct_labor"]),
                special_tooling=special_tooling,
                qty=float(row["qty"])
            )

        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request.'})

def import_engineering(request):
    if request.method == "POST":
        overwrite = request.POST.get("overwrite") == "1"
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({'status': 'error', 'message': 'No file uploaded.'})

        try:
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Could not read file: {str(e)}'})

        required_fields = ["item", "discipline", "document", "jobcard_number", "rev", "status"]
        missing = [f for f in required_fields if f not in df.columns]
        extra = [f for f in df.columns if f not in required_fields]
        if missing:
            return JsonResponse({'status': 'error', 'message': f"Missing columns: {', '.join(missing)}"})
        if extra:
            return JsonResponse({'status': 'error', 'message': f"Extra/unexpected columns: {', '.join(extra)}"})

        # Verificar duplicatas por jobcard_number
        jobcards_in_file = df['jobcard_number'].unique().tolist()
        duplicates = EngineeringBase.objects.filter(jobcard_number__in=jobcards_in_file).exists()

        if duplicates and not overwrite:
            return JsonResponse({
                'status': 'duplicate',
                'message': 'Some jobcards already exist in the system. Do you want to overwrite them?'
            })

        if overwrite:
            EngineeringBase.objects.filter(jobcard_number__in=jobcards_in_file).delete()

        for _, row in df.iterrows():
            EngineeringBase.objects.create(
                item=row["item"],
                discipline=row["discipline"].strip(),
                document=row["document"].strip(),
                jobcard_number=row["jobcard_number"].strip(),
                rev=str(row["rev"]).strip(),
                status=row["status"].strip()
            )

        return JsonResponse({'status': 'ok'})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})


def import_taskbase(request):
    if request.method == "POST":
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({'status': 'error', 'message': 'No file uploaded.'})

        try:
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Could not read file: {str(e)}'})

        required_fields = [
            "item",
            "discipline",
            "working_code",
            "typical_task",
            "order"
        ]

        file_fields = list(df.columns)
        missing = [f for f in required_fields if f not in file_fields]
        extra = [f for f in file_fields if f not in required_fields]
        if missing:
            return JsonResponse({'status': 'error', 'message': f"Missing columns: {', '.join(missing)}"})
        if extra:
            return JsonResponse({'status': 'error', 'message': f"Extra/unexpected columns: {', '.join(extra)}"})

        empty_fields = []
        for field in required_fields:
            if df[field].isnull().any() or (df[field].astype(str).str.strip() == '').any():
                empty_fields.append(field)
        if empty_fields:
            return JsonResponse({'status': 'error', 'message': f"Please fill all required fields: {', '.join(empty_fields)}"})

        # Descobrir todos os pares discipline+working_code do arquivo
        discipline_codes = set(
            (str(row["discipline"]).strip().upper(), str(row["working_code"]).strip().upper())
            for _, row in df.iterrows()
        )

        # Apagar todos os registros existentes para esses pares
        for discipline, working_code in discipline_codes:
            TaskBase.objects.filter(
                discipline__iexact=discipline,
                working_code__iexact=working_code
            ).delete()

        # Criar os novos registros
        for _, row in df.iterrows():
            TaskBase.objects.create(
                item=int(row["item"]),
                discipline=str(row["discipline"]).strip().upper(),
                working_code=str(row["working_code"]).strip().upper(),
                typical_task=str(row["typical_task"]).strip().upper(),
                order=int(row["order"])
            )

        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request.'})

#REFERENCES FUNÇÕES 
class DisciplineListView(ListView):
    model = Discipline
    template_name = 'sistema/references/discipline_list.html'
    context_object_name = 'disciplines'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = DisciplineForm()
        return context

    def post(self, request, *args, **kwargs):
        form = DisciplineForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('disciplines_list')

class AreaListView(ListView):
    model = Area
    template_name = 'sistema/references/area_list.html'
    context_object_name = 'areas'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = AreaForm()
        return context

    def post(self, request, *args, **kwargs):
        form = AreaForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('areas_list')

class WorkingCodeListView(ListView):
    model = WorkingCode
    template_name = 'sistema/references/workingcode_list.html'
    context_object_name = 'workingcodes'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = WorkingCodeForm()
        return context

    def post(self, request, *args, **kwargs):
        form = WorkingCodeForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('workingcodes_list')

class SystemListView(ListView):
    model = System
    template_name = 'sistema/references/system_list.html'
    context_object_name = 'systems'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = SystemForm()
        return context

    def post(self, request, *args, **kwargs):
        form = SystemForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('systems_list')
    
    # Discipline Delete
  
  
# --------- DELETAR DADOS -------------- #



def delete_discipline(request, pk):
    discipline = get_object_or_404(Discipline, pk=pk)
    discipline.delete()
    return redirect('disciplines_list')  # Nome da URL que lista Disciplines

# System Delete

def delete_system(request, pk):
    system = get_object_or_404(System, pk=pk)
    system.delete()
    return redirect('systems_list')  # Nome da URL que lista Systems

# Working Code Delete


def delete_working_code(request, pk):
    working_code = get_object_or_404(WorkingCode, pk=pk)
    working_code.delete()
    return redirect('workingcodes_list')  # Nome da URL que lista Working Codes

# Area Delete


def delete_area(request, pk):
    area = get_object_or_404(Area, pk=pk)
    area.delete()
    return redirect('areas_list')  # Nome da URL que lista Areas


# --------- BAIXAR EXPORTAÇÕES --------- #


# imports necessários (ajuste o arquivo conforme seu projeto)
import pandas as pd
from django.http import HttpResponse
from django.db.models import Q, OuterRef, Subquery

from .models import MaterialBase, JobCard


def export_materials_excel(request):
    """
    Exporta o banco de materiais em XLSX aplicando os mesmos filtros da tela
    e incluindo os campos do JobCard logo após o Job Card Number:
      - jc_status (status "novo" do JobCard)
      - jc_status_legacy (status legado, se existir no model)
      - jc_rev (revisão)

    Parâmetros (GET):
      material      -> filtra Material Segmentation
      status        -> filtra em jc_status OU jc_status_legacy
      project_code  -> filtra Project Code
      search        -> busca global em diversos campos (inclui os anotados do JC)
    """

    # CAPTURA OS FILTROS ENVIADOS PELO FRONT
    material = (request.GET.get('material', '') or '').strip()
    status = (request.GET.get('status', '') or '').strip()
    project_code = (request.GET.get('project_code', '') or '').strip()
    global_search = (request.GET.get('search', '') or '').strip()

    # Subconsulta para "juntar" o JobCard por job_card_number
    jc_qs = JobCard.objects.filter(job_card_number=OuterRef('job_card_number'))

    # Inicia a query de materiais já com as ANOTAÇÕES do JobCard
    qs = (
        MaterialBase.objects
        .annotate(
            jc_status=Subquery(jc_qs.values('jobcard_status')[:1]),   # status "novo"
            jc_status_legacy=Subquery(jc_qs.values('status')[:1]),    # status "legacy" (se existir)
            jc_rev=Subquery(jc_qs.values('rev')[:1])                  # revisão do JC
        )
    )

    # FILTROS POR CAMPO INDIVIDUAL
    if material:
        qs = qs.filter(material_segmentation__icontains=material)

    if status:
        # Agora o filtro 'status' considera os status do JobCard (novo e legacy)
        qs = qs.filter(
            Q(jc_status__icontains=status) |
            Q(jc_status_legacy__icontains=status)
        )

    if project_code:
        qs = qs.filter(project_code__icontains=project_code)

    # FILTRO GLOBAL (BUSCA EM VÁRIOS CAMPOS, incluindo os anotados do JC)
    if global_search:
        qs = qs.filter(
            Q(item__icontains=global_search) |
            Q(job_card_number__icontains=global_search) |
            Q(jc_status__icontains=global_search) |
            Q(jc_status_legacy__icontains=global_search) |
            Q(jc_rev__icontains=global_search) |
            Q(working_code__icontains=global_search) |
            Q(discipline__icontains=global_search) |
            Q(tag_jobcard_base__icontains=global_search) |
            Q(material_segmentation__icontains=global_search) |
            Q(project_code__icontains=global_search) |
            Q(description__icontains=global_search) |
            Q(unit_req_qty__icontains=global_search) |
            Q(weight_kg__icontains=global_search) |
            Q(comments__icontains=global_search) |
            Q(status_procurement__icontains=global_search) |
            Q(mr_number__icontains=global_search) |
            Q(basic_material__icontains=global_search) |
            Q(nps1__icontains=global_search) |
            Q(unit__icontains=global_search) |
            Q(po__icontains=global_search) |
            Q(reference_documents__icontains=global_search)
        )

    # ORDENAÇÃO (opcional, pra ficar próximo do que aparece na UI)
    qs = qs.order_by('job_card_number', 'item')

    # MONTA O DATAFRAME COM OS CAMPOS NA MESMA ORDEM DO HTML:
    data = qs.values(
        'item',
        'job_card_number',
        'jc_status',            # novo: imediatamente após JC Number
        'working_code',
        'discipline',
        'tag_jobcard_base',
        'jobcard_required_qty',
        'unit_req_qty',
        'weight_kg',
        'material_segmentation',
        'comments',
        'sequenc_no_procurement',
        'status_procurement',
        'mr_number',
        'basic_material',
        'description',
        'project_code',
        'nps1',
        'qty',
        'unit',
        'po',
        'reference_documents'
    )

    df = pd.DataFrame(list(data)).fillna('')

    # FORMATAÇÃO NUMÉRICA: duas casas, vírgula (mantém compatível com seu padrão atual)
    for col in ['jobcard_required_qty', 'weight_kg', 'qty']:
        if col in df.columns:
            df[col] = df[col].apply(
                lambda x: '{:.2f}'.format(float(x)).replace('.', ',') if str(x).strip() != '' else ''
            )

    # GERA O EXCEL PARA DOWNLOAD
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=materials.xlsx'
    # Observação: requer openpyxl/xlsxwriter no ambiente
    df.to_excel(response, index=False)
    return response



def export_manpower_excel (request):
    discipline = request.GET.get('discipline', '')
    working_code = request.GET.get('working_code', '')
    direct_labor = request.GET.get('direct_labor', '')
    global_search = request.GET.get('search', '')

    qs = ManpowerBase.objects.all()

    if discipline:
        qs = qs.filter(discipline__icontains=discipline)
    if working_code:
        qs = qs.filter(working_code__icontains=working_code)
    if direct_labor:
        qs = qs.filter(direct_labor__icontains=direct_labor)
    if global_search:
        qs = qs.filter(
            Q(discipline__icontains=global_search) |
            Q(working_code__icontains=global_search) |
            Q(working_description__icontains=global_search) |
            Q(direct_labor__icontains=global_search)
        )

    data = qs.values(
        'item',
        'discipline',
        'working_code',
        'working_description',
        'direct_labor',
        'qty'
    )

    df = pd.DataFrame(list(data))

    # Formatação de qty com 2 casas
    if 'qty' in df.columns:
        df['qty'] = df['qty'].apply(lambda x: '{:.2f}'.format(x) if pd.notnull(x) else '')

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=manpower_base_export.xlsx'
    df.to_excel(response, index=False)
    return response


def export_toolsbase_excel(request):
    discipline = request.GET.get('discipline', '')
    working_code = request.GET.get('working_code', '')
    direct_labor = request.GET.get('direct_labor', '')
    global_search = request.GET.get('search', '')

    qs = ToolsBase.objects.all()

    if discipline:
        qs = qs.filter(discipline__icontains=discipline)
    if working_code:
        qs = qs.filter(working_code__icontains=working_code)
    if direct_labor:
        qs = qs.filter(direct_labor__icontains=direct_labor)
    if global_search:
        qs = qs.filter(
            Q(discipline__icontains=global_search) |
            Q(working_code__icontains=global_search) |
            Q(direct_labor__icontains=global_search) |
            Q(special_tooling__icontains=global_search)
        )

    data = qs.values(
        'item',
        'discipline',
        'working_code',
        'direct_labor',
        'qty_direct_labor',
        'special_tooling',
        'qty'
    )

    df = pd.DataFrame(list(data))

    # Formatação de qty com 2 casas decimais
    for col in ['qty_direct_labor', 'qty']:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: '{:.2f}'.format(x) if pd.notnull(x) else '')

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=toolsbase_export.xlsx'
    df.to_excel(response, index=False)
    return response


def export_jobcard_excel(request):
    search_number = request.GET.get('search_number', '').strip()
    search_discipline = request.GET.get('search_discipline', '').strip()
    search_prepared_by = request.GET.get('search_prepared_by', '').strip()
    search_location = request.GET.get('search_location', '').strip()
    search_status = request.GET.get('search_status', '').strip()
    global_search = request.GET.get('global_search', '').strip()

    qs = JobCard.objects.all()

    if search_number:
        qs = qs.filter(job_card_number__icontains=search_number)
    if search_discipline:
        qs = qs.filter(discipline__icontains=search_discipline)
    if search_prepared_by:
        qs = qs.filter(prepared_by__icontains=search_prepared_by)
    if search_location:
        qs = qs.filter(location__icontains=search_location)
    if search_status:
        qs = qs.filter(jobcard_status__icontains=search_status)
    if global_search:
        qs = qs.filter(
            Q(job_card_number__icontains=global_search) |
            Q(discipline__icontains=global_search) |
            Q(prepared_by__icontains=global_search) |
            Q(location__icontains=global_search) |
            Q(jobcard_status__icontains=global_search)
        )

    # Campos do seu model, exatamente como estão
    data = qs.values(
        'item',
        'seq_number',
        'discipline',
        'discipline_code',
        'location',
        'level',
        'activity_id',
        'start',
        'finish',
        'system',
        'subsystem',
        'workpack_number',
        'working_code',
        'tag',
        'working_code_description',
        'job_card_number',
        'rev',
        'jobcard_status',
        'job_card_description',
        'completed',
        'total_weight',
        'unit',
        'total_duration_hs',
        'indice_kpi',
        'total_man_hours',
        'prepared_by',
        'date_prepared',
        'approved_br',
        'date_approved',
        'hot_work_required',
        'status',
        'comments',
        'last_modified_by',
        'last_modified_at',
        'offshore_field_check',
        'checked_preliminary_by',
        'checked_preliminary_at'
    )

    df = pd.DataFrame(list(data))

    # Função para checar se é float, mas exporta como está se não for
    def format_float(val):
        try:
            return '{:.2f}'.format(float(val))
        except (TypeError, ValueError):
            return val  # retorna o valor original sem mexer

    # Formata datas (deixa em branco se nulo)
    for col in ['start', 'finish', 'date_prepared', 'date_approved', 'checked_preliminary_at', 'last_modified_at']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d %H:%M:%S').replace('NaT', '')

    # Só tenta formatar como número se realmente for possível, senão mantém string original
    for col in ['total_weight', 'total_duration_hs', 'indice_kpi', 'total_man_hours']:
        if col in df.columns:
            df[col] = df[col].apply(format_float)

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=jobcards_export.xlsx'
    df.to_excel(response, index=False)
    return response

# --------- AREA REPORTS --------------- #


def jobcards_tam(request):
    qs = JobCard.objects.filter(comments__icontains='TAM')

    search = request.GET.get('search', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    items_per_page = request.GET.get('items_per_page', '10')

    if search:
        qs = qs.filter(
            Q(discipline__icontains=search) |
            Q(job_card_number__icontains=search) |
            Q(prepared_by__icontains=search)
        )

    if start_date:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        qs = qs.filter(start__gte=start_date_obj)

    if end_date:
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        qs = qs.filter(start__lte=end_date_obj)

    try:
        items_per_page = int(items_per_page)
    except (ValueError, TypeError):
        items_per_page = 10

    paginator = Paginator(qs, items_per_page)
    page = request.GET.get('page')
    try:
        jobcards_page = paginator.page(page)
    except PageNotAnInteger:
        jobcards_page = paginator.page(1)
    except EmptyPage:
        jobcards_page = paginator.page(paginator.num_pages)

    backups_dir = os.path.join(settings.BASE_DIR, 'jobcard_backups')
    available_pdfs = {f for f in os.listdir(backups_dir) if f.endswith('.pdf')}

    context = {
        'jobcards': jobcards_page,
        'search': search,
        'start_date': start_date,
        'end_date': end_date,
        'items_per_page': items_per_page,
        'paginator': paginator,
        'available_pdfs': available_pdfs,
        'tam_only': True,
        'request': request,  # Para usar request.GET.* no template
    }
    return render(request, 'sistema/jobcards.html', context)



# --------- AVANÇAR JOBCARDS --------------- #


def jobcard_progress(request):
    return render(request, 'sistema/avancar/jobcard_progress.html')


@require_GET
def api_jobcard_detail(request, jobcard_number):
    try:
        job = JobCard.objects.get(job_card_number=jobcard_number)
    except JobCard.DoesNotExist:
        return JsonResponse({'error': 'JobCard not found'}, status=404)

    data = {
        'job_card_number': job.job_card_number,
        'discipline': job.discipline,
        'location': job.location,
        'tag': job.tag,
        'jobcard_status': job.jobcard_status,
        'completed': job.completed,
        'prepared_by': job.prepared_by,
        'date_prepared': job.date_prepared.strftime('%Y-%m-%d') if job.date_prepared else '',
        'prepared_by': job.prepared_by,
        'working_code_description': job.working_code_description,
    }
    return JsonResponse(data)


@require_POST
def api_jobcard_advance(request, jobcard_number):
    try:
        job = JobCard.objects.get(job_card_number=jobcard_number)
    except JobCard.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'JobCard not found'}, status=404)
    job.completed = 'YES'
    job.save()
    return JsonResponse({'success': True})


# --------- AREA DE IMPEDIMENTOS --------------- #


def create_impediment(request):
    error = None
    jobcard = None
    jobcard_number = ''

    if request.method == 'POST':
        form = ImpedimentsForm(request.POST)
        jobcard_number = request.POST.get('jobcard_number', '').strip()
        if not JobCard.objects.filter(job_card_number=jobcard_number).exists():
            error = f"JobCard '{jobcard_number}' does not exist."
        if form.is_valid() and not error:
            impediment = form.save(commit=False)
            impediment.created_by = request.user.username  # <--- SÓ O NOME
            impediment.save()
            return render(request, 'sistema/impediments/impediments.html', {'form': ImpedimentsForm(), 'success': True})

    else:
        form = ImpedimentsForm()

    return render(request, 'sistema/impediments/impediments.html', {
        'form': form,
        'jobcard_number': jobcard_number,
        'jobcard': jobcard,
        'error': error,
    })
        
def impediments_list(request):
    search = request.GET.get('search', '')
    items_per_page = int(request.GET.get('items_per_page', 10))
    page = request.GET.get('page', 1)

    qs = Impediments.objects.all().order_by('-created_at')

    if search:
        qs = qs.filter(
            Q(jobcard_number__icontains=search) |
            Q(notes__icontains=search) |
            Q(other__icontains=search)
        )

    paginator = Paginator(qs, items_per_page)
    impediments = paginator.get_page(page)
    
    items_options = [5, 10, 20, 50, 100]

    context = {
        'impediments': impediments,
        'search': search,
        'items_per_page': items_per_page,
        'paginator': paginator,
        'items_options': items_options,
    }
    return render(request, 'sistema/impediments/impediments_list.html', context)

@csrf_exempt
def impediment_update(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        try:
            imped = Impediments.objects.get(id=id)
        except Impediments.DoesNotExist:
            return JsonResponse({'success': False, 'msg': 'Not found'})

        imped.jobcard_number = request.POST.get('jobcard_number', imped.jobcard_number)
        imped.scaffold      = request.POST.get('scaffold') == 'true'
        imped.material      = request.POST.get('material') == 'true'
        imped.engineering   = request.POST.get('engineering') == 'true'
        imped.other         = request.POST.get('other', '')
        imped.origin_shell  = request.POST.get('origin_shell') == 'true'
        imped.origin_utc    = request.POST.get('origin_utc') == 'true'
        imped.notes         = request.POST.get('notes', '')
        imped.mainpower = request.POST.get('mainpower') == 'true'
        imped.tools = request.POST.get('tools') == 'true'
        imped.access = request.POST.get('access') == 'true'
        imped.pwt = request.POST.get('pwt') == 'true'
        imped.save()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})

@csrf_exempt
def impediment_delete(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        try:
            imped = Impediments.objects.get(id=id)
            imped.delete()
            return JsonResponse({'success': True})
        except Impediments.DoesNotExist:
            return JsonResponse({'success': False, 'msg': 'Not found'})
    return JsonResponse({'success': False})

# --------- AREA DE PMTO --------------- #

@login_required(login_url='login')
def pmto_list(request):
    pmto_items = PMTOBase.objects.all()
    return render(request, 'sistema/pmto/pmto_list.html', {'pmto_items': pmto_items})

@login_required(login_url='login')
@csrf_exempt
def import_pmto(request):
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            file = request.FILES['file']
            df = pd.read_excel(file) if file.name.endswith('.xlsx') else pd.read_csv(file)
            force = request.POST.get('force_update') == 'true'

            # Coleta todos os PMTOCODE do arquivo
            pmto_codes = set(df['PMTOCODE'])
            # Checa se já existem no banco
            existing_codes = list(PMTOBase.objects.filter(pmto_code__in=pmto_codes).values_list('pmto_code', flat=True))

            # Se existe e não está forçando update, retorna para pedir confirmação
            if existing_codes and not force:
                return JsonResponse({
                    'status': 'need_confirm',
                    'duplicated_codes': list(existing_codes)
                })

            # Agora, valida DESCRIPTION+MATERIAL para cada linha
            for _, row in df.iterrows():
                other_pmto = PMTOBase.objects.filter(
                    description=row['DESCRITIVO'],
                    material=row['MATERIAL']
                ).exclude(pmto_code=row['PMTOCODE'])
                if other_pmto.exists():
                    return JsonResponse({
                        'status': 'descmat_duplicate',
                        'detail': (
                            f"<b>DESCRIPTION:</b> {row['DESCRITIVO']}<br>"
                            f"<b>MATERIAL:</b> {row['MATERIAL']}<br>"
                            f"Already registered for <b>PMTO CODE:</b> {other_pmto.first().pmto_code}"
                        )
                    })

            # Agora, faz o import normal (update_or_create)
            for _, row in df.iterrows():
                PMTOBase.objects.update_or_create(
                    pmto_code=row['PMTOCODE'],
                    defaults={
                        'description': row['DESCRITIVO'],
                        'material': row['MATERIAL'],
                        'qty': row['QTY'],
                        'weight': row['WEIGHT'],
                        'unit': row['unit']
                    }
                )

            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'invalid', 'message': 'No file uploaded'})

@login_required(login_url='login')
def export_pmto_excel(request):
    # CAPTURA OS FILTROS ENVIADOS PELO FRONT
    pmto_code = request.GET.get('pmto_code', '')
    description = request.GET.get('description', '')
    material = request.GET.get('material', '')
    global_search = request.GET.get('search', '')

    # INICIA A QUERY
    qs = PMTOBase.objects.all()

    # FILTROS POR CAMPO INDIVIDUAL
    if pmto_code:
        qs = qs.filter(pmto_code__icontains=pmto_code)
    if description:
        qs = qs.filter(description__icontains=description)
    if material:
        qs = qs.filter(material__icontains=material)

    # FILTRO GLOBAL (BUSCA EM VÁRIOS CAMPOS)
    if global_search:
        qs = qs.filter(
            Q(pmto_code__icontains=global_search) |
            Q(description__icontains=global_search) |
            Q(material__icontains=global_search) |
            Q(unit__icontains=global_search)
        )

    # MONTA O DATAFRAME COM OS CAMPOS QUE QUER EXPORTAR
    data = qs.values(
        'pmto_code',
        'description',
        'material',
        'qty',
        'weight',
        'unit'
    )

    df = pd.DataFrame(list(data))

    # FORMATAÇÃO NUMÉRICA: duas casas, vírgula
    for col in ['qty', 'weight']:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: '{:.2f}'.format(x).replace('.', ',') if pd.notnull(x) else '')

    # GERA O EXCEL PARA DOWNLOAD
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=pmto_base_export.xlsx'
    df.to_excel(response, index=False)
    return response

# --------- AREA DE MATERIAL REQUISITION --------------- #

@login_required(login_url='login')
def mr_list(request):
    mr_items = MRBase.objects.all()
    return render(request, 'sistema/materialRequest/mr_list.html', {'mr_items': mr_items})

@login_required(login_url='login')
@csrf_exempt
def import_mr(request):
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            file = request.FILES['file']
            df = pd.read_excel(file) if file.name.endswith('.xlsx') else pd.read_csv(file)
            mr_rev_set = set(zip(df['MR_NUMBER'], df['REV']))

            # Verifica se é preview ou overwrite:
            force = request.POST.get('force_overwrite') == 'true'

            # Verifica duplicidade
            existentes = [
                (mr, rev) for (mr, rev) in mr_rev_set
                if MRBase.objects.filter(mr_number=mr, rev=rev).exists()
            ]

            if existentes and not force:
                # Retorna para o frontend a lista de MR+REV que já existem
                return JsonResponse({'status': 'need_confirm', 'mr_rev': existentes})

            # Agora pode apagar e inserir
            for mr_number, rev in mr_rev_set:
                MRBase.objects.filter(mr_number=mr_number, rev=rev).delete()

            for _, row in df.iterrows():
                MRBase.objects.create(
                    mr_number=row['MR_NUMBER'],
                    pmto_code=row['PMTOCODE'],
                    type_items=row['TYPE ITEMS'],
                    basic_material=row['BASIC MATERIAL'],
                    description=row['DESCRIPTION'],
                    nps1=row['NPS 1'],
                    length_ft_inch=row["LENGTH (FT' INCH\")"],
                    thk_mm=row['THK (mm)'],
                    pid=row['P&ID'],
                    line_number=row['LINE Nº'],
                    qty=row['QTY'],
                    unit=row['UNIT'],
                    design_pressure_bar=row['DESIGN PRESSURE (Bar)'],
                    design_temperature_c=row['DESIGN TEMPERATURE (ºC)'],
                    service=row['SERVICE'],
                    spec=row['SPEC'],
                    proposer_sap_code=row['PROPOSER CODE (SAP CODE)'],
                    rev=row['REV'],
                    notes=row.get('NOTES', '')
                )
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'invalid', 'message': 'No file uploaded'})

@login_required(login_url='login')
def export_mr_excel(request):
    mr_number = request.GET.get('mr_number', '')
    pmto_code = request.GET.get('pmto_code', '')
    basic_material = request.GET.get('basic_material', '')
    global_search = request.GET.get('search', '')

    qs = MRBase.objects.all()

    if mr_number:
        qs = qs.filter(mr_number__icontains=mr_number)
    if pmto_code:
        qs = qs.filter(pmto_code__icontains=pmto_code)
    if basic_material:
        qs = qs.filter(basic_material__icontains=basic_material)

    if global_search:
        qs = qs.filter(
            Q(mr_number__icontains=global_search) |
            Q(pmto_code__icontains=global_search) |
            Q(basic_material__icontains=global_search) |
            Q(description__icontains=global_search)
        )

    data = qs.values(
        'mr_number', 'pmto_code', 'type_items', 'basic_material', 'description',
        'nps1', 'length_ft_inch', 'thk_mm', 'pid', 'line_number',
        'qty', 'unit', 'design_pressure_bar', 'design_temperature_c',
        'service', 'spec', 'proposer_sap_code', 'rev', 'notes'
    )

    df = pd.DataFrame(list(data))
    if 'qty' in df.columns:
        df['qty'] = df['qty'].apply(lambda x: '{:.2f}'.format(x) if pd.notnull(x) else '')

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=mr_base_export.xlsx'
    df.to_excel(response, index=False)
    return response


# --------- AREA DE PROCUREMENT BASE --------------- #

@login_required(login_url='login')
def procurement_list(request):
    procurement_items = ProcurementBase.objects.all()
    return render(request, 'sistema/ProcurementBase/procurement_list.html', {'procurement_items': procurement_items})

@login_required(login_url='login')
@csrf_exempt
@login_required(login_url='login')
@csrf_exempt
def import_procurement(request):
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            file = request.FILES['file']
            df = pd.read_excel(file) if file.name.endswith('.xlsx') else pd.read_csv(file)

            # 1. Apaga toda a base antes de importar (base absoluta)
            ProcurementBase.objects.all().delete()

            # 2. Insere todos os registros da planilha, seguindo as regras de unicidade
            for _, row in df.iterrows():
                ProcurementBase.objects.create(
                    po_number=row['PO Number'],
                    po_status=row.get("Status", ""),
                    po_date=row.get("PO Date"),
                    vendor=row.get("Vendor", ""),
                    expected_delivery_date=row.get("Expected Delivery Date"),
                    mr_number=row.get("MR Number", ""),
                    mr_rev=row.get("MR Rev", ""),
                    qty_mr=row.get("Qty MR"),
                    qty_mr_unit=row.get("Qty MR [UNIT]", ""),
                    item_type=row.get("Item Type", ""),
                    discipline=row.get("Discipline", ""),
                    tam_2026=row.get("TAM 2026", ""),
                    pmto_code=row.get("PMTO CODE", ""),
                    tag=row.get("TAG", ""),
                    detailed_description=row.get("Detailed Description", ""),
                    qty_purchased=row.get("Qty Purchased"),
                    qty_purchased_unit=row.get("Qty Purchased [UNIT]", ""),
                    qty_received=row.get("Qty Received", 0)
                )
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'invalid', 'message': 'No file uploaded'})

@login_required(login_url='login')
def export_procurement_excel(request):
    po_number = request.GET.get('po_number', '')
    pmto_code = request.GET.get('pmto_code', '')
    mr_number = request.GET.get('mr_number', '')

    qs = ProcurementBase.objects.all()
    if po_number:
        qs = qs.filter(po_number__icontains=po_number)
    if pmto_code:
        qs = qs.filter(pmto_code__icontains=pmto_code)
    if mr_number:
        qs = qs.filter(mr_number__icontains=mr_number)

    data = qs.values(
        'po_number',
        'po_status',
        'po_date',
        'vendor',
        'expected_delivery_date',
        'mr_number',
        'mr_rev',
        'qty_mr',
        'qty_mr_unit',
        'item_type',
        'discipline',
        'tam_2026',
        'pmto_code',
        'tag',
        'detailed_description',
        'qty_purchased',
        'qty_purchased_unit'
    )

    df = pd.DataFrame(list(data))

    # Formatação de QTYs (duas casas, vírgula)
    for col in ['qty_mr', 'qty_purchased']:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: '{:.2f}'.format(x).replace('.', ',') if pd.notnull(x) else '')

    # Formatação de data para exportação
    for col in ['po_date', 'expected_delivery_date']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col]).dt.strftime('%Y-%m-%d')

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=procurement_base_export.xlsx'
    df.to_excel(response, index=False)
    return response


# AREA PARA NOTIFICAÇÕES DO USUARIO

from jobcards.models import EngineeringBase, DocumentoRevisaoAlterada

@login_required(login_url='login')
def api_revisoes_ultimas(request):
    limit = int(request.GET.get('limit', 30))
    codigos_validos = EngineeringBase.objects.values_list('document', flat=True)
    qs = DocumentoRevisaoAlterada.objects.filter(
        codigo__in=codigos_validos
    ).order_by('-data_mudanca')

    notificacoes_lidas_data = request.session.get('notificacoes_lidas_data')
    if notificacoes_lidas_data:
        # Mostra só as posteriores (novas)
        qs = qs.filter(data_mudanca__gt=notificacoes_lidas_data)

    revisoes = qs[:limit] if limit > 0 else qs
    data = [
        {
            "codigo": rev.codigo,
            "nome_projeto": rev.nome_projeto,
            "revisao_anterior": rev.revisao_anterior,
            "revisao_nova": rev.revisao_nova,
            "data_mudanca": rev.data_mudanca.strftime("%d/%m/%Y %H:%M"),
            "data_mudanca_db": rev.data_mudanca.strftime("%Y-%m-%d %H:%M:%S"),
        }
        for rev in revisoes
    ]
    return JsonResponse({"revisoes": data})

@csrf_exempt
@login_required
@require_POST
def api_notificacoes_lidas(request):
    import json
    data = json.loads(request.body)
    ultima_data = data.get('ultima_data')
    if ultima_data:
        request.session['notificacoes_lidas_data'] = ultima_data
        return JsonResponse({'ok': True})
    return JsonResponse({'ok': False}, status=400)

# ALLOCAÇÃO DE MAO DE OBRA

from django.db import transaction
from django.http import JsonResponse, HttpRequest


def _taskbase_description(jobcard_id: str, task_order: int) -> str:
    """
    Retorna a descrição *atual* da TaskBase para o working_code do JobCard + order.
    Se não houver TaskBase correspondente, devolve string vazia.
    """
    job = JobCard.objects.only("working_code").get(job_card_number=jobcard_id)
    tb = (
        TaskBase.objects
        .filter(working_code=job.working_code, order=task_order)
        .only("typical_task")
        .first()
    )
    return tb.typical_task if tb else ""


@transaction.atomic
def save_allocation(request: HttpRequest, jobcard_id: str, task_order: int):
    """
    Salva a alocação de manpowers da task e SINCRONIZA a descrição do AllocatedTask
    com a descrição *atual* da TaskBase (reflete renomes/revisões feitos pelo usuário).
    """
    if request.method != "POST":
        return JsonResponse({'error': 'Invalid method'}, status=400)

    data = json.loads(request.body or "{}")

    # 1) Limpa manpowers antigos dessa task
    AllocatedManpower.objects.filter(
        jobcard_number=jobcard_id,
        task_order=task_order
    ).delete()

    # 2) Recria manpowers e calcula totais (fallbacks seguros)
    total_hours = Decimal("0")
    max_hours = Decimal("0")

    for mp in data.get('manpowers', []):
        # Parse seguro para Decimal
        qty = Decimal(str(mp.get("qty", "0") or "0"))
        hours = Decimal(str(mp.get("hours", "0") or "0"))
        total_hours += qty * hours
        if hours > max_hours:
            max_hours = hours

        # OBS: Se seu AllocatedManpower exigir 'discipline' e 'working_code',
        # adicione abaixo esses campos a partir do JobCard ou do frontend.
        AllocatedManpower.objects.update_or_create(
            jobcard_number=jobcard_id,
            task_order=task_order,
            direct_labor=mp.get("direct_labor", "").strip(),
            defaults={
                "qty": qty,
                "hours": hours,
            }
        )

    # 3) Totais vindos do frontend (mantém sua lógica atual)
    max_hh = Decimal(str(data.get('max_hh', max_hours)))
    total_hh = Decimal(str(data.get('total_hh', total_hours)))
    percent = float(data.get('percent', 0) or 0)
    not_applicable = bool(data.get('not_applicable', False))

    # 4) SINCRONIZA a descrição com a TaskBase SEMPRE (update or create)
    #    - Se o usuário renomear a típica na TaskBase, aqui já reflete.
    description = _taskbase_description(jobcard_id, task_order)

    # Se quiser permitir que o frontend sobrescreva em casos específicos:
    # description = (data.get("description") or description).strip()

    AllocatedTask.objects.update_or_create(
        jobcard_number=jobcard_id,
        task_order=task_order,
        defaults={
            "description": description,          # <- atualiza sempre
            "max_hours": float(max_hh),
            "total_hours": float(total_hh),
            "percent": percent,
            "not_applicable": not_applicable,
        }
    )

    return JsonResponse({'success': True})

# --------- AREA DE SUPRIMENTOS --------------- #

def po_tracking(request):
    statuses = [
        "Pending MR",
        "Ordered",
        "In Production",
        "In Transit",
        "Delivered",
        "Ready for Inspection",
        "Ready for Receipt at Warehouse",
        "On Hold",
        "Cancelled",
    ]


    # Um dict status -> lista de POs
    KANBAN_LIMIT = 100
    kanban = {
        status: ProcurementBase.objects.filter(po_status=status)[:KANBAN_LIMIT]
        for status in statuses
    }
    return render(request, 'sistema/procurement/po_tracking.html', {"kanban": kanban, "statuses": statuses})

@csrf_exempt
def po_tracking_update_status(request):
    if request.method == "POST":
        data = json.loads(request.body)
        po = ProcurementBase.objects.get(id=data['id'])
        po.po_status = data['status']
        po.save()
        return JsonResponse({"success": True})
    return JsonResponse({"success": False}, status=400)

def po_tracking_detail(request, po_id):
    po = ProcurementBase.objects.get(id=po_id)
    # Renderize um pedaço de HTML com detalhes do PO (pode usar um template parcial)
    return render(request, "sistema/procurement/partials/po_tracking_detail.html", {"po": po})

def po_tracking_search(request):
    status = request.GET.get('status')
    query = request.GET.get('q', '').strip()

    qs = ProcurementBase.objects.filter(po_status=status)
    if query:
        qs = qs.filter(
            Q(po_number__icontains=query) |
            Q(pmto_code__icontains=query) |
            Q(tag__icontains=query) |
            Q(mr_number__icontains=query) |
            Q(mr_rev__icontains=query) |
            Q(vendor__icontains=query) |
            Q(detailed_description__icontains=query)
        )
    # Exibe todos os resultados do filtro!
    cards_html = render_to_string(
        "sistema/procurement/partials/kanban_cards.html",
        {"pos": qs}
    )
    return JsonResponse({'html': cards_html})


# --------- CONTROLE DE ESTOQUE --------------- #

def warehouse_kanban(request):
    # Mostra cada item de ProcurementBase como um card
    ready_qs = ProcurementBase.objects.filter(po_status="Ready for Receipt at Warehouse", qty_purchased__gt=F('qty_received'))

    received_qs = ProcurementBase.objects.filter(qty_received__gte=F('qty_purchased'))

    kanban = {
        "Ready for Receipt at Warehouse": ready_qs,
        "Received at Warehouse": received_qs,
    }

    # Stocks para detalhamento (WarehouseStock)
    stocks = WarehouseStock.objects.all()

    return render(request, 'sistema/warehouse/warehouse_kanban.html', {
        'statuses': ["Ready for Receipt at Warehouse", "Received at Warehouse"],
        'kanban': kanban,
        'stocks': stocks,
    })
    
@csrf_exempt
def update_warehouse_status(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        po = get_object_or_404(ProcurementBase, id=data['id'])
        po.po_status = data['status']
        po.save()
        return JsonResponse({'result': 'ok'})
    return JsonResponse({'result': 'error'}, status=400)

def warehouse_detail(request, pk):
    po = get_object_or_404(ProcurementBase, id=pk)
    ws = WarehouseStock.objects.filter(po_number=po.po_number, pmto_code=po.pmto_code).first()
    return render(request, 'sistema/warehouse/partials/warehouse_detail.html', {
        'po': po,
        'ws': ws,   # pode ser None
    })
    
@csrf_exempt
def warehouse_receive(request, pk):
    if request.method == 'POST':
        data = json.loads(request.body)
        po = get_object_or_404(ProcurementBase, id=pk)
        qty_to_add = Decimal(str(data.get('qty_received', 0)).replace(',', '.'))

        if po.qty_received is None:
            po.qty_received = Decimal('0.0')

        new_qty = po.qty_received + qty_to_add
        if po.qty_purchased is not None and new_qty > po.qty_purchased:
            qty_to_add = po.qty_purchased - po.qty_received

        po.qty_received += qty_to_add

        if po.qty_purchased and po.qty_received >= po.qty_purchased:
            po.po_status = "Received at Warehouse"
        po.save()

        # Corrigido: defaults preenchido corretamente
        ws, created = WarehouseStock.objects.get_or_create(
            po_number=po.po_number,
            pmto_code=po.pmto_code,
            tag=po.tag,
            defaults={
                'item_type': po.item_type,
                'vendor': po.vendor,
                'discipline': po.discipline,
                'detailed_description': po.detailed_description,
                'qty_purchased': po.qty_purchased,
                'qty_purchased_unit': po.qty_purchased_unit,
                'qty_received': po.qty_received,
                'balance_to_receive': (po.qty_purchased or 0) - (po.qty_received or 0),
                'registration_type': data.get('registration_type', 'unit'),
                'received_by': request.user.get_username() if request.user.is_authenticated else '',
                'last_received_at': timezone.now(),
                'notes': data.get('notes', ''),
            }
        )

        if not created:
            ws.qty_received = po.qty_received
            ws.balance_to_receive = (po.qty_purchased or 0) - (po.qty_received or 0)
            ws.last_received_at = timezone.now()
            ws.received_by = request.user.get_username() if request.user.is_authenticated else ''
            ws.notes = data.get('notes', '')
            ws.registration_type = data.get('registration_type', 'unit')
            ws.save()

        return JsonResponse({
            'result': 'ok',
            'new_status': po.po_status,
            'qty_received': str(po.qty_received)
        })
    return JsonResponse({'result': 'error'}, status=400)

def warehouse_receive_form(request, pk):
    """
    Renderiza o formulário de recebimento (partial) para ser carregado no modal.
    """
    po = get_object_or_404(ProcurementBase, id=pk)

    # Corrija aqui! Busque por TODOS os campos identificadores.
    ws = WarehouseStock.objects.filter(
        po_number=po.po_number,
        pmto_code=po.pmto_code,
        tag=po.tag,
        # Adicione os campos extras que garantem unicidade:
        # mr_number e mr_rev (precisa existir no model WarehouseStock! Se não existir, adicione!)
        # Exemplo:
        # mr_number=po.mr_number,
        # mr_rev=po.mr_rev,
    ).order_by('-id').first()

    if not ws:
        ws = WarehouseStock.objects.create(
            po_number=po.po_number,
            pmto_code=po.pmto_code,
            tag=po.tag,
            # mr_number=po.mr_number,
            # mr_rev=po.mr_rev,
            item_type=po.item_type,
            vendor=po.vendor,
            discipline=po.discipline,
            detailed_description=po.detailed_description,
            qty_purchased=po.qty_purchased,
            qty_purchased_unit=po.qty_purchased_unit,
        )

    saldo = (po.qty_purchased or 0) - (po.qty_received or 0)
    
    return render(request, 'sistema/warehouse/partials/warehouse_receive_form.html', {
        'po': po,
        'ws': ws,
        'saldo': saldo,
    })
      
def warehouse_rfid(request):
    qs = WarehouseStock.objects.all().prefetch_related('pieces')
    po = request.GET.get('po')
    pmto = request.GET.get('pmto')
    tag = request.GET.get('tag')
    if po:
        qs = qs.filter(po_number__icontains=po)
    if pmto:
        qs = qs.filter(pmto_code__icontains=pmto)
    if tag:
        qs = qs.filter(tag__icontains=tag)
    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(qs, 50)  # 50 cards por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Adiciona saldo (to_register) em cada stock
    for stock in page_obj:
        total_registered = stock.pieces.aggregate(total=models.Sum('lot_qty'))['total'] or 0
        stock.to_register = float(stock.qty_received or 0) - float(total_registered or 0)
        stock.registered_qty = float(total_registered or 0)

    return render(request, 'sistema/warehouse/warehouse_rfid.html', {
        'warehouse_stocks': page_obj,  # agora itere sobre page_obj
        'page_obj': page_obj,
        'po': po or '',
        'pmto': pmto or '',
        'tag': tag or '',
    })
    
def rfid_modal(request, stock_id):
    stock = get_object_or_404(WarehouseStock, pk=stock_id)
    pieces = stock.pieces.all()
    registered_qty = pieces.aggregate(total=models.Sum('lot_qty'))['total'] or 0
    to_register = float(stock.qty_received or 0) - float(registered_qty or 0)
    return render(request, 'sistema/warehouse/partials/rfid_modal.html', {
        'stock': stock,
        'pieces': pieces,
        'to_register': to_register,
        'registered_qty': registered_qty,
    })
    

@csrf_exempt
@csrf_exempt
def rfid_add(request, stock_id):
    if request.method == "POST":
        stock = get_object_or_404(WarehouseStock, pk=stock_id)
        data = json.loads(request.body)
        rfid_tag = data.get("rfid_tag")
        location = data.get("location") or "Warehouse at Aveon Yard"
        try:
            lot_qty = Decimal(str(data.get("lot_qty", 1)))
        except (TypeError, ValueError, decimal.InvalidOperation):
            return JsonResponse({'result': 'error', 'msg': 'Invalid quantity!'}, status=400)
        received_by = request.user.get_username() if request.user.is_authenticated else ""

        # Do not allow registration if limit has been reached
        registered = stock.pieces.count()
        if registered >= stock.qty_received:
            return JsonResponse({'result': 'error', 'msg': 'RFID registration limit for this item has been reached!'})

        # Prevent duplicate RFID
        if WarehousePiece.objects.filter(rfid_tag=rfid_tag).exists():
            return JsonResponse({'result': 'error', 'msg': 'RFID already exists.'})

        # Optional: If using lot_qty > 1, ensure total does not exceed qty_received
        total_qty = stock.pieces.aggregate(total=models.Sum('lot_qty'))['total'] or Decimal('0')
        if (total_qty + lot_qty) > Decimal(str(stock.qty_received)):
            return JsonResponse({'result': 'error', 'msg': 'The sum of the batches exceeds the received quantity!'})

        WarehousePiece.objects.create(
            stock=stock,
            rfid_tag=rfid_tag,
            lot_qty=lot_qty,
            location=location,
            received_by=received_by,
        )
        return JsonResponse({'result': 'ok'})
    return JsonResponse({'result': 'error', 'msg': 'Invalid request'}, status=400)


# --------- IMPORTAÇÕES BANCOS --------------- #

LOGISTIC_AREAS = [
    'warehouse_aveon',
    'dock_aveon',
    'flotel',
    'laydown_area',
]

def warehouse_logistics(request):
    logistics = {
        area: WarehousePiece.objects.filter(location=area)
        for area in LOGISTIC_AREAS
    }
    return render(request, "sistema/warehouse/warehouse_logistics.html", {
        "areas": LOGISTIC_AREAS,
        "logistics": logistics,
    })
    
@csrf_exempt
def warehouse_logistics_update_area(request):
    if request.method == "POST":
        import json
        data = json.loads(request.body)
        piece = WarehousePiece.objects.get(id=data['id'])
        piece.location = data['area']
        piece.save()
        return JsonResponse({'result': 'ok'})
    return JsonResponse({'result': 'error'}, status=400)


def warehouse_logistics_search(request):
    area = request.GET.get('area')
    query = request.GET.get('q', '').strip()
    qs = WarehousePiece.objects.filter(location=area)
    if query:
        qs = qs.filter(
            Q(rfid_tag__icontains=query) |
            Q(stock__po_number__icontains=query) |
            Q(stock__pmto_code__icontains=query) |
            Q(stock__tag__icontains=query) |
            Q(stock__vendor__icontains=query) |
            Q(stock__detailed_description__icontains=query)
        )
    cards_html = render_to_string(
        "sistema/warehouse/partials/warehouse_logistics_cards.html",
        {"pieces": qs}
    )
    return JsonResponse({'html': cards_html})

def export_warehouse_pieces_excel(request):
    # Filtra tudo (ou adicione filtros, se quiser depois)
    qs = WarehousePiece.objects.select_related('stock').all()

    data = []
    for piece in qs:
        data.append({
            "RFID Tag": piece.rfid_tag,
            "Lot Qty": piece.lot_qty,
            "PO Number": piece.stock.po_number,
            "Vendor": piece.stock.vendor,
            "PMTO Code": piece.stock.pmto_code,
            "Tag": piece.stock.tag,
            "Description": piece.stock.detailed_description,
            "Qty Purchased": piece.stock.qty_purchased,
            "Qty Received": piece.stock.qty_received,
            "Location": dict(WarehousePiece.LOCATION_CHOICES).get(piece.location, piece.location),
            "Created At": piece.created_at.strftime('%Y-%m-%d %H:%M'),
            "Notes": piece.notes,
        })

    df = pd.DataFrame(data)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=warehouse_pieces.xlsx'
    df.to_excel(response, index=False)
    return response

# --------- JOBCARD PLANNING --------------- #

def jobcards_planning_list(request):
    # Busca apenas JobCards com status específico
    jobcards = JobCard.objects.filter(jobcard_status='PRELIMINARY JOBCARD CHECKED')

    # Filtro simples por GET (adapte se quiser)
    search = request.GET.get('search', '').strip()
    if search:
        jobcards = jobcards.filter(
            Q(job_card_number__icontains=search) |
            Q(discipline__icontains=search) |
            Q(working_code__icontains=search) |
            Q(workpack_number__icontains=search)
        )

    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(jobcards, int(request.GET.get('items_per_page', 10)))
    page = request.GET.get('page')
    jobcards_page = paginator.get_page(page)

    # Monta dados para cada jobcard
    jobcards_data = []
    for job in jobcards_page:
        # Materiais
        allocated_materials = AllocatedMaterial.objects.filter(jobcard_number=job.job_card_number)
        materials_status = []
        all_materials_available = True

        for mat in allocated_materials:
            pieces = WarehousePiece.objects.filter(stock__pmto_code=mat.pmto_code)
            available = pieces.exists()
            materials_status.append({
                "description": mat.description,
                "pmto_code": mat.pmto_code,
                "qty": mat.qty,
                "available": available,
            })
            if not available:
                all_materials_available = False

        # Documentos
        allocated_docs = AllocatedEngineering.objects.filter(jobcard_number=job.job_card_number)
        docs_status = []
        all_docs_available = True

        # for doc in allocated_docs:
        #     doc_ok = DocumentoControle.objects.filter(
        #         codigo=doc.document,
        #         revisao='AFC'
        #     ).exists()
        #     docs_status.append({
        #         "document": doc.document,
        #         "rev": doc.rev,
        #         "available": doc_ok,
        #     })
        #     if not doc_ok:
        #         all_docs_available = False

        for doc in allocated_docs:
            doc_ok = doc.status == 'AFC'
            docs_status.append({
                "document": doc.document,
                "rev": doc.rev,
                "available": doc_ok,
            })
            if not doc_ok:
                all_docs_available = False

        can_advance = all_materials_available and all_docs_available

        jobcards_data.append({
            "job": job,
            "start": job.start,
            "finish": job.finish,
            "total_materials": allocated_materials.count(),
            "materials_status": materials_status,
            "all_materials_available": all_materials_available,
            "total_docs": allocated_docs.count(),
            "docs_status": docs_status,
            "all_docs_available": all_docs_available,
            "can_advance": can_advance,
        })

    context = {
        "jobcards_data": jobcards_data,
        "jobcards": jobcards_page,  # Para a paginação padrão
        "search": search,
        "items_per_page": request.GET.get('items_per_page', 10),
    }

    return render(request, 'sistema/jobcards_planning_list/jobcards_planning_list.html', context)

def change_jobcard_status(request, jobcard_id):
    if request.method == "POST":
        job = get_object_or_404(JobCard, job_card_number=jobcard_id)
        # (Você pode personalizar o novo status)
        job.jobcard_status = "PLANNING JOBCARD CHECKED"
        job.save()
        messages.success(request, "Status updated!")
    return redirect('jobcards_planning_list')

# --------- API PARA APLICATIVO --------------- #


@api_view(["GET"])
#@authentication_classes([JWTAuthentication])
#@permission_classes([IsAuthenticated])
def api_jobcard_list(request):
    qs = JobCard.objects.all().order_by('job_card_number')  # ajuste se quiser
    return Response(JobCardSerializer(qs, many=True).data)


@api_view(["GET"])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def api_jobcard_detail(request, jobcard_number: str):
    """
    Returns JobCard JSON by its number (requires JWT).
    """
    try:
        jc = JobCard.objects.get(job_card_number=jobcard_number)
    except JobCard.DoesNotExist:
        return Response({"detail": "JobCard not found."}, status=status.HTTP_404_NOT_FOUND)

    return Response(JobCardSerializer(jc).data, status=status.HTTP_200_OK)


@api_view(["POST"])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def api_jobcard_advance(request, jobcard_number: str):
    """
    Advances/updates JobCard status without CSRF (JWT only).
    Optional payload: {"status": "COMPLETED"}
    """
    try:
        jc = JobCard.objects.get(job_card_number=jobcard_number)
    except JobCard.DoesNotExist:
        return Response({"detail": "JobCard not found."}, status=status.HTTP_404_NOT_FOUND)

    new_status = request.data.get("status", "COMPLETED")

    # adjust field name if your model uses a different one:
    jc.jobcard_status = new_status
    jc.save(update_fields=["jobcard_status"])

    return Response(
        {
            "ok": True,
            "job_card_number": jc.job_card_number,
            "jobcard_status": jc.jobcard_status,
        },
        status=status.HTTP_200_OK,
    )
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def jobcard_pdfs(request):
    base_url = request.build_absolute_uri(settings.MEDIA_URL)

    pdfs = []
    for jc in JobCard.objects.all():
        name = f"JobCard_{jc.job_card_number}_Rev_{jc.rev}.pdf"
        file_path = os.path.join(settings.MEDIA_ROOT, name)
        if os.path.exists(file_path):
            pdfs.append({
                "name": name,
                "url": base_url + name
            })

    return Response({"pdfs": pdfs})

# ESTOQUE E RFIDs conexões

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def check_rfid(request):
    rfids = request.data.get("rfids", [])
    results = []
    for tag in rfids:
        piece = WarehousePiece.objects.filter(rfid_tag=tag).first()
        if piece:
            results.append({
                "rfid": tag,
                "status": "Found",
                "location": piece.location,
                "stock_id": piece.stock.id,
            })
        else:
            results.append({"rfid": tag, "status": "Not Registered"})
    return Response(results)

def api_rfid_all(request):
    rfids = list(WarehousePiece.objects.values_list('rfid_tag', flat=True))
    return JsonResponse(rfids, safe=False)

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def update_location(request):
    user = request.user
    rfids = request.data.get("rfids", [])
    new_location = request.data.get("location")
    updated = []

    for tag in rfids:
        piece = WarehousePiece.objects.filter(rfid_tag=tag).first()
        if piece:
            piece.location = new_location
            piece.received_by = user.username
            piece.save()
            updated.append(tag)

    return Response({"updated": updated})

# Criar impedimento
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_create_impediment(request):
    serializer = ImpedimentSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save(created_by=request.user.username if request.user else None)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Listar todos impedimentos
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_list_impediments(request):
    impediments = Impediments.objects.all().order_by('-created_at')
    serializer = ImpedimentSerializer(impediments, many=True)
    return Response(serializer.data)


# Detalhar um impedimento específico
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_impediment_detail(request, pk):
    try:
        imp = Impediments.objects.get(pk=pk)
    except Impediments.DoesNotExist:
        return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

    serializer = ImpedimentSerializer(imp)
    return Response(serializer.data)


# Atualizar impedimento
@api_view(['PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
def api_update_impediment(request, pk):
    try:
        imp = Impediments.objects.get(pk=pk)
    except Impediments.DoesNotExist:
        return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

    serializer = ImpedimentSerializer(imp, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Deletar impedimento
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def api_delete_impediment(request, pk):
    try:
        imp = Impediments.objects.get(pk=pk)
    except Impediments.DoesNotExist:
        return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

    imp.delete()
    return Response({'message': 'Deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def api_jobcard_manpowers(request, jobcard_number):
    jobcard = get_object_or_404(JobCard, job_card_number=jobcard_number)

    # Manpowers dessa jobcard
    manpowers = AllocatedManpower.objects.filter(
        jobcard_number=jobcard_number
    ).order_by("task_order")

    serializer = AllocatedManpowerSerializer(manpowers, many=True)

    # Catálogo global de direct_labor (sem repetir)
    # Se preferir, filtre por disciplina do jobcard: .filter(discipline=jobcard.discipline)
    all_labors = list(
        AllocatedManpower.objects
        .exclude(direct_labor__isnull=True)
        .exclude(direct_labor__exact="")
        .values_list("direct_labor", flat=True)
        .distinct()
        .order_by("direct_labor")
    )

    return Response({
        "jobcard": jobcard.job_card_number,
        "discipline": jobcard.discipline,
        "working_code": jobcard.working_code,
        "manpowers": serializer.data,
        "all_labors": all_labors,  # 👈 catálogo global, sem repetir
    })

@api_view(["POST"])
@permission_classes([IsAuthenticated])
@transaction.atomic
def api_dfr_close(request, jobcard_number: str):
    """
    Payload esperado:
    {
      "report_date": "2025-08-20",
      "items": [
        {"task_description":"...", "task_order":1, "direct_labor":"...", "hours": 1.5, "qty":1, "source": "MANUAL"},
        ...
      ],
      "notes": "optional",
      "snapshot": {...}  // optional
    }
    """
    data = request.data or {}
    report_date = data.get("report_date")
    items = data.get("items") or []
    notes = data.get("notes")
    snapshot = data.get("snapshot")

    if not report_date:
        return Response({"detail": "report_date is required."}, status=400)
    if not items:
        return Response({"detail": "items is required and cannot be empty."}, status=400)

    # metadados do JobCard no momento do fechamento
    discipline = None
    working_code = None
    try:
        jc = JobCard.objects.get(job_card_number=jobcard_number)
        discipline = jc.discipline
        working_code = jc.working_code
    except JobCard.DoesNotExist:
        pass

    # gera número do DFR uma vez por relatório
    dfr_number = DailyFieldReport.next_dfr_number()

    # calcula totais
    total_hours = 0.0
    total_lines = 0
    for it in items:
        h = float(it.get("hours") or 0)
        total_hours += h
        total_lines += 1

    total_hours = round(total_hours, 2)

    # cria cada linha (mesma tabela)
    line_seq = 1
    created_rows = []
    for it in items:
        row = DailyFieldReport.objects.create(
            dfr_number=dfr_number,
            line_seq=line_seq,
            jobcard_number=jobcard_number,
            discipline=discipline,
            working_code=working_code,
            report_date=report_date,
            total_hours=total_hours,
            total_lines=total_lines,
            created_by=str(request.user),
            notes=notes,
            snapshot=snapshot,

            task_description=it.get("task_description") or "",
            task_order=it.get("task_order"),
            direct_labor=it.get("direct_labor") or "",
            hours=round(float(it.get("hours") or 0), 2),
            qty=int(it.get("qty") or 1),
            source=(it.get("source") or "MANUAL")[:12],
        )
        created_rows.append(row.id)
        line_seq += 1

    summary = {
        "dfr_number": dfr_number,
        "report_date": report_date,
        "jobcard_number": jobcard_number,
        "total_hours": total_hours,
        "total_lines": total_lines,
    }
    return Response(summary, status=status.HTTP_201_CREATED)

from django.core.files.storage import default_storage, FileSystemStorage
import os, re
import time
from django.utils.text import slugify

def _save_jobcard_image(job, uploaded_file, index):
    field = f'image_{index}'
    base = re.sub(r'[^A-Za-z0-9_.-]', '_', (job.job_card_number or str(job.id)))
    ext = os.path.splitext(uploaded_file.name)[1] or '.png'
    filename = f"{base}_{index}{ext}"
    path = os.path.join('jobcard_images', filename)

    # remover existente
    existing = getattr(job, field)
    try:
        if existing and default_storage.exists(existing.name):
            default_storage.delete(existing.name)
    except Exception:
        pass

    saved_path = default_storage.save(path, uploaded_file)
    setattr(job, field, saved_path)
    return saved_path


@login_required(login_url='login')
@permission_required('jobcards.change_jobcard', raise_exception=True)
def delete_jobcard_image(request):
    """
    POST: {'jobcard_number': 'A01.PS-EL-0001', 'index': 1}
    Remove arquivo do storage e zera campo no banco.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)

    jobcard_number = request.POST.get('jobcard_number')
    try:
        index = int(request.POST.get('index'))
    except Exception:
        return JsonResponse({'success': False, 'error': 'Invalid index'}, status=400)

    job = get_object_or_404(JobCard, job_card_number=jobcard_number)
    field = f'image_{index}'
    f = getattr(job, field, None)
    if f:
        try:
            if hasattr(f, 'name') and default_storage.exists(f.name):
                default_storage.delete(f.name)
        except Exception:
            pass
        setattr(job, field, None)
        job.save(update_fields=[field])
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': 'No image to delete'}, status=404)

ALLOWED_EXT = ".pdf"
MAX_SIZE_MB = 50  # ajuste se quiser

def _is_pdf(file_obj) -> bool:
    """
    Validação robusta: extensão .pdf e header %PDF (magic bytes).
    """
    name_ok = file_obj.name.lower().endswith(ALLOWED_EXT)
    pos = file_obj.tell()
    head = file_obj.read(5)  # b'%PDF-'
    file_obj.seek(pos)
    header_ok = head.startswith(b"%PDF")
    return name_ok and header_ok

@login_required(login_url="login")
@permission_required("jobcards.change_jobcard", raise_exception=True)
def upload_documents(request):
    if request.method == "GET":
        # Renderiza a página de upload
        return render(request, "sistema/upload_documents/upload_documents.html")

    if request.method == "POST":
        files = request.FILES.getlist("documents")
        if not files:
            messages.error(request, "No files selected.")
            return redirect("upload_documents")  # mantenha o nome da sua rota aqui

        # Pasta de destino: MEDIA_ROOT/documents_jobcards/
        rel_dir = os.path.join("documents_jobcards")
        abs_dir = os.path.join(settings.MEDIA_ROOT, rel_dir)
        os.makedirs(abs_dir, exist_ok=True)

        storage = FileSystemStorage(
            location=abs_dir,
            base_url=settings.MEDIA_URL + rel_dir + "/"
        )

        saved, errors = [], []

        for f in files:
            # Tamanho
            if f.size > MAX_SIZE_MB * 1024 * 1024:
                errors.append(f"{f.name}: file too large (> {MAX_SIZE_MB}MB).")
                continue

            # Apenas PDF (ext + magic bytes)
            if not _is_pdf(f):
                errors.append(f"{f.name}: only PDF files are allowed.")
                continue

            # Nome seguro (sem timestamp para permitir overwrite "1:1")
            base, _ = os.path.splitext(f.name)
            safe_name = f"{slugify(base)[:80]}.pdf"

            # OVERWRITE: se existir, apaga antes de salvar
            abs_target = os.path.join(abs_dir, safe_name)
            if os.path.exists(abs_target):
                try:
                    os.remove(abs_target)
                except OSError as e:
                    errors.append(f"{f.name}: failed to overwrite ({e}).")
                    continue

            # Salva (FileSystemStorage não sobrescreve por padrão)
            filename = storage.save(safe_name, f)
            saved.append(storage.url(filename))

        if saved:
            messages.success(request, f"{len(saved)} PDF(s) uploaded successfully (overwritten when same name).")
        if errors:
            messages.warning(request, " | ".join(errors))

        return redirect("upload_documents")

    # Outros métodos → volta pra página
    messages.error(request, "Method not allowed.")
    return redirect("upload_documents")



# --------- AREA DE MODIFICAÇÃO DA JOBCARDS --------------- #

from django.contrib import messages
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required, permission_required

from .forms import JobCardForm, EDITABLE_FIELDS
from .models import JobCard
from django.utils.translation import override

from datetime import datetime
import csv
import io

import logging
from django.urls import reverse
from django.utils import timezone
from urllib.parse import urlencode

def _parse_date_maybe(value):
    """
    Converte strings comuns de data para date (YYYY-MM-DD, DD/MM/YYYY, etc.).
    Retorna None se vazio; retorna string original se formato não reconhecido.
    """
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "date"):  # ex.: Timestamp
        try:
            return value.date()
        except Exception:
            pass
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return text  # deixa como veio; o ModelForm validará se necessário


@login_required(login_url="login")
@permission_required("jobcards.change_jobcard", raise_exception=True)
def modify_jobcard_entry(request):
    """
    Página de entrada: usuário informa o número da JobCard.
    """
    if request.method == "POST":
        number = (request.POST.get("job_card_number") or "").strip().upper()
        if not number:
            messages.error(request, "Informe o número da JobCard.")
        elif not JobCard.objects.filter(job_card_number=number).exists():
            messages.error(request, f"JobCard '{number}' não encontrada.")
        else:
            return redirect("modify_jobcard_edit", jobcard=number)

    recent = JobCard.objects.order_by("-last_modified_at")[:20]
    return render(request, "sistema/modify_jobcard/modify_jobcard_entry.html", {"recent_jobcards": recent})


logger = logging.getLogger(__name__)

@login_required(login_url="login")
@permission_required("jobcards.change_jobcard", raise_exception=True)
@transaction.atomic
def modify_jobcard_edit(request, jobcard):
    """
    Tela de edição via 'Modify': usuário escolhe STATUS nos chips.
    - Salva no banco
    - Redireciona para generate_pdf com src=modify & status=<escolhido> & apply_status=1
      para SOBRESCREVER o status também no generate_pdf.
    """
    obj = get_object_or_404(JobCard, job_card_number=jobcard)

    if request.method == "POST":
        form = JobCardForm(request.POST, request.FILES or None, instance=obj)

        if 'prepared_by' in form.fields:
            form.fields['prepared_by'].required = False

        chosen_status = (
            request.POST.get("jobcard_status")
            or request.POST.get("JOBCARD_STATUS")
            or obj.jobcard_status
        )

        if form.is_valid():
            inst = form.save(commit=False)

            # aplica status escolhido
            inst.jobcard_status   = chosen_status
            inst.last_modified_by = request.user.username
            inst.last_modified_at = timezone.now()

            # se virou PRELIMINARY e nunca carimbou, carimba agora
            if (
                inst.jobcard_status == "PRELIMINARY JOBCARD CHECKED"
                and not inst.checked_preliminary_by
                and not inst.checked_preliminary_at
            ):
                inst.checked_preliminary_by = request.user.username
                inst.checked_preliminary_at = timezone.now()

            inst.save()
            form.save_m2m()

            # >>> redireciona pedindo que o generate_pdf APLIQUE o status escolhido
            from urllib.parse import urlencode
            from django.urls import reverse

            qs = urlencode({
                "src": "modify",
                "status": inst.jobcard_status,
                "apply_status": "1",
            })
            return redirect(f"{reverse('generate_pdf', args=[inst.job_card_number])}?{qs}")

        messages.error(request, "Please correct the errors and try again.")
    else:
        form = JobCardForm(instance=obj)

    return render(
        request,
        "sistema/modify_jobcard/modify_jobcard.html",
        {"form": form, "job": obj},
    )


@login_required(login_url="login")
@permission_required("jobcards.change_jobcard", raise_exception=True)
@transaction.atomic
def modify_jobcard_excel_patch(request, jobcard):
    """
    Patch rápido por Excel (.xlsx) ou CSV:
    - Lê a PRIMEIRA linha de dados.
    - Cabeçalhos devem casar com nomes dos campos (ex.: prepared_by, jobcard_status, comments...).
    - Atualiza apenas campos em EDITABLE_FIELDS.
    """
    job = get_object_or_404(JobCard, job_card_number=jobcard)

    if request.method != "POST" or "file" not in request.FILES:
        messages.error(request, "No file received.")
        return redirect("modify_jobcard_edit", jobcard=jobcard)

    f = request.FILES["file"]
    filename = f.name.lower()

    # Monta dict {campo: valor} da primeira linha
    row_dict = {}

    try:
        if filename.endswith(".csv"):
            # CSV em UTF-8 (aceita BOM)
            data = io.TextIOWrapper(f.file, encoding="utf-8-sig")
            reader = csv.DictReader(data)
            first = next(reader, None)
            if not first:
                messages.error(request, "Empty CSV.")
                return redirect("modify_jobcard_edit", jobcard=jobcard)
            row_dict = {k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in first.items() if k}

        elif filename.endswith(".xlsx"):
            try:
                import openpyxl
            except ImportError:
                messages.error(request, "openpyxl not installed. Install with: pip install openpyxl")
                return redirect("modify_jobcard_edit", jobcard=jobcard)

            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                headers.append((str(cell.value).strip() if cell.value is not None else ""))

            # pega apenas a primeira linha de dados
            data_row = next(ws.iter_rows(min_row=2, values_only=True), None)
            if not data_row:
                messages.error(request, "Excel has no data rows.")
                return redirect("modify_jobcard_edit", jobcard=jobcard)

            for i, head in enumerate(headers):
                if head:
                    row_dict[head] = data_row[i] if i < len(data_row) else None

        else:
            messages.error(request, "Unsupported file. Use .xlsx or .csv.")
            return redirect("modify_jobcard_edit", jobcard=jobcard)

        # Aplica patch somente nos campos permitidos
        changed = []
        for field, value in row_dict.items():
            if field in EDITABLE_FIELDS:
                if field in {"start", "finish", "date_prepared", "date_approved"}:
                    value = _parse_date_maybe(value)
                setattr(job, field, value)
                changed.append(field)

        if changed:
            job.last_modified_by = request.user.username
            job.save()
            messages.success(request, f"Patched fields: {', '.join(changed)}")
        else:
            messages.info(request, "No editable fields found in file.")

    except Exception as e:
        messages.error(request, f"Error reading file: {e}")

    return redirect("modify_jobcard_edit", jobcard=jobcard)


# --- CONFIGURAÇÃO DO IMPORT/MODIFY ------------------------------------------
# (deixe isso perto dos seus imports)

# Campos que NÃO podem ser atualizados (a sua coluna vermelha)
FORBIDDEN_FIELDS = {
    "seq_number","discipline","discipline_code","location","level",
    "total_duration_hs","indice_kpi","total_man_hours","prepared_by",
    "date_prepared","approved_br","date_approved","status",
    "checked_preliminary_by","checked_preliminary_at",
}

# Campos que PODEM aparecer no template e ser atualizados (a sua coluna da direita)
ALLOWED_FOR_MODIFY = [
    "activity_id","start","finish","system","subsystem","workpack_number",
    "working_code","tag","working_code_description",
    "rev","jobcard_status","job_card_description","completed",
    "total_weight","unit","hot_work_required","comments",
]

# Campos que vamos ignorar mesmo que apareçam
IGNORED_ON_IMPORT = {"image_1","image_2","image_3","image_4","last_modified_at"}

DATE_FIELDS = {"start", "finish", "date_prepared", "date_approved", "checked_preliminary_at"}

def _parse_date_any(val):
    """
    Converte diferentes formatos em date.
    Retorna None para vazio/NULL/NONE ou se não reconhecer.
    """
    if val is None:
        return None

    # already date/datetime
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()

    s = str(val).strip()
    if s == "":
        return None

    # palavras-chave
    lowered = s.lower()
    if lowered in {"today", "hoje", "now", "agora"}:
        return date.today()

    # ISO com 'T'
    if "t" in lowered:
        try:
            return datetime.fromisoformat(s.replace("Z", "")).date()
        except Exception:
            pass

    # serial numérico do Excel (ex.: 45123, 45123.0)
    # Excel (Windows) usa base 1899-12-30 por conta do bug do ano 1900
    try:
        # aceita "45123", "45123.0", "45123,0"
        s_num = s.replace(",", ".")
        serial = float(s_num)
        if 1 <= serial < 100000:  # guarda-chuva razoável
            base = date(1899, 12, 30)
            return base + timedelta(days=int(round(serial)))
    except Exception:
        pass

    # formatos comuns
    fmts = (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
        "%Y-%m-%d %H:%M:%S",
        "%d.%m.%Y",
        "%m/%d/%Y",            # se alguém mandar em US
        "%Y-%m-%d %H:%M",      # com hora sem segundos
    )
    for fmt in fmts:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue

    # última tentativa: só data da parte antes do espaço
    if " " in s:
        try:
            return datetime.strptime(s.split(" ")[0], "%Y-%m-%d").date()
        except Exception:
            pass

    # não reconhecido → None (e você decide se ignora ou trata como erro)
    return None

def _yn(val, default="NO"):
    if val in (None, ""):
        return default
    s = str(val).strip().upper()
    return "YES" if s in {"YES","Y","SIM","TRUE","1"} else "NO"

def _hotwork(val):
    if val in (None, ""):
        return "NO"
    s = str(val).strip().upper()
    if s in {"YES","Y","SIM","TRUE","1"}: return "YES"
    if s in {"NO","N","NAO","FALSE","0"}: return "NO"
    return "NO"

# === NOVA VIEW: baixar template com todos os campos ==========================
@login_required(login_url='login')
@permission_required('jobcards.change_jobcard', raise_exception=True)
def download_jobcard_modify_template(request):
    """
    Gera .xlsx para o import_jobcard_modify:
    - Cabeçalho = job_card_number + SOMENTE os campos de ALLOWED_FOR_MODIFY.
    """
    import io, pandas as pd

    headers = ["job_card_number"] + ALLOWED_FOR_MODIFY
    example_row = ["A01.PS-EL-0001"] + [""] * len(ALLOWED_FOR_MODIFY)

    df = pd.DataFrame([example_row], columns=headers)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="ModifyJobCards", index=False)

        info = pd.DataFrame({
            "Dica": [
                "Preencha apenas as colunas que quer atualizar.",
                'Deixe a célula vazia para NÃO alterar o valor atual.',
                'Escreva "NULL" para LIMPAR (salvar None) um campo.',
                "Não altere o job_card_number — ele precisa existir no banco.",
            ]
        })
        info.to_excel(writer, sheet_name="Instrucoes", index=False)

    buf.seek(0)
    return HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="jobcard_modify_template.xlsx"'},
    )



# === MODIFICAÇÃO DAS JOBCARDS  VIA EXCEL ====================

@login_required(login_url='login')
@permission_required('jobcards.change_jobcard', raise_exception=True)
def import_jobcard_modify(request):
    """
    Atualiza JobCards existentes a partir de .xlsx/.csv.
    - NÃO cria novas.
    - Só aplica colunas em ALLOWED_FOR_MODIFY (vazias são ignoradas; "NULL"/"NONE" zera).
    """
    if request.method != "POST":
        return JsonResponse({'status': 'error', 'message': 'Invalid request.'}, status=405)

    f = request.FILES.get('file')
    if not f:
        return JsonResponse({'status': 'error', 'message': 'No file uploaded.'}, status=400)

    import pandas as pd, re
    try:
        df = (pd.read_csv(f, dtype=str) if f.name.lower().endswith('.csv')
              else pd.read_excel(f, dtype=str)).fillna("")
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Could not read file: {e}'}, status=400)

    # normaliza nomes de coluna
    df.columns = [c.strip() for c in df.columns]
    if 'job_card_number' not in df.columns:
        return JsonResponse({'status': 'error', 'message': "Missing column: 'job_card_number'."}, status=400)

    # só vamos considerar as colunas realmente permitidas
    cols_to_apply = [c for c in df.columns if c in ALLOWED_FOR_MODIFY]
    if not cols_to_apply:
        return JsonResponse({'status': 'error', 'message': 'No updatable columns found.'}, status=400)

    # valida duplicadas e formato do número
    pattern = r"^[A-Z0-9]{3}\.[A-Z0-9]{2}-[A-Z0-9]{2}-\d{4}$"
    dupes = df['job_card_number'].str.upper().duplicated(keep=False)
    if dupes.any():
        dups_list = df.loc[dupes, 'job_card_number'].str.upper().unique().tolist()
        return JsonResponse({'status': 'error', 'message': f"Duplicated job_card_number in file: {', '.join(dups_list)}"}, status=400)

    payload = []
    invalid = []
    for _, row in df.iterrows():
        jobnum = str(row['job_card_number']).strip().upper()
        if not jobnum:
            continue
        if not re.match(pattern, jobnum):
            invalid.append(jobnum); continue

        entry = {'job_card_number': jobnum}
        for k in cols_to_apply:
            entry[k] = row.get(k, "")
        payload.append(entry)

    if invalid:
        return JsonResponse({'status': 'error', 'message': f"Invalid JobCard format: {', '.join(invalid)}"}, status=400)
    if not payload:
        return JsonResponse({'status': 'error', 'message': 'No useful rows to import.'}, status=400)

    # existem todas?
    jobnums = [r['job_card_number'] for r in payload]
    existing = set(JobCard.objects.filter(job_card_number__in=jobnums).values_list('job_card_number', flat=True))
    missing = sorted(set(jobnums) - existing)
    if missing:
        return JsonResponse({'status': 'not_found', 'missing': missing, 'message': 'Some JobCards do not exist.'}, status=400)

    # aplica
    from django.db import transaction
    updated = 0
    who = request.user.username or request.user.email or "importer"

    with transaction.atomic():
        for row in payload:
            jobnum = row['job_card_number']

            # monta updates (regras: vazio = não mexe; "NULL"/"NONE" = None)
            updates = {}
            for k, v in row.items():
                if k == 'job_card_number' or k in IGNORED_ON_IMPORT:
                    continue

                val = (v or "").strip()
                if val == "":
                    continue
                if val.upper() in {"NULL", "NONE"}:
                    updates[k] = None
                    continue

                # normalizações específicas
                if k in DATE_FIELDS:
                    updates[k] = _parse_date_any(val)
                elif k == "completed":
                    updates[k] = _yn(val, default="NO")
                elif k == "hot_work_required":
                    updates[k] = _hotwork(val)
                elif k == "jobcard_status":
                    updates[k] = val.upper()
                else:
                    updates[k] = val

            if not updates:
                continue

            updates['last_modified_by'] = who
            JobCard.objects.filter(job_card_number=jobnum).update(**updates)
            updated += 1

    return JsonResponse({'status': 'ok', 'updated': updated})




#====================== RENDERIZA NOVAMENTE OS PDFS ==========================#

# views.py (imports extras)
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from django.core.cache import cache
from django.views.decorators.http import require_POST, require_GET
from django.template.loader import render_to_string
from django.core.files.storage import default_storage
import tempfile, os
from django.views.decorators.http import require_POST
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import os, time
from concurrent.futures import ThreadPoolExecutor, as_completed
from django.core.cache import cache
from django.db import close_old_connections   # 👈 AQUI!

def _render_jobcard_pdf_to_disk(job_card_number: str):
    """
    Renderiza o PDF da JobCard atualizada e grava em:
      jobcard_backups/JobCard_{num}_Rev_{rev}.pdf
    Retorna (True, None) em sucesso, ou (False, 'erro...') em falha.
    """
    try:
        job = JobCard.objects.get(job_card_number=job_card_number)
    except JobCard.DoesNotExist:
        return (False, "JobCard not found")

    # temporários a serem limpos no finally (logo + watermark + header/footer HTMLs)
    _tmp_paths = []

    # helpers locais
    def _file_url(p: str) -> str:
        return f'file:///{p.replace("\\", "/")}'

    def _to_local_path(field_or_path):
        base_dir = str(getattr(settings, "BASE_DIR", ""))
        if hasattr(field_or_path, "path") and field_or_path.path:
            return field_or_path.path if os.path.exists(field_or_path.path) else None
        if hasattr(field_or_path, "name") and field_or_path.name:
            try:
                if default_storage.exists(field_or_path.name):
                    return default_storage.path(field_or_path.name)
            except Exception:
                return None
        if isinstance(field_or_path, str) and field_or_path:
            return field_or_path if os.path.isabs(field_or_path) else os.path.join(base_dir, field_or_path)
        return None

    def _copy_to_temp(src_path: str) -> tuple[str, str]:
        ext = os.path.splitext(src_path)[1] or ".png"
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
        with open(src_path, "rb") as s, open(tmp.name, "wb") as d:
            d.write(s.read())
        return _file_url(tmp.name), tmp.name

    try:
        area_info = Area.objects.filter(area_code=job.location).first() if job.location else None

        # ----- Código de barras -----
        barcode_folder = os.path.join(settings.BASE_DIR, 'static', 'barcodes')
        os.makedirs(barcode_folder, exist_ok=True)
        barcode_filename = f'{job.job_card_number}.png'
        barcode_path = os.path.join(barcode_folder, barcode_filename)
        if not os.path.exists(barcode_path):
            CODE128 = barcode.get_barcode_class('code128')
            code128 = CODE128(job.job_card_number, writer=ImageWriter())
            with open(barcode_path, 'wb') as bf:
                code128.write(bf, options={'write_text': False})
        barcode_url = _file_url(barcode_path)

        # ----- Dados alocados -----
        allocated_manpowers   = AllocatedManpower.objects.filter(jobcard_number=job_card_number).order_by('task_order')
        allocated_materials   = AllocatedMaterial.objects.filter(jobcard_number=job.job_card_number)
        allocated_tools       = AllocatedTool.objects.filter(jobcard_number=job_card_number)
        allocated_tasks       = AllocatedTask.objects.filter(jobcard_number=job_card_number).order_by('task_order')
        allocated_engineering = AllocatedEngineering.objects.filter(jobcard_number=job.job_card_number)

        # ----- LOGO (image_path) com cache-buster -----
        base_dir = str(getattr(settings, "BASE_DIR", ""))  # fallback
        logo_src = (
            _to_local_path(getattr(job, "image_1", None)) or
            _to_local_path(getattr(settings, "PDF_HEADER_LOGO", None)) or
            os.path.join(base_dir, "static", "assets", "img", "3.jpg")
        )
        image_url, tmp_logo = _copy_to_temp(logo_src)
        _tmp_paths.append(tmp_logo)

        # ----- WATERMARK (watermark_url) com cache-buster -----
        # Se você usa sempre um arquivo fixo (ex.: static/assets/img/utc_vazio.png),
        # copiar para temp garante refresh ao substituir mantendo o nome.
        wm_src = os.path.join(settings.BASE_DIR, 'static', 'assets', 'img', 'utc_vazio.png')
        if os.path.exists(wm_src):
            watermark_url, tmp_wm = _copy_to_temp(wm_src)
            _tmp_paths.append(tmp_wm)
        else:
            watermark_url = None  # mantém compatível se não existir

        # ----- Limpa referências quebradas (image_1..image_4 do corpo) -----
        updated_fields = []
        for i in range(1, 5):
            field = f'image_{i}'
            f = getattr(job, field, None)
            if f:
                exists = False
                try:
                    if hasattr(f, 'path') and os.path.exists(f.path):
                        exists = True
                    elif hasattr(f, 'name') and default_storage.exists(f.name):
                        exists = True
                except Exception:
                    exists = False
                if not exists:
                    setattr(job, field, None)
                    updated_fields.append(field)
        if updated_fields:
            job.save(update_fields=updated_fields)

        # ----- Monta file:/// para imagens do corpo -----
        image_files = {}
        for i in range(1, 5):
            field = f'image_{i}'
            f = getattr(job, field, None)
            if f:
                try:
                    if hasattr(f, 'path') and os.path.exists(f.path):
                        image_files[field] = _file_url(f.path)
                    elif hasattr(f, 'name') and default_storage.exists(f.name):
                        storage_path = default_storage.path(f.name)
                        image_files[field] = _file_url(storage_path)
                    elif hasattr(f, 'url') and getattr(f, 'url', None):
                        image_files[field] = f.url
                    else:
                        image_files[field] = None
                except Exception:
                    image_files[field] = None
            else:
                image_files[field] = None

        half = len(allocated_tools) // 2 if allocated_tools else 0

        # ----- wkhtmltopdf -----
        def _find_wkhtmltopdf():
            bin_path = getattr(settings, "WKHTMLTOPDF_BIN", None)
            if bin_path and os.path.exists(bin_path):
                return bin_path
            candidate = os.path.join(str(getattr(settings, "BASE_DIR", "")), "wkhtmltopdf", "bin", "wkhtmltopdf.exe")
            if os.path.exists(candidate):
                return candidate
            for c in [
                r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe",
                r"C:\Program Files (x86)\wkhtmltopdf\bin\wkhtmltopdf.exe",
                "/usr/bin/wkhtmltopdf",
                "/usr/local/bin/wkhtmltopdf",
            ]:
                if os.path.exists(c):
                    return c
            from shutil import which
            return which("wkhtmltopdf")

        path_wkhtmltopdf = _find_wkhtmltopdf()
        local_config = None
        try:
            if path_wkhtmltopdf and os.path.exists(path_wkhtmltopdf):
                local_config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
        except Exception:
            local_config = None

        # ----- Context (mantém SUAS variáveis image_path e watermark_url) -----
        context = {
            'job': job,
            'allocated_manpowers': allocated_manpowers,
            'allocated_materials': allocated_materials,
            'allocated_tools': allocated_tools,
            'allocated_tools_left': allocated_tools[:half],
            'allocated_tools_right': allocated_tools[half:],
            'allocated_tasks': allocated_tasks,
            'allocated_engineerings': allocated_engineering,
            'image_path': image_url,         # usado no <img src="{{ image_path }}">
            'barcode_image': barcode_url,
            'area_info': area_info,
            'image_files': image_files,
            'watermark_url': watermark_url,  # usado no <img src="{{ watermark_url }}">
        }

        # ----- Render HTMLs -----
        html_string         = render_to_string('sistema/jobcard_pdf.html', context)
        header_html_string  = render_to_string('sistema/header.html', context)
        footer_html_string  = render_to_string('sistema/footer.html', context)

        header_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.html')
        footer_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.html')
        header_temp.write(header_html_string.encode('utf-8'))
        footer_temp.write(footer_html_string.encode('utf-8'))
        header_temp.close(); footer_temp.close()
        _tmp_paths += [header_temp.name, footer_temp.name]

        # ----- Geração do PDF -----
        try:
            pdf_bytes = pdfkit.from_string(
                html_string,
                False,
                configuration=local_config,
                options={
                    'enable-local-file-access': '',  # precisa ser "truthy"
                    'margin-top': '35mm',
                    'margin-bottom': '30mm',
                    'header-html': _file_url(header_temp.name),
                    'footer-html': _file_url(footer_temp.name),
                    'header-spacing': '5',
                    'footer-spacing': '5',
                    'quiet': '',
                }
            )
        except Exception as e:
            return (False, f"wkhtmltopdf error: {e}")

        # ----- Persistência do PDF -----
        backups_dir = os.path.join(settings.BASE_DIR, 'jobcard_backups')
        os.makedirs(backups_dir, exist_ok=True)
        rev_tag = (job.rev or "R00").replace("/", "_").replace("\\", "_")
        backup_filename = f'JobCard_{job.job_card_number}_Rev_{rev_tag}.pdf'
        backup_path     = os.path.join(backups_dir, backup_filename)
        with open(backup_path, 'wb') as f:
            f.write(pdf_bytes)

        return (True, None)

    except Exception as e:
        return (False, str(e))
    finally:
        # limpa todos os temporários criados (logo, watermark, header/footer)
        for p in _tmp_paths:
            try:
                os.unlink(p)
            except Exception:
                pass


# views.py — inicia uma corrida e devolve run_id + total
@login_required(login_url='login')
@permission_required('jobcards.change_jobcard', raise_exception=True)
@require_POST
def api_pdf_run_start(request):
    ids = list(
        JobCard.objects
        .exclude(jobcard_status__in=['NO CHECKED','NOT CHECKED'])
        .order_by('job_card_number')
        .values_list('job_card_number', flat=True)
    )
    run_id = uuid.uuid4().hex[:8].upper()

    # zera contadores na cache (use Redis/Memcached p/ produção)
    ttl = 2 * 60 * 60  # 2h
    cache.set(f'pdf:{run_id}:total', len(ids), ttl)
    cache.set(f'pdf:{run_id}:done',  0, ttl)
    cache.set(f'pdf:{run_id}:ok',    0, ttl)
    cache.set(f'pdf:{run_id}:err',   0, ttl)

    return JsonResponse({'status':'ok','run_id':run_id,'total':len(ids)})

# views.py — consulta progresso da corrida
@login_required(login_url='login')
@permission_required('jobcards.change_jobcard', raise_exception=True)
@require_GET
def api_pdf_run_progress(request):
    run_id = request.GET.get('run_id','')
    total = cache.get(f'pdf:{run_id}:total', 0) or 0
    done  = cache.get(f'pdf:{run_id}:done',  0) or 0
    ok    = cache.get(f'pdf:{run_id}:ok',    0) or 0
    err   = cache.get(f'pdf:{run_id}:err',   0) or 0
    return JsonResponse({'status':'ok','run_id':run_id,'total':total,'done':done,'ok':ok,'err':err})


# Limite global de renders simultâneos por **processo** (ajuste via env em produção)
PDF_SEMAPHORE = threading.BoundedSemaphore(int(os.environ.get('PDF_MAX_CONCURRENCY', '4')))


# views.py — processa um lote (batch) com paralelismo no servidor
@login_required(login_url='login')
@permission_required('jobcards.change_jobcard', raise_exception=True)
@require_POST
def api_regenerate_jobcards_pdfs(request):
    """
    POST: limit, offset, parallel, (opcional) run_id
    - Nunca estoura 500 por erro de item: qualquer falha entra em failed[].
    - Fecha conexões de BD nas threads e limita paralelismo de render.
    """
    # ---- parâmetros seguros ----
    try:
        limit    = max(0, int(request.POST.get('limit', 200)))
        offset   = max(0, int(request.POST.get('offset', 0)))
        parallel = max(1, min(int(request.POST.get('parallel', 3)), 8))  # bound de segurança
    except ValueError:
        return JsonResponse({'status':'error','message':'invalid params'}, status=400)

    run_id = request.POST.get('run_id') or None

    # ---- queryset base ----
    base_qs = (JobCard.objects
               .exclude(jobcard_status__in=['NO CHECKED','NOT CHECKED'])
               .order_by('job_card_number'))
    total = base_qs.count()
    if limit == 0:
        # preflight: só devolve total
        return JsonResponse({'status':'ok','total': total})

    # fatia do lote
    batch = list(base_qs.values_list('job_card_number', flat=True)[offset:offset+limit])

    processed_ids, failed = [], []

    # ---- worker que trata UM item com isolamento ----
    def _one(jobnum: str):
        """
        - Garante que a thread tem/fecha sua própria conexão de BD.
        - Aplica teto global de concorrência de PDF (semaforo).
        - Converte qualquer exceção em registro de falha, sem derrubar a view.
        """
        # cada thread usa sua conexão de forma segura
        close_old_connections()
        PDF_SEMAPHORE.acquire()
        try:
            try:
                ok, err = _render_jobcard_pdf_to_disk(jobnum)
            except Exception as e:
                # qualquer erro no render vira falha do item
                ok, err = False, f"{type(e).__name__}: {e}"
        finally:
            PDF_SEMAPHORE.release()
            close_old_connections()

        # contadores de progresso por run_id (se houver)
        if run_id:
            try:
                cache.incr(f'pdf:{run_id}:done', 1)
                cache.incr(f'pdf:{run_id}:ok' if ok else f'pdf:{run_id}:err', 1)
            except Exception:
                # cache não pode derrubar a requisição
                pass

        return (ok, err, jobnum)

    # ---- executa o lote com paralelismo bounded ----
    max_workers = max(1, min(parallel, 8))
    try:
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            futures = [ex.submit(_one, j) for j in batch]
            for fut in as_completed(futures):
                try:
                    ok, err, jobnum = fut.result()
                except Exception as e:
                    # se a própria future der exceção, captura aqui
                    ok, err, jobnum = False, f"ThreadError: {type(e).__name__}: {e}", "?"
                if ok:
                    processed_ids.append(jobnum)
                else:
                    failed.append({'job': jobnum, 'error': str(err)})
    except Exception as e:
        # se algo MUITO atípico escapar (ex.: erro de serialização), ainda assim responde 200 com contexto
        failed.append({'job': '*batch*', 'error': f'BatchError: {type(e).__name__}: {e}'})

    return JsonResponse({
        'status': 'ok',
        'total': total,
        'processed_batch': len(processed_ids) + len(failed),
        'processed_ids': processed_ids,
        'failed': failed
    })




@login_required(login_url='login')
def ajax_tools_for_manpowers(request):
    working_code = request.GET.get('working_code')
    direct_labors = request.GET.getlist('direct_labors[]')

    # Filtra só ferramentas do working code E dos manpowers alocados
    tools = ToolsBase.objects.filter(
        working_code=working_code,
        direct_labor__in=direct_labors
    ).order_by('item')

    data = [
        {
            'item': t.item,
            'discipline': t.discipline,
            'working_code': t.working_code,
            'direct_labor': t.direct_labor,
            'qty_direct_labor': t.qty_direct_labor,
            'special_tooling': t.special_tooling,
            'qty': t.qty,
        }
        for t in tools
    ]
    return JsonResponse({'tools': data})

from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse
import subprocess, sys

@staff_member_required
def pdf_diag(request):
    info = {}
    try:
        who = subprocess.run(["whoami"], capture_output=True, text=True)
        info["whoami"] = (who.stdout or who.stderr).strip()
    except Exception as e:
        info["whoami"] = f"ERR: {e}"

    info["python"] = sys.executable
    info["WKHTMLTOPDF_BIN"] = settings.WKHTMLTOPDF_BIN
    info["TEMP"] = os.environ.get("TEMP")
    info["TMP"] = os.environ.get("TMP")

    checks = []
    for p in [settings.PDF_TEMP_DIR, settings.LOGS_DIR, settings.BARCODES_DIR, settings.BACKUPS_DIR]:
        try:
            tf = tempfile.NamedTemporaryFile(dir=p, delete=True)
            tf.write(b"ok"); tf.flush(); tf.close()
            checks.append(f"OK write: {p}")
        except Exception as e:
            checks.append(f"FAIL write: {p} -> {e}")
    info["write_checks"] = checks

    try:
        r = subprocess.run([settings.WKHTMLTOPDF_BIN, "--version"], capture_output=True, text=True, timeout=10)
        info["wkhtmltopdf_version"] = (r.stdout or r.stderr).strip()
    except Exception as e:
        info["wkhtmltopdf_version"] = f"ERROR: {e}"

    return JsonResponse(info)


# ===================== AJAX PARA TOOLS BASED ON MANPOWERS ========================== #