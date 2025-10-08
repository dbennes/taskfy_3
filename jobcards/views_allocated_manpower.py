# jobcards/views_allocated_manpower.py
from __future__ import annotations

import csv
import io
import os
import difflib
from collections import defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Dict, List, Set, Tuple

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.db import models, transaction
from django.db.models import Q, Sum, Max, F, ExpressionWrapper, FloatField
from django.db.models.functions import Coalesce
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.shortcuts import render, redirect
from django.utils.safestring import mark_safe
from django.db.models import Q, Sum, Max, F, ExpressionWrapper, DecimalField, Value
from django.db.models.functions import Coalesce, Cast

# MODELS
from .models import (
    AllocatedManpower,
    AllocatedTask,
    JobCard,         # fields: job_card_number, total_man_hours, total_duration_hs, discipline, working_code
    ManpowerBase,    # lista “válida” de direct_labor
    TaskBase,        # catálogo de tasks por (working_code, order)
)

# XLSX opcional
try:
    import openpyxl  # type: ignore
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False


# =========================
# Helpers numéricos e I/O
# =========================
def _to_decimal(value) -> Decimal:
    """Converte string -> Decimal aceitando vírgula/ponto. Vazio -> 0."""
    if value is None:
        return Decimal("0")
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return Decimal("0")
    s = str(value).strip()
    if not s:
        return Decimal("0")
    # normaliza "1.234,56" -> "1234.56"
    s = s.replace(".", "").replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0")


def _q2(val: float | Decimal) -> Decimal:
    """Arredonda para 2 casas com HALF_UP em Decimal."""
    try:
        d = val if isinstance(val, Decimal) else Decimal(str(val or 0))
    except Exception:
        d = Decimal("0")
    return d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _coerce_for_field(val: Decimal, model_field: models.Field):
    """
    Converte Decimal(2 casas) para o tipo do campo alvo do JobCard.
    - DecimalField -> Decimal
    - Float/Integer -> float
    - Char/Text -> string "0.00"
    """
    if isinstance(model_field, models.DecimalField):
        return val
    if isinstance(model_field, (models.FloatField, models.IntegerField)):
        return float(val)
    return f"{val:.2f}"


def _read_rows_from_upload(fobj, filename) -> List[dict]:
    """
    Lê CSV/XLSX e retorna lista de dicts com colunas:
    jobcard_number, discipline, working_code, direct_labor, task_order, qty, hours
    - CSV: auto-detecta delimitador (; , \t |) e remove linha "sep=;"
    - XLSX: usa openpyxl (se instalado)
    """
    rows: List[dict] = []
    wanted = [
        "jobcard_number",
        "discipline",
        "working_code",
        "direct_labor",
        "task_order",
        "qty",
        "hours",
    ]
    ext = os.path.splitext(filename or "")[1].lower()

    if ext == ".csv":
        data = fobj.read()
        # decodifica com BOM (utf-8-sig) e fallback Latin-1
        try:
            text = data.decode("utf-8-sig")
        except Exception:
            text = data.decode("latin-1")

        # remove primeira linha "sep=;"
        lines = text.splitlines()
        csv_sep_hint = None
        if lines and lines[0].lower().startswith("sep="):
            try:
                csv_sep_hint = lines[0].split("=", 1)[1].strip() or ";"
            except Exception:
                csv_sep_hint = ";"
            text = "\n".join(lines[1:])  # drop da linha sep=

        # autodetecta delimitador
        sample = text[:4096]
        if csv_sep_hint in (";", ",", "\t", "|"):
            delimiter = csv_sep_hint
        else:
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
                delimiter = dialect.delimiter
            except Exception:
                delimiter = ";" if text.count(";") >= text.count(",") else ","

        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)

        # normaliza header -> lower/strip
        src_headers = [h.strip() for h in (reader.fieldnames or [])]
        header_map = {h.lower(): h for h in src_headers}

        for raw in reader:
            obj = {}
            for k in wanted:
                src = header_map.get(k)
                val = raw.get(src) if src else None
                obj[k] = (val or "").strip() if isinstance(val, str) else val
            rows.append(obj)

    elif ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        if not _HAS_OPENPYXL:
            raise ValueError("To import XLSX, install openpyxl or upload a CSV file.")
        wb = openpyxl.load_workbook(fobj, read_only=True, data_only=True)
        ws = wb.active
        header: List[str] = []
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                header = [(str(c or "")).strip().lower() for c in row]
                first = False
                continue
            raw = dict(zip(header, row))
            obj = {}
            for k in wanted:
                v = raw.get(k)
                obj[k] = (str(v).strip() if isinstance(v, str) else v)
            rows.append(obj)
    else:
        raise ValueError("Unsupported file format. Please upload CSV or XLSX.")

    # Normalização e tipagem
    norm: List[dict] = []
    for r in rows:
        job = (r.get("jobcard_number") or "").strip()
        if not job:
            continue
        # task_order
        t = r.get("task_order")
        try:
            task_order = int(str(t).strip()) if t not in (None, "") else 0
        except Exception:
            task_order = 0
        # qty/hours
        qty = _to_decimal(r.get("qty"))
        hours = _to_decimal(r.get("hours"))
        norm.append(
            {
                "jobcard_number": job,
                "discipline": (r.get("discipline") or "").strip(),
                "working_code": (r.get("working_code") or "").strip(),
                "direct_labor": (r.get("direct_labor") or "").strip(),
                "task_order": task_order,
                "qty": qty,
                "hours": hours,
            }
        )
    return norm


def _manpower_valid_names_set() -> Set[str]:
    """
    Retorna um set (lowercased) com os nomes válidos de direct_labor em ManpowerBase.
    Tenta 'name', depois 'direct_labor', senão pega o primeiro CharField disponível.
    """
    field_name = None
    try:
        ManpowerBase._meta.get_field("name")
        field_name = "name"
    except Exception:
        try:
            ManpowerBase._meta.get_field("direct_labor")
            field_name = "direct_labor"
        except Exception:
            for f in ManpowerBase._meta.get_fields():
                if isinstance(f, models.CharField):
                    field_name = f.name
                    break
    if not field_name:
        return set()

    vals = ManpowerBase.objects.values_list(field_name, flat=True)
    return {(v or "").strip().lower() for v in vals if v}


# =========================
# Regras de negócio (Tasks)
# =========================
def _recalc_task_totals(affected_pairs: Set[Tuple[str, int]]) -> int:
    """
    Para cada (jobcard_number, task_order):
      - total_hours = Σ(qty * hours) dos manpowers da task  ✅
      - max_hours   = MAIOR 'hours' (sem qty) dentre os manpowers da task  ✅
      - description = TaskBase.typical_task (por working_code + order), se houver
      - not_applicable = False (toda task presente no import é aplicável)
      - completed = False (sempre)
      - percent = NÃO calculamos aqui (é por participação no total do JobCard)
    """
    if not affected_pairs:
        return 0

    jobcards = {jc for jc, _ in affected_pairs}
    orders = {o for _, o in affected_pairs}

    # Padroniza tudo para DECIMAL no banco
    dec6 = DecimalField(max_digits=20, decimal_places=6)

    # (qty * hours) com CAST para Decimal e COALESCE para nulos
    qty_dec   = Cast(Coalesce(F("qty"),   Value(0)), dec6)
    hours_dec = Cast(Coalesce(F("hours"), Value(0)), dec6)
    qty_hours_expr = ExpressionWrapper(qty_dec * hours_dec, output_field=dec6)

    # Σ(qty * hours) por (jobcard, task_order)
    totals_qs = (
        AllocatedManpower.objects
        .filter(jobcard_number__in=jobcards, task_order__in=orders)
        .values("jobcard_number", "task_order")
        .annotate(total=Coalesce(Sum(qty_hours_expr, output_field=dec6), Value(0), output_field=dec6))
    )
    total_map = {(r["jobcard_number"], r["task_order"]): float(_q2(r["total"])) for r in totals_qs}

    # max_hours = MAIOR 'hours' (sem qty), também forçando Decimal
    max_qs = (
        AllocatedManpower.objects
        .filter(jobcard_number__in=jobcards, task_order__in=orders)
        .values("jobcard_number", "task_order")
        .annotate(maxh=Coalesce(Max(Cast(F("hours"), dec6), output_field=dec6), Value(0), output_field=dec6))
    )
    max_map = {(r["jobcard_number"], r["task_order"]): float(_q2(r["maxh"])) for r in max_qs}

    # working_code por par (para buscar descrição no TaskBase)
    wk_qs = (
        AllocatedManpower.objects
        .filter(jobcard_number__in=jobcards, task_order__in=orders)
        .values("jobcard_number", "task_order")
        .annotate(working_code=Max("working_code"))
    )
    wk_map = {(r["jobcard_number"], r["task_order"]): (r["working_code"] or "") for r in wk_qs}

    # TaskBase: (working_code, order) -> typical_task
    wks = {wk for wk in wk_map.values() if wk}
    tb_qs = TaskBase.objects.filter(working_code__in=wks, order__in=orders).values("working_code", "order", "typical_task")
    tb_map = {(r["working_code"], r["order"]): (r["typical_task"] or "") for r in tb_qs}

    # Atualiza/Cria AllocatedTask — ORDENADO por (jobcard, task_order)
    existing = {
        (t.jobcard_number, t.task_order): t
        for t in AllocatedTask.objects.filter(jobcard_number__in=jobcards, task_order__in=orders)
    }

    to_update, to_create = [], []
    for key in sorted(affected_pairs, key=lambda x: (x[0], x[1])):
        jc, order = key
        totalh = float(total_map.get(key, 0.0))
        maxh = float(max_map.get(key, 0.0))
        wk = wk_map.get(key, "")
        desc = tb_map.get((wk, order), "")

        if key in existing:
            inst = existing[key]
            fields = []
            if float(inst.total_hours or 0.0) != totalh:
                inst.total_hours = totalh; fields.append("total_hours")
            if float(inst.max_hours or 0.0) != maxh:
                inst.max_hours = maxh; fields.append("max_hours")
            if (not inst.description) and desc:
                inst.description = desc; fields.append("description")
            if inst.not_applicable:
                inst.not_applicable = False; fields.append("not_applicable")
            if inst.completed:
                inst.completed = False; fields.append("completed")
            if fields:
                to_update.append(inst)
        else:
            to_create.append(
                AllocatedTask(
                    jobcard_number=jc,
                    task_order=order,
                    description=desc or "",
                    max_hours=maxh,
                    total_hours=totalh,
                    completed=False,
                    percent=0.0,
                    not_applicable=False,
                )
            )

    if to_create:
        to_create.sort(key=lambda t: (t.jobcard_number, t.task_order))
        AllocatedTask.objects.bulk_create(to_create, batch_size=1000)
    if to_update:
        AllocatedTask.objects.bulk_update(
            to_update,
            ["description", "max_hours", "total_hours", "completed", "not_applicable"],
            batch_size=1000,
        )

    return len(to_create) + len(to_update)



def _recalc_jobcard_percentages(jobcards: Set[str]) -> int:
    if not jobcards:
        return 0

    dec6 = DecimalField(max_digits=20, decimal_places=6)
    totals = (
        AllocatedTask.objects.filter(jobcard_number__in=jobcards)
        .values("jobcard_number")
        .annotate(
            gt=Coalesce(Sum(Cast(F("total_hours"), dec6), output_field=dec6), Value(0), output_field=dec6)
        )
    )
    gt_map = {r["jobcard_number"]: float(r["gt"] or 0.0) for r in totals}

    tasks = list(
        AllocatedTask.objects
        .filter(jobcard_number__in=jobcards)
        .order_by('jobcard_number', 'task_order')
    )
    to_update = []
    for t in tasks:
        grand = gt_map.get(t.jobcard_number, 0.0)
        new_percent = 0.0 if grand <= 0.0 else min(round(float(t.total_hours or 0.0) / grand * 100.0, 2), 100.0)
        if float(t.percent or 0.0) != new_percent:
            t.percent = new_percent
            to_update.append(t)

    if to_update:
        AllocatedTask.objects.bulk_update(to_update, ["percent"], batch_size=1000)
    return len(to_update)

# =========================
# Helpers: PRUNE órfãs (AllocatedTask x TaskBase)
# =========================
def _prune_orphan_allocated_tasks(jobcards: Set[str], job_map: Dict[str, dict]) -> int:
    """
    Remove de AllocatedTask as tasks cujo task_order NÃO existe mais no TaskBase
    para o working_code da JobCard.
    - Se o working_code não tiver nenhuma task no TaskBase, apaga TODAS as tasks da JobCard.
    Retorna quantidade de linhas deletadas.
    """
    if not jobcards:
        return 0

    # working_code por JobCard
    wc_by_jc = {jc: (job_map.get(jc, {}).get("working_code") or "") for jc in jobcards}
    working_codes = {wc for wc in wc_by_jc.values() if wc}

    # Mapa de orders esperados por working_code, conforme TaskBase
    expected_by_wc: Dict[str, Set[int]] = defaultdict(set)
    if working_codes:
        tb_qs = TaskBase.objects.filter(working_code__in=working_codes).values("working_code", "order")
        for r in tb_qs:
            expected_by_wc[r["working_code"]].add(int(r["order"]))

    deleted_total = 0
    # Para cada JobCard importado, apaga o que não estiver no TaskBase
    for jc in sorted(jobcards):
        wc = wc_by_jc.get(jc, "")
        expected_orders = expected_by_wc.get(wc, set())

        if not expected_orders:
            # Não há tasks no TaskBase para este working_code -> apaga todas
            deleted, _ = AllocatedTask.objects.filter(jobcard_number=jc).delete()
            deleted_total += deleted
        else:
            # Apaga apenas as que não pertencem ao TaskBase
            deleted, _ = (
                AllocatedTask.objects
                .filter(jobcard_number=jc)
                .exclude(task_order__in=list(expected_orders))
                .delete()
            )
            deleted_total += deleted

    return deleted_total


def _ensure_missing_tasks_as_na(valid_rows: List[dict], job_map: Dict[str, dict]) -> int:
    """
    replace_jobcard:
      Compara orders importados x orders do TaskBase (pelo working_code da JobCard).
      Faltantes: N/A (percent=0, completed=False, max/total=0). Mantém ordem.
    """
    imported_by_jc: Dict[str, Set[int]] = defaultdict(set)
    for r in valid_rows:
        jc = (r.get("jobcard_number") or "").strip()
        order = int(r.get("task_order") or 0)
        if jc:
            imported_by_jc[jc].add(order)

    if not imported_by_jc:
        return 0

    wc_by_jc = {jc: (job_map.get(jc, {}).get("working_code") or "") for jc in imported_by_jc.keys()}
    working_codes = {wc for wc in wc_by_jc.values() if wc}

    tb_qs = TaskBase.objects.filter(working_code__in=working_codes).values("working_code", "order", "typical_task")
    expected_by_wc: Dict[str, Dict[int, str]] = defaultdict(dict)
    for r in tb_qs:
        expected_by_wc[r["working_code"]][int(r["order"])] = (r["typical_task"] or "")

    existing_tasks = {
        (t.jobcard_number, t.task_order): t
        for t in AllocatedTask.objects.filter(jobcard_number__in=list(imported_by_jc.keys()))
    }

    to_update, to_create = [], []
    for jc in sorted(imported_by_jc.keys()):
        imported_orders = imported_by_jc[jc]
        wc = wc_by_jc.get(jc, "")
        expected_orders_map = expected_by_wc.get(wc, {})
        expected_orders = set(expected_orders_map.keys())
        if not expected_orders:
            continue

        missing_orders = expected_orders - imported_orders
        if not missing_orders:
            continue

        for order in sorted(missing_orders):
            key = (jc, order)
            if key in existing_tasks:
                inst = existing_tasks[key]
                fields = []
                if not inst.not_applicable:
                    inst.not_applicable = True; fields.append("not_applicable")
                if inst.completed:
                    inst.completed = False; fields.append("completed")
                if float(inst.percent or 0.0) != 0.0:
                    inst.percent = 0.0; fields.append("percent")
                if float(inst.max_hours or 0.0) != 0.0:
                    inst.max_hours = 0.0; fields.append("max_hours")
                if float(inst.total_hours or 0.0) != 0.0:
                    inst.total_hours = 0.0; fields.append("total_hours")
                desc = expected_orders_map.get(order, "")
                if (not inst.description) and desc:
                    inst.description = desc; fields.append("description")
                if fields:
                    to_update.append(inst)
            else:
                to_create.append(
                    AllocatedTask(
                        jobcard_number=jc,
                        task_order=order,
                        description=expected_orders_map.get(order, ""),
                        max_hours=0.0,
                        total_hours=0.0,
                        completed=False,
                        percent=0.0,
                        not_applicable=True,
                    )
                )

    if to_create:
        to_create.sort(key=lambda t: (t.jobcard_number, t.task_order))
        AllocatedTask.objects.bulk_create(to_create, batch_size=1000)
    if to_update:
        AllocatedTask.objects.bulk_update(
            to_update,
            ["description", "max_hours", "total_hours", "percent", "completed", "not_applicable"],
            batch_size=1000,
        )
    return len(to_create) + len(to_update)


# =========================
# Sincronismo JobCard ↔ Tasks
# =========================
def _sync_jobcard_from_allocatedtask(jobcards: Set[str]) -> int:
    if not jobcards:
        return 0

    dec6 = DecimalField(max_digits=20, decimal_places=6)
    aggs = (
        AllocatedTask.objects
        .filter(jobcard_number__in=jobcards)
        .values("jobcard_number")
        .annotate(
            sum_total=Coalesce(Sum(Cast(F("total_hours"), dec6), output_field=dec6), Value(0), output_field=dec6),
            sum_dur  = Coalesce(Sum(Cast(F("max_hours"),   dec6), output_field=dec6), Value(0), output_field=dec6),
        )
    )
    by_jc = {r["jobcard_number"]: (_q2(r["sum_total"]), _q2(r["sum_dur"])) for r in aggs}

    f_tmh = JobCard._meta.get_field("total_man_hours")
    f_dur = JobCard._meta.get_field("total_duration_hs")

    to_update = []
    for job in JobCard.objects.filter(job_card_number__in=jobcards):
        tmh, dur = by_jc.get(job.job_card_number, (_q2(0), _q2(0)))
        new_tmh = _coerce_for_field(tmh, f_tmh)
        new_dur = _coerce_for_field(dur, f_dur)

        if (str(job.total_man_hours) != str(new_tmh)) or (str(job.total_duration_hs) != str(new_dur)):
            job.total_man_hours  = new_tmh
            job.total_duration_hs = new_dur
            to_update.append(job)

    if to_update:
        JobCard.objects.bulk_update(to_update, ["total_man_hours", "total_duration_hs"], batch_size=500)
    return len(to_update)



# =========================
# Views: LISTA (server-side)
# =========================
@login_required(login_url="login")
@permission_required("jobcards.change_jobcard", raise_exception=True)
def allocated_manpower_list(request):
    """Renderiza o template. Dados via AJAX (server-side)."""
    return render(request, "sistema/allocated/allocated_manpower_list.html")


@login_required(login_url="login")
@permission_required("jobcards.change_jobcard", raise_exception=True)
def allocated_manpower_table_data(request):
    """
    Endpoint para DataTables (server-side processing).
    Suporta: paginação, busca global, filtros por coluna e ordenação.
    """
    # Parâmetros DataTables
    try:
        draw = int(request.GET.get("draw", "1"))
    except Exception:
        draw = 1
    try:
        start = int(request.GET.get("start", "0"))
    except Exception:
        start = 0
    try:
        length = int(request.GET.get("length", "18"))
    except Exception:
        length = 18

    search_value = (request.GET.get("search[value]") or "").strip()

    # Filtros customizados
    f_job = (request.GET.get("filter_job") or "").strip()
    f_disc = (request.GET.get("filter_disc") or "").strip()
    f_labor = (request.GET.get("filter_labor") or "").strip()

    # Ordenação
    order_col_idx = request.GET.get("order[0][column]", "0")
    order_dir = request.GET.get("order[0][dir]", "asc")
    columns = [
        "jobcard_number",  # 0
        "discipline",      # 1
        "working_code",    # 2
        "direct_labor",    # 3
        "qty",             # 4
        "hours",           # 5
        "task_order",      # 6
    ]
    try:
        order_field = columns[int(order_col_idx)]
    except Exception:
        order_field = "jobcard_number"
    if order_dir == "desc":
        order_field = f"-{order_field}"

    qs = AllocatedManpower.objects.all()

    # Filtros por campo
    if f_job:
        qs = qs.filter(jobcard_number__icontains=f_job)
    if f_disc:
        qs = qs.filter(discipline__icontains=f_disc)
    if f_labor:
        qs = qs.filter(direct_labor__icontains=f_labor)

    # Busca global
    if search_value:
        sv = search_value
        qs = qs.filter(
            Q(jobcard_number__icontains=sv)
            | Q(discipline__icontains=sv)
            | Q(working_code__icontains=sv)
            | Q(direct_labor__icontains=sv)
        )

    records_total = AllocatedManpower.objects.count()
    records_filtered = qs.count()

    # Ordena e pagina retornando somente campos necessários
    qs = (
        qs.order_by(order_field)
        .only(
            "jobcard_number",
            "discipline",
            "working_code",
            "direct_labor",
            "qty",
            "hours",
            "task_order",
        )
        .values_list(
            "jobcard_number",
            "discipline",
            "working_code",
            "direct_labor",
            "qty",
            "hours",
            "task_order",
        )
    )[start : start + length]

    data = []
    for job, disc, wk, labor, qty, hrs, order in qs:
        data.append(
            [
                job or "",
                disc or "",
                wk or "",
                labor or "",
                f"{(qty or 0):.2f}",
                f"{(hrs or 0):.2f}",
                order if order is not None else "",
            ]
        )

    return JsonResponse(
        {
            "draw": draw,
            "recordsTotal": records_total,
            "recordsFiltered": records_filtered,
            "data": data,
        }
    )


# =========================
# Views: IMPORT & TEMPLATE (VERSÃO ATUALIZADA)
# =========================
@login_required(login_url="login")
@permission_required("jobcards.change_jobcard", raise_exception=True)
def import_allocated_manpower(request):
    """
    Importa CSV/XLSX para AllocatedManpower COM REGRAS:
      1) direct_labor deve existir em ManpowerBase
      2) discipline e working_code SEMPRE vêm da JobCard base (ignoramos arquivo)
      3) jobcard_number deve existir em JobCard base

    Modos:
      - merge: upsert por (jobcard_number, task_order, direct_labor)
      - replace_jobcard:
          * apaga todos os AllocatedManpower dos jobcards importados
          * zera horas das AllocatedTask remanescentes
          * 🔥 remove AllocatedTask que NÃO existem mais no TaskBase (prune órfãs)
          * cria/ajusta faltantes como N/A (com base no TaskBase)
          * recalcula percentuais e sincroniza JobCard (Σ totals / Σ max)
    """
    if request.method != "POST":
        return HttpResponseBadRequest("Invalid method.")

    upload = request.FILES.get("file")
    mode = (request.POST.get("mode") or "merge").strip()
    if not upload:
        messages.error(request, "Please upload a CSV or XLSX file.")
        return redirect("allocated_manpower_list")

    try:
        rows = _read_rows_from_upload(upload, upload.name)
    except Exception as e:
        messages.error(request, f"Failed to read file: {e}")
        return redirect("allocated_manpower_list")

    if not rows:
        messages.error(request, "Import cancelled: no valid rows found in file.")
        return redirect("allocated_manpower_list")

    # Pré-carregamentos
    incoming_jobs = {(r.get("jobcard_number") or "").strip() for r in rows if (r.get("jobcard_number") or "").strip()}
    jc_qs = JobCard.objects.filter(job_card_number__in=incoming_jobs).values("job_card_number", "discipline", "working_code")
    job_map = {
        j["job_card_number"]: {"discipline": j["discipline"] or "", "working_code": j["working_code"] or ""}
        for j in jc_qs
    }
    valid_labor = _manpower_valid_names_set()

    # Validação de linhas
    valid_rows: List[dict] = []
    invalid_jobs: List[str] = []
    invalid_labor: List[str] = []

    for r in rows:
        jc = (r.get("jobcard_number") or "").strip()
        if not jc or jc not in job_map:
            invalid_jobs.append(jc or "(empty)")
            continue

        labor = (r.get("direct_labor") or "").strip()
        if not labor or labor.strip().lower() not in valid_labor:
            invalid_labor.append(labor or "(empty)")
            continue

        # Overwrite SEMPRE com a base da JobCard
        r["discipline"] = job_map[jc]["discipline"]
        r["working_code"] = job_map[jc]["working_code"]
        valid_rows.append(r)

    # Abort on any error
    if invalid_jobs or invalid_labor:
        blocks = []
        if invalid_jobs:
            uniq = sorted({x for x in invalid_jobs if x})
            preview = ", ".join(uniq[:20]) + ("…" if len(uniq) > 20 else "")
            blocks.append(f"<li><b>Unknown JobCards</b>: {preview}</li>")

        if invalid_labor:
            uniq = sorted({x for x in invalid_labor if x})
            labor_base = sorted({v for v in _manpower_valid_names_set()})
            sug_items = []
            for name in uniq[:20]:
                close = difflib.get_close_matches(name.lower(), labor_base, n=1, cutoff=0.72)
                if close:
                    sug_items.append(f"<li><code>{name}</code> → did you mean <b>{close[0]}</b>?</li>")
                else:
                    sug_items.append(f"<li><code>{name}</code></li>")
            suggestion_html = "<ul>" + "".join(sug_items) + "</ul>"
            blocks.append("<li><b>Direct Labor not found in ManpowerBase</b>:" f"{suggestion_html}</li>")

        hint = (
            "<p class='mb-1'>How to fix:</p>"
            "<ul>"
            "<li>Make sure every <b>JobCard</b> exists in the JobCard base.</li>"
            "<li>Use only <b>Direct Labor</b> registered in <b>ManpowerBase</b>.</li>"
            "<li>Download the <i>CSV Template</i> from the top bar to guarantee correct headers and separator.</li>"
            "</ul>"
        )
        html = "<p>Import aborted due to validation errors:</p>"
        if blocks:
            html += f"<ul>{''.join(blocks)}</ul>"
        html += hint
        messages.error(request, mark_safe(html))
        return redirect("allocated_manpower_list")

    # Ordem determinística antes de salvar
    valid_rows.sort(
        key=lambda r: (
            r["jobcard_number"],
            int(r.get("task_order") or 0),
            (r.get("direct_labor") or "").lower()
        )
    )

    affected_pairs: Set[Tuple[str, int]] = set()
    created_na = 0
    pct_updates = 0
    agg_updates = 0
    pruned_orphans = 0

    try:
        with transaction.atomic():
            if mode == "replace_jobcard":
                # 1) Apaga manpowers dos jobcards existentes
                existing_jobcards = {r["jobcard_number"] for r in valid_rows}
                if existing_jobcards:
                    AllocatedManpower.objects.filter(jobcard_number__in=existing_jobcards).delete()

                    # 2) Zera horas de tasks (as remanescentes)
                    AllocatedTask.objects.filter(jobcard_number__in=existing_jobcards).update(total_hours=0)

                    # 3) 🔥 Remove tasks órfãs (que não existem mais no TaskBase)
                    pruned_orphans = _prune_orphan_allocated_tasks(existing_jobcards, job_map)

            # MERGE/UPSERT por (jobcard_number, task_order, direct_labor)
            for r in valid_rows:
                key = dict(
                    jobcard_number=r["jobcard_number"],
                    task_order=r["task_order"],
                    direct_labor=(r["direct_labor"] or "").strip(),
                )
                defaults = dict(
                    discipline=r["discipline"],      # sempre da JobCard
                    working_code=r["working_code"],  # sempre da JobCard
                    qty=r["qty"],
                    hours=r["hours"],
                )
                obj, created = AllocatedManpower.objects.get_or_create(**key, defaults=defaults)
                if not created:
                    obj.discipline = defaults["discipline"]
                    obj.working_code = defaults["working_code"]
                    obj.qty = defaults["qty"]
                    obj.hours = defaults["hours"]
                    obj.save(update_fields=["discipline", "working_code", "qty", "hours"])

                affected_pairs.add((r["jobcard_number"], r["task_order"]))

            # Recalcula totals/max/description, marca aplicáveis e não concluídas
            updated_tasks = _recalc_task_totals(affected_pairs)

            # Em replace_jobcard, completa/ajusta faltantes como N/A
            if mode == "replace_jobcard":
                created_na = _ensure_missing_tasks_as_na(valid_rows, job_map)

            # Recalcula percentuais por participação no total do JobCard
            affected_jobcards = {r["jobcard_number"] for r in valid_rows}
            pct_updates = _recalc_jobcard_percentages(affected_jobcards)

            # Sincroniza JobCard com AllocatedTask (Σ totals / Σ max)
            agg_updates = _sync_jobcard_from_allocatedtask(affected_jobcards)

        # ===== Mensagens =====
        base_msg = (
            f"<p><b>Import completed</b>.</p>"
            f"<ul>"
            f"<li>{len(valid_rows)} valid row(s)</li>"
            f"<li>{len(affected_pairs)} affected task(s)</li>"
            f"<li>{updated_tasks} task(s) updated (total/max/description)</li>"
            f"<li>{pct_updates} task(s) with % recalculated</li>"
            f"<li><b>{agg_updates}</b> JobCard(s) synced from AllocatedTask</li>"
        )
        if mode == "replace_jobcard":
            base_msg += f"<li>{pruned_orphans} orphan task(s) removed (no longer in TaskBase)</li>"
            base_msg += f"<li>{created_na} missing task(s) were set as <b>N/A</b></li>"
        base_msg += "</ul>"
        messages.success(request, mark_safe(base_msg))

    except Exception as e:
        messages.error(request, f"An unexpected error occurred during import: {e}")

    return redirect("allocated_manpower_list")


@login_required(login_url="login")
@permission_required("jobcards.change_jobcard", raise_exception=True)
def allocated_manpower_template(request):
    """
    Gera um CSV modelo com separador ';' e BOM UTF-8, compatível com Excel pt-BR.
    Colunas: jobcard_number, discipline, working_code, direct_labor, task_order, qty, hours
    """
    headers = [
        "jobcard_number",
        "discipline",
        "working_code",
        "direct_labor",
        "task_order",
        "qty",
        "hours",
    ]
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="allocated_manpower_template.csv"'
    response.write("\ufeff")   # BOM
    response.write("sep=;\n")  # dica de separador para Excel

    writer = csv.writer(response, delimiter=";", lineterminator="\n", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    # writer.writerow(["JC-0001","PIPING","W1234","PIPEFITTER",1,"2,00","16,00"])  # exemplo opcional
    return response


@login_required(login_url="login")
@permission_required("jobcards.change_jobcard", raise_exception=True)
def export_allocated_manpower_csv(request):
    """
    Exporta CSV respeitando filtros atuais; delim=comma|semicolon.
    &excel_hint=1 para incluir linha 'sep=;'.
    """
    delim = (request.GET.get("delim") or "comma").lower()
    delimiter = ";" if delim in (";", "semicolon", "scsv", "semi") else ","
    excel_hint = (request.GET.get("excel_hint") or "").lower() in ("1", "true", "yes", "y")

    f_job = (request.GET.get("filter_job") or "").strip()
    f_disc = (request.GET.get("filter_disc") or "").strip()
    f_lab = (request.GET.get("filter_labor") or "").strip()
    q = (request.GET.get("q") or "").strip()

    qs = AllocatedManpower.objects.all()
    if f_job:
        qs = qs.filter(jobcard_number__icontains=f_job)
    if f_disc:
        qs = qs.filter(discipline__icontains=f_disc)
    if f_lab:
        qs = qs.filter(direct_labor__icontains=f_lab)
    if q:
        qs = qs.filter(
            Q(jobcard_number__icontains=q)
            | Q(discipline__icontains=q)
            | Q(working_code__icontains=q)
            | Q(direct_labor__icontains=q)
        )

    # ✅ exporta em ordem
    qs = qs.order_by('jobcard_number', 'task_order', 'direct_labor')

    filename = f'allocated_manpower_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # BOM + dica de separador (opcional)
    response.write("\ufeff")
    if excel_hint and delimiter == ";":
        response.write("sep=;\n")

    writer = csv.writer(response, delimiter=delimiter, lineterminator="\n")
    writer.writerow(
        ["jobcard_number", "discipline", "working_code", "direct_labor", "task_order", "qty", "hours"]
    )
    for r in qs.only(
        "jobcard_number", "discipline", "working_code", "direct_labor", "task_order", "qty", "hours"
    ):
        writer.writerow(
            [
                r.jobcard_number,
                r.discipline,
                r.working_code,
                r.direct_labor,
                r.task_order,
                f"{r.qty:.2f}" if r.qty is not None else "",
                f"{r.hours:.2f}" if r.hours is not None else "",
            ]
        )
    return response


@login_required(login_url="login")
@permission_required("jobcards.change_jobcard", raise_exception=True)
def export_allocated_manpower_xlsx(request):
    """
    Exporta XLSX respeitando os filtros atuais.
    Colunas: jobcard_number, discipline, working_code, direct_labor, task_order, qty, hours
    - Ordenado por (jobcard_number, task_order, direct_labor)
    - Cabeçalho estilizado, AutoFilter, Freeze Pane, AutoFit, formatação numérica 2 casas
    """
    if not _HAS_OPENPYXL:
        return HttpResponse(
            "To export XLSX, install openpyxl (pip install openpyxl).",
            status=500,
            content_type="text/plain; charset=utf-8",
        )

    # === Filtros iguais aos do CSV ===
    f_job = (request.GET.get("filter_job") or "").strip()
    f_disc = (request.GET.get("filter_disc") or "").strip()
    f_lab  = (request.GET.get("filter_labor") or "").strip()
    q      = (request.GET.get("q") or "").strip()

    qs = AllocatedManpower.objects.all()
    if f_job:
        qs = qs.filter(jobcard_number__icontains=f_job)
    if f_disc:
        qs = qs.filter(discipline__icontains=f_disc)
    if f_lab:
        qs = qs.filter(direct_labor__icontains=f_lab)
    if q:
        qs = qs.filter(
            Q(jobcard_number__icontains=q)
            | Q(discipline__icontains=q)
            | Q(working_code__icontains=q)
            | Q(direct_labor__icontains=q)
        )

    qs = qs.order_by("jobcard_number", "task_order", "direct_labor").only(
        "jobcard_number", "discipline", "working_code",
        "direct_labor", "task_order", "qty", "hours"
    )

    # === Monta XLSX em memória ===
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AllocatedManpower"

    headers = ["jobcard_number","discipline","working_code","direct_labor","task_order","qty","hours"]
    ws.append(headers)

    # Linhas
    for r in qs:
        ws.append([
            r.jobcard_number or "",
            r.discipline or "",
            r.working_code or "",
            r.direct_labor or "",
            r.task_order if r.task_order is not None else "",
            float(r.qty or 0.0),
            float(r.hours or 0.0),
        ])

    # Estilo cabeçalho
    header_fill  = PatternFill("solid", fgColor="F3F5F8")
    header_font  = Font(bold=True)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        cell.border = border

    # AutoFilter e congelar linha 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    ws.freeze_panes = "A2"

    # Formatação numérica (2 casas) para qty (F) e hours (G)
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = "0.00"

    # Auto-ajuste de largura (simples)
    for col_idx in range(1, len(headers) + 1):
        col = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col]:
            v = cell.value
            l = len(str(v)) if v is not None else 0
            if l > max_len:
                max_len = l
        ws.column_dimensions[col].width = min(60, max(10, max_len + 2))

    # Retorna resposta
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f'allocated_manpower_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
